import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt
from io import BytesIO
from Bio import SeqIO
import json
import plotly.express as px
import plotly.graph_objects as go
import seaborn as sns
import numpy as np
from matplotlib.patches import Rectangle
import math
import zipfile
import random
from Bio.Seq import Seq
from Bio.SeqRecord import SeqRecord
import io
import openpyxl

# Custom CSS for professional styling
def load_css():
    st.markdown("""
    <style>
    /* Main background with gradient colors only */
    .main {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 25%, #1a1a1a 50%, #2d2d2d 75%, #1a1a1a 100%) !important;
        color: #ffffff !important;
        position: relative;
    }
    
    .stApp {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 25%, #1a1a1a 50%, #2d2d2d 75%, #1a1a1a 100%) !important;
        color: #ffffff !important;
    }
    
    /* Force background on all main containers */
    .block-container {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 25%, #1a1a1a 50%, #2d2d2d 75%, #1a1a1a 100%) !important;
        color: #ffffff !important;
    }
    
    /* Override any Streamlit default backgrounds */
    div[data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 25%, #1a1a1a 50%, #2d2d2d 75%, #1a1a1a 100%) !important;
    }
    
    div[data-testid="stAppViewBlockContainer"] {
        background: linear-gradient(135deg, #1a1a1a 0%, #2d2d2d 25%, #1a1a1a 50%, #2d2d2d 75%, #1a1a1a 100%) !important;
    }
    
    /* Ensure all text is white */
    .stMarkdown, .stText, .stDataFrame, .stSelectbox, .stTextInput, .stTextArea, .stNumberInput {
        color: #ffffff !important;
    }
    
    /* Override any dark text */
    p, h1, h2, h3, h4, h5, h6, span, div, label {
        color: #ffffff !important;
    }
    
    /* Sidebar styling - Overlay mode */
    .css-1d391kg {
        background: linear-gradient(180deg, #1a1a1a 0%, #2d2d2d 50%, #1a1a1a 100%) !important;
        border-right: 2px solid #3498db;
        position: fixed !important;
        top: 0;
        left: 0;
        height: 100vh;
        z-index: 1000;
        width: 21rem;
        overflow-y: auto;
        box-shadow: 2px 0 10px rgba(0,0,0,0.3);
        transition: transform 0.3s ease;
        color: #ffffff !important;
    }
    
    /* Main content area - adjust for overlay sidebar */
    .main .block-container {
        margin-left: 0 !important;
        padding-left: 1rem !important;
        max-width: none !important;
    }
    
    /* Sidebar toggle button styling */
    .css-1rs6os {
        position: fixed !important;
        top: 1rem;
        left: 1rem;
        z-index: 1001;
        background: rgba(52, 152, 219, 0.9) !important;
        border-radius: 50% !important;
        width: 3rem !important;
        height: 3rem !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        box-shadow: 0 2px 10px rgba(0,0,0,0.3) !important;
    }
    
    /* Header styling */
    .main-header {
        background: linear-gradient(135deg, #404040 0%, #505050 50%, #404040 100%) !important;
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        box-shadow: 0 8px 32px rgba(0,0,0,0.4);
        border: 1px solid #3498db;
        color: #ffffff !important;
    }
    
    .main-header h1 {
        color: #ffffff !important;
        text-align: center;
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
    }
    
    /* Section styling */
    .section-container {
        background: linear-gradient(135deg, #404040 0%, #505050 50%, #404040 100%) !important;
        padding: 2rem;
        border-radius: 15px;
        margin: 1rem 0;
        border: 1px solid rgba(52, 152, 219, 0.3);
        backdrop-filter: blur(10px);
        box-shadow: 0 8px 32px rgba(0,0,0,0.3);
        color: #ffffff !important;
    }
    
    .section-header {
        color: #ffffff !important;
        font-size: 1.5rem;
        font-weight: 600;
        margin-bottom: 1rem;
        text-align: center;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    /* Button styling */
    .stButton > button {
        background: linear-gradient(45deg, #3498db, #2980b9) !important;
        color: #ffffff !important;
        border: none;
        border-radius: 25px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
    }
    
    .stButton > button:hover {
        background: linear-gradient(45deg, #2980b9, #1f5f8b) !important;
        color: #ffffff !important;
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
    }
    
    /* Input styling */
    .stTextInput > div > div > input {
        background: #505050 !important;
        border: 1px solid rgba(52, 152, 219, 0.5);
        border-radius: 10px;
        color: #ffffff !important;
        padding: 0.75rem;
    }
    
    .stTextArea > div > div > textarea {
        background: #505050 !important;
        border: 1px solid rgba(52, 152, 219, 0.5);
        border-radius: 10px;
        color: #ffffff !important;
        padding: 0.75rem;
    }
    
    .stSelectbox > div > div > select {
        background: #505050 !important;
        border: 1px solid rgba(52, 152, 219, 0.5);
        border-radius: 10px;
        color: #ffffff !important;
        padding: 0.75rem;
    }
    
    /* File uploader styling */
    .stFileUploader > div {
        background: #404040 !important;
        border: 2px dashed rgba(52, 152, 219, 0.5);
        border-radius: 15px;
        padding: 2rem;
        text-align: center;
        color: #ffffff !important;
    }
    
    /* Radio button styling */
    .stRadio > div {
        background: #404040 !important;
        border-radius: 10px;
        padding: 1rem;
        border: 1px solid rgba(52, 152, 219, 0.3);
        color: #ffffff !important;
    }
    
    /* Number input styling */
    .stNumberInput > div > div > input {
        background: #505050 !important;
        border: 1px solid rgba(52, 152, 219, 0.5);
        border-radius: 10px;
        color: #ffffff !important;
        padding: 0.75rem;
    }
    
    /* Expander styling */
    .streamlit-expanderHeader {
        background: #505050 !important;
        color: #ffffff !important;
        border-radius: 10px;
        font-weight: 600;
        border: 1px solid rgba(52, 152, 219, 0.3);
    }
    
    /* Success/Info/Warning styling */
    .stSuccess {
        background: rgba(46, 204, 113, 0.2);
        border: 1px solid #2ecc71;
        border-radius: 10px;
        padding: 1rem;
        color: #2ecc71;
    }
    
    .stInfo {
        background: rgba(52, 152, 219, 0.2);
        border: 1px solid #3498db;
        border-radius: 10px;
        padding: 1rem;
        color: #3498db;
    }
    
    .stWarning {
        background: rgba(241, 196, 15, 0.2);
        border: 1px solid #f1c40f;
        border-radius: 10px;
        padding: 1rem;
        color: #f1c40f;
    }
    
    /* Dataframe styling */
    .dataframe {
        background: #404040 !important;
        border-radius: 10px;
        overflow: hidden;
        border: 1px solid rgba(52, 152, 219, 0.3);
        color: #ffffff !important;
    }
    
    /* Download button styling */
    .stDownloadButton > button {
        background: linear-gradient(45deg, #27ae60, #229954) !important;
        color: #ffffff !important;
        border: none;
        border-radius: 25px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
    }
    
    .stDownloadButton > button:hover {
        background: linear-gradient(45deg, #229954, #1e8449) !important;
        color: #ffffff !important;
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
    }
    
    /* Progress bar styling */
    .stProgress > div > div > div {
        background: linear-gradient(90deg, #3498db, #2980b9);
        border-radius: 10px;
    }
    
    /* Sidebar expander styling */
    .css-1d391kg .streamlit-expanderHeader {
        background: #505050 !important;
        color: #ffffff !important;
        border-radius: 8px;
        font-weight: 600;
        border: 1px solid rgba(52, 152, 219, 0.3);
    }
    
    /* Footer styling */
    .footer {
        text-align: center;
        color: #ffffff !important;
        font-size: 0.9rem;
        margin-top: 3rem;
        padding: 1rem;
        border-top: 1px solid rgba(52, 152, 219, 0.3);
    }
    
    /* Tool card styling */
    .tool-card {
        background: #404040 !important;
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1rem 0;
        border: 1px solid rgba(52, 152, 219, 0.3);
        box-shadow: 0 4px 15px rgba(0,0,0,0.3);
        color: #ffffff !important;
    }
    
    /* Fragmenter specific styling */
    .fragmenter-section {
        background: #404040 !important;
        border: 2px solid rgba(52, 152, 219, 0.5);
        border-radius: 20px;
        padding: 2rem;
        margin: 1rem 0;
        color: #ffffff !important;
    }
    
    /* Responsive design */
    @media (max-width: 768px) {
        .main-header h1 {
            font-size: 2rem;
        }
        
        .section-container {
            padding: 1rem;
        }
        
        .css-1d391kg {
            width: 100vw;
        }
    }
    </style>
    """, unsafe_allow_html=True)

# --- Instructions ---
def show_instructions():
    st.markdown("""
    ### 📋 Instructions
    
    1. **Choose Input Method**:
       - **Upload CSV/Excel**: Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns
       - **Upload FASTA**: Upload a FASTA file containing DNA sequences
       - **Upload GenBank**: Upload a GenBank file containing DNA sequences
       - **Manual Entry**: Enter sequences manually with gene names
    
    2. **Input Format**:
       - Only A, T, G, C nucleotides are accepted (case insensitive)
       - Invalid characters will be automatically removed
       - Each sequence must have a unique gene name
    
    3. **Analysis**:
       - Click "Calculate GC Content" to process sequences
       - View results in the interactive table
       - Explore visualizations of GC content and nucleotide composition
    
    4. **Export Options**:
       - Export results in multiple formats (Excel, CSV, JSON)
       - Use "Export All Results" for a complete dataset export
       - Customize output filename before downloading
    """)

# --- Sequence Sanitization and Validation ---
def sanitize_sequence(seq):
    # Convert to string if not already
    if not isinstance(seq, str):
        seq = str(seq)
    return ''.join(filter(lambda x: x.upper() in ['A', 'T', 'G', 'C'], seq.upper()))

def is_valid_sequence(seq):
    # Convert to string if not already
    if not isinstance(seq, str):
        seq = str(seq)
    return all(base in ['A', 'T', 'G', 'C'] for base in seq.upper())

# --- Z-DNA Prediction Function ---
def calculate_z_dna_propensity(sequence):
    """
    Calculate Z-DNA propensity based on alternating purine-pyrimidine patterns
    and GC-rich regions that are prone to Z-DNA formation.
    """
    if not sequence or len(sequence) < 6:
        return 0.0
    
    sequence = sequence.upper()
    z_score = 0.0
    total_positions = 0
    
    # Check for alternating purine-pyrimidine patterns (Z-DNA prone)
    for i in range(len(sequence) - 5):
        window = sequence[i:i+6]
        if len(window) == 6:
            # Check for alternating pattern
            alternating = True
            for j in range(1, 6):
                current = window[j] in 'AG'  # Purine
                prev = window[j-1] in 'AG'   # Purine
                if current == prev:  # Not alternating
                    alternating = False
                    break
            
            if alternating:
                # Calculate GC content in this window
                gc_count = window.count('G') + window.count('C')
                gc_content = gc_count / 6.0
                
                # Z-DNA is more likely in GC-rich alternating regions
                if gc_content >= 0.5:  # 50% or more GC
                    z_score += gc_content * 2.0  # Higher weight for GC-rich
                else:
                    z_score += gc_content * 1.0  # Lower weight for AT-rich
                
                total_positions += 1
    
    # Normalize by sequence length and total positions checked
    if total_positions > 0:
        return (z_score / total_positions) * 100  # Convert to percentage
    else:
        return 0.0

# --- Enhanced Nucleotide Analysis with Z-DNA ---
def analyze_sequence(name, seq):
    # Convert sequence to string if it's not already
    if not isinstance(seq, str):
        seq = str(seq)
    
    seq = sanitize_sequence(seq)
    length = len(seq)
    if length == 0:
        return {"Gene Name": name, "Error": "Invalid sequence or empty after sanitization"}

    a_count = seq.count('A')
    t_count = seq.count('T')
    g_count = seq.count('G')
    c_count = seq.count('C')
    gc_count = g_count + c_count
    at_count = a_count + t_count

    # Calculate Z-DNA propensity
    z_dna_propensity = calculate_z_dna_propensity(seq)

    return {
        "Gene Name": name,
        "Sequence": seq,
        "Length": length,
        "A Count": a_count,
        "T Count": t_count,
        "G Count": g_count,
        "C Count": c_count,
        "A %": round((a_count / length) * 100, 2),
        "T %": round((t_count / length) * 100, 2),
        "G %": round((g_count / length) * 100, 2),
        "C %": round((c_count / length) * 100, 2),
        "GC %": round((gc_count / length) * 100, 2),
        "AT %": round((at_count / length) * 100, 2),
        "Z-DNA Propensity %": round(z_dna_propensity, 2),
    }

def process_fasta(file, max_sequences=1000):
    sequences = []
    try:
        # Ensure the file is read from the beginning
        file.seek(0)
        # Read binary and decode to text for FASTA
        text = file.read().decode("utf-8")
        text_io = io.StringIO(text)
        records = SeqIO.parse(text_io, "fasta")
        count = 0
        for record in records:
            if count >= max_sequences:
                break
            # Convert Seq object to string
            sequences.append((str(record.id), str(record.seq)))
            count += 1
        if count == 0:
            st.warning("No sequences found in the uploaded FASTA file after parsing.")
        elif count >= max_sequences:
             st.info(f"Processed the first {max_sequences} sequences from the FASTA file.")
    except Exception as e:
        st.error(f"Error parsing FASTA file: {e}")
        # Return empty list on error
        return []

    return sequences

def process_genbank(file, max_sequences=1000):
    sequences = []
    try:
        # Ensure the file is read from the beginning
        file.seek(0)
        # Read binary and decode to text for GenBank
        text = file.read().decode("utf-8")
        text_io = io.StringIO(text)
        records = SeqIO.parse(text_io, "genbank")
        count = 0
        for record in records:
             if count >= max_sequences:
                 break
             # Ensure molecule_type is set for GenBank export compatibility later
             # For uploaded GenBank, try to get molecule_type from annotations, default to DNA
             mol_type = record.annotations.get("molecule_type", "DNA")
             sequences.append((str(record.id), str(record.seq)))
             count += 1
        if count == 0:
             st.warning("No sequences found in the uploaded GenBank file after parsing.")
        elif count >= max_sequences:
             st.info(f"Processed the first {max_sequences} sequences from the GenBank file.")

    except Exception as e:
        st.error(f"Error parsing GenBank file: {e}")
        # Return empty list on error
        return []

    return sequences

def display_visuals(df):
    st.subheader("📊 GC Content Distribution")
    fig, ax = plt.subplots()
    df.plot.bar(x='Gene Name', y='GC %', ax=ax, color='purple', legend=False)
    plt.ylabel('GC %')
    plt.xticks(rotation=45, ha='right')
    st.pyplot(fig)

    st.subheader("🧬 Nucleotide Composition Per Sequence")
    for _, row in df.iterrows():
        fig, ax = plt.subplots()
        ax.pie(
            [row['A %'], row['T %'], row['G %'], row['C %']],
            labels=['A %', 'T %', 'G %', 'C %'],
            autopct='%1.1f%%',
            startangle=90
        )
        ax.set_title(f"{row['Gene Name']} - Base % Composition")
        st.pyplot(fig)

def export_data(df, format_type):
    signature = "Made by Shubh Rakesh Nahar / Troy University"
    if format_type == "Excel":
        towrite = BytesIO()
        with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['GC Content Analysis']
            
            # Add signature as a comment in cell A1
            worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
            # Auto-adjust columns' width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Add some basic formatting
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = cell.font.copy(bold=True)
            
            # Add signature in the last row
            last_row = len(df) + 2
            worksheet.cell(row=last_row, column=1, value=signature)
        
        towrite.seek(0)
        return towrite, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    elif format_type == "CSV":
        towrite = BytesIO()
        # Add signature as a comment in the first line
        towrite.write(f"# {signature}\n".encode())
        df.to_csv(towrite, index=False)
        towrite.seek(0)
        return towrite, "text/csv", "csv"
    else:  # JSON
        towrite = BytesIO()
        # Add signature to the JSON metadata
        json_data = {
            "metadata": {
                "signature": signature,
                "timestamp": pd.Timestamp.now().isoformat()
            },
            "data": df.to_dict(orient='records')
        }
        json_str = json.dumps(json_data, indent=2)
        towrite.write(json_str.encode())
        towrite.seek(0)
        return towrite, "application/json", "json"

def create_gc_heatmap(df):
    st.subheader("🌡️ GC Content Heatmap")
    
    # Create a matrix of GC content for each position
    sequences = df['Sequence'].tolist()
    max_len = max(len(seq) for seq in sequences)
    
    # Initialize matrix
    gc_matrix = np.zeros((len(sequences), max_len))
    
    # Fill matrix with GC content for each position
    for i, seq in enumerate(sequences):
        for j in range(len(seq)):
            if j < len(seq):
                window = seq[max(0, j-10):min(len(seq), j+11)]
                gc_count = window.count('G') + window.count('C')
                gc_matrix[i, j] = (gc_count / len(window)) * 100
    
    # Create heatmap using plotly
    fig = go.Figure(data=go.Heatmap(
        z=gc_matrix,
        x=list(range(max_len)),
        y=df['Gene Name'].tolist(),
        colorscale='Viridis',
        colorbar=dict(title='GC %')
    ))
    
    fig.update_layout(
        title='GC Content Distribution Across Sequences',
        xaxis_title='Position',
        yaxis_title='Gene Name',
        height=400 + (len(sequences) * 20)  # Adjust height based on number of sequences
    )
    
    st.plotly_chart(fig, use_container_width=True)

def calculate_information_content(freq):
    """Calculate information content in bits."""
    if freq == 0:
        return 0
    return freq * math.log2(freq * 4)  # 4 for number of nucleotides

def create_sequence_logo(sequences):
    st.subheader("🎨 Sequence Logo")
    
    # Calculate position frequency matrix
    max_len = max(len(seq) for seq in sequences)
    pfm = np.zeros((4, max_len))  # 4 nucleotides
    
    for seq in sequences:
        for i, base in enumerate(seq):
            if i < max_len:
                if base == 'A':
                    pfm[0, i] += 1
                elif base == 'T':
                    pfm[1, i] += 1
                elif base == 'G':
                    pfm[2, i] += 1
                elif base == 'C':
                    pfm[3, i] += 1
    
    # Normalize
    pfm = pfm / len(sequences)
    
    # Calculate information content
    ic = np.zeros(max_len)
    for i in range(max_len):
        ic[i] = sum(calculate_information_content(freq) for freq in pfm[:, i])
    
    # Create the plot
    fig, ax = plt.subplots(figsize=(12, 4))
    
    # Colors for nucleotides
    colors = {'A': '#2ecc71', 'T': '#e74c3c', 'G': '#f1c40f', 'C': '#3498db'}
    bases = ['A', 'T', 'G', 'C']
    
    # Plot each position
    for i in range(max_len):
        # Sort frequencies for this position
        freqs = pfm[:, i]
        sorted_indices = np.argsort(freqs)
        
        # Plot each base
        y_bottom = 0
        for idx in sorted_indices:
            if freqs[idx] > 0:
                height = freqs[idx] * ic[i]
                rect = Rectangle((i, y_bottom), 1, height,
                               facecolor=colors[bases[idx]],
                               edgecolor='black',
                               linewidth=0.5)
                ax.add_patch(rect)
                y_bottom += height
    
    # Customize the plot
    ax.set_xlim(0, max_len)
    ax.set_ylim(0, max(ic) * 1.1)
    ax.set_xlabel('Position')
    ax.set_ylabel('Bits')
    ax.set_title('Sequence Logo')
    
    # Add legend
    legend_elements = [Rectangle((0, 0), 1, 1, facecolor=color, edgecolor='black')
                      for color in colors.values()]
    ax.legend(legend_elements, bases, loc='upper right')
    
    # Remove spines
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    
    # Show the plot
    st.pyplot(fig)
    plt.close(fig)

def create_interactive_plots(df):
    st.subheader("📊 Interactive Plots")
    
    # GC Content Distribution
    fig_gc = px.box(df, y='GC %', title='GC Content Distribution',
                    hover_data=['Gene Name', 'Length', 'GC %'])
    fig_gc.update_traces(marker_color='purple')
    st.plotly_chart(fig_gc, use_container_width=True)
    
    # Nucleotide Composition
    fig_comp = px.bar(df, 
                     x='Gene Name',
                     y=['A %', 'T %', 'G %', 'C %'],
                     title='Nucleotide Composition by Gene',
                     barmode='group',
                     hover_data=['Length', 'GC %'])
    st.plotly_chart(fig_comp, use_container_width=True)
    
    # Length vs GC Content
    fig_scatter = px.scatter(df,
                           x='Length',
                           y='GC %',
                           color='GC %',
                           hover_data=['Gene Name', 'Length', 'GC %'],
                           title='Sequence Length vs GC Content')
    st.plotly_chart(fig_scatter, use_container_width=True)

def export_all_to_zip(df, output_filename):
    signature = "Made by Shubh Rakesh Nahar / Troy University"
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w") as zip_file:
        # Excel with formatting
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
            # Get the workbook and the worksheet
            workbook = writer.book
            worksheet = writer.sheets['GC Content Analysis']
            
            # Add signature as a comment in cell A1
            worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
            # Auto-adjust columns' width
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Add some basic formatting
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.font = cell.font.copy(bold=True)
            
            # Add signature in the last row
            last_row = len(df) + 2
            worksheet.cell(row=last_row, column=1, value=signature)
        
        excel_buffer.seek(0)
        zip_file.writestr(f"{output_filename}.xlsx", excel_buffer.read())
        
        # CSV
        csv_buffer = BytesIO()
        # Add signature as a comment in the first line
        csv_buffer.write(f"# {signature}\n".encode())
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        zip_file.writestr(f"{output_filename}.csv", csv_buffer.read())
        
        # JSON
        json_data = {
            "metadata": {
                "signature": signature,
                "timestamp": pd.Timestamp.now().isoformat()
            },
            "data": df.to_dict(orient='records')
        }
        json_str = json.dumps(json_data, indent=2)
        json_buffer = BytesIO()
        json_buffer.write(json_str.encode())
        json_buffer.seek(0)
        zip_file.writestr(f"{output_filename}.json", json_buffer.read())
    
    zip_buffer.seek(0)
    return zip_buffer

def export_to_fasta(sequences, result_df):
    """Export sequences to FASTA format with analysis results in description"""
    signature = "Made by Shubh Rakesh Nahar / Troy University"
    output = io.StringIO()
    # Add signature as a comment at the top
    output.write(f"# {signature}\n")
    for idx, row in result_df.iterrows():
        # Create a description with analysis results
        description = f"GC Content: {row['GC %']:.2f}%, Length: {row['Length']}, "
        description += f"A: {row['A Count']}, T: {row['T Count']}, G: {row['G Count']}, C: {row['C Count']}"
        
        # Write sequence with description
        output.write(f">{row['Gene Name']} {description}\n")
        output.write(f"{row['Sequence']}\n")
    
    # Add signature at the end
    output.write(f"\n# {signature}\n")
    return output.getvalue()

def export_to_genbank(sequences, result_df):
    """Export sequences to GenBank format with analysis results in features"""
    signature = "Made by Shubh Rakesh Nahar / Troy University"
    output = io.StringIO()
    for idx, row in result_df.iterrows():
        # Write GenBank header
        output.write(f"LOCUS       {row['Gene Name']}              {row['Length']} bp    DNA     linear\n")
        output.write(f"DEFINITION  {row['Gene Name']}\n")
        output.write(f"ACCESSION   {row['Gene Name']}\n")
        output.write(f"VERSION     {row['Gene Name']}\n")
        output.write(f"SOURCE      .\n")
        output.write(f"  ORGANISM  .\n")
        output.write(f"            .\n")
        
        # Add signature in the features section
        output.write("FEATURES             Location/Qualifiers\n")
        output.write("     source          1..{}\n".format(row['Length']))
        output.write('                     /organism="."\n')
        output.write('                     /mol_type="genomic DNA"\n')
        output.write('                     /note="{}"\n'.format(signature))
        output.write('     misc_feature    1..{}\n'.format(row['Length']))
        output.write('                     /note="GC Content: {:.2f}%"\n'.format(row['GC %']))
        output.write('                     /note="A Count: {}"\n'.format(row['A Count']))
        output.write('                     /note="T Count: {}"\n'.format(row['T Count']))
        output.write('                     /note="G Count: {}"\n'.format(row['G Count']))
        output.write('                     /note="C Count: {}"\n'.format(row['C Count']))
        
        # Write sequence
        output.write("ORIGIN\n")
        sequence = row['Sequence']
        for i in range(0, len(sequence), 60):
            chunk = sequence[i:i+60]
            output.write(f"     {i+1:9d} {chunk}\n")
        output.write("//\n")
    
    return output.getvalue()

def fragment_sequences(sequences, fragment_length=200):
    """
    Splits each sequence into fragments of the given length (default 200 bases) and calculates comprehensive analysis.
    Returns a list of (Fragment Name, Fragment Sequence, Parent Name, Fragment Index, Start Position, Analysis Data).
    """
    fragments = []
    for name, seq in sequences:
        seq = sanitize_sequence(seq)
        if not seq or len(seq) < fragment_length:
            # If sequence is too short, keep as is
            analysis = analyze_sequence(name, seq)
            fragments.append((name, seq, name, 1, 1, analysis))
            continue
            
        # Split into fragments
        num_fragments = (len(seq) + fragment_length - 1) // fragment_length
        
        for i in range(num_fragments):
            start = i * fragment_length
            end = min(start + fragment_length, len(seq))
            frag_seq = seq[start:end]
            
            # Create fragment name with position range
            frag_name = f"{name}_F{i+1}[{start+1}-{end}]"
            
            # Calculate comprehensive analysis for the fragment
            analysis = analyze_sequence(frag_name, frag_seq)
            
            fragments.append((frag_name, frag_seq, name, i+1, start+1, analysis))
    return fragments

def analyze_sequence_with_fragments(name, sequence, fragment_length=200):
    """
    Analyze a sequence and automatically create fragments with comprehensive analysis.
    Returns original analysis and list of fragment analyses.
    """
    # Analyze original sequence
    original_analysis = analyze_sequence(name, sequence)
    
    # Create fragments
    fragments = fragment_sequences([(name, sequence)], fragment_length)
    
    # Extract fragment analyses
    fragment_analyses = []
    for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
        fragment_analyses.append({
            "Fragment_Name": frag_name,
            "Parent_Gene": parent_name,
            "Fragment_Number": frag_num,
            "Start_Position": start_pos,
            "End_Position": start_pos + len(frag_seq) - 1,
            "Fragment_Length": len(frag_seq),
            "Fragment_Sequence": frag_seq,
            **analysis  # Include all analysis data
        })
    
    return original_analysis, fragment_analyses

def export_fragments_to_excel(fragments):
    # Create a comprehensive DataFrame with all analysis data including Z-DNA
    data = []
    for frag_name, frag_seq, parent, idx, start_pos, analysis in fragments:
        row = {
            "Gene Name": frag_name,
            "Sequence": frag_seq,
            "Parent Name": parent,
            "Fragment Index": idx,
            "Start Position": start_pos,
            "End Position": start_pos + len(frag_seq) - 1,
            "Fragment Length": len(frag_seq),
            "Length": analysis.get("Length", len(frag_seq)),
            "A Count": analysis.get("A Count", 0),
            "T Count": analysis.get("T Count", 0),
            "G Count": analysis.get("G Count", 0),
            "C Count": analysis.get("C Count", 0),
            "A %": analysis.get("A %", 0),
            "T %": analysis.get("T %", 0),
            "G %": analysis.get("G %", 0),
            "C %": analysis.get("C %", 0),
            "GC %": analysis.get("GC %", 0),
            "AT %": analysis.get("AT %", 0),
            "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
        }
        data.append(row)
    
    df = pd.DataFrame(data)
    towrite = BytesIO()
    with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Fragments with Analysis')
        workbook = writer.book
        worksheet = writer.sheets['Fragments with Analysis']
        
        # Auto-adjust columns' width
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Add formatting to headers
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = cell.font.copy(bold=True)
    
    towrite.seek(0)
    return towrite


def export_fragments_to_fasta(fragments):
    output = io.StringIO()
    for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
        # Create a comprehensive description with all analysis data
        description = f"parent={parent} idx={idx} len={frag_len} "
        description += f"GC_Content={analysis.get('GC %', 0):.2f}% "
        description += f"A={analysis.get('A Count', 0)}({analysis.get('A %', 0):.2f}%) "
        description += f"T={analysis.get('T Count', 0)}({analysis.get('T %', 0):.2f}%) "
        description += f"G={analysis.get('G Count', 0)}({analysis.get('G %', 0):.2f}%) "
        description += f"C={analysis.get('C Count', 0)}({analysis.get('C %', 0):.2f}%)"
        
        output.write(f">{frag_name} {description}\n{frag_seq}\n")
    return output.getvalue()


def export_fragments_to_genbank(fragments):
    output = io.StringIO()
    for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
        output.write(f"LOCUS       {frag_name} {frag_len} bp    DNA     linear\n")
        output.write(f"DEFINITION  Fragment of {parent}, index {idx}\n")
        output.write(f"ACCESSION   {frag_name}\n")
        output.write(f"VERSION     {frag_name}\n")
        output.write(f"SOURCE      .\n")
        output.write(f"  ORGANISM  .\n")
        output.write(f"            .\n")
        output.write("FEATURES             Location/Qualifiers\n")
        output.write(f"     source          1..{frag_len}\n")
        output.write('                     /organism="."\n')
        output.write('                     /mol_type="genomic DNA"\n')
        output.write(f'                     /note="Fragment of {parent}, index {idx}"\n')
        output.write(f'                     /note="GC Content: {analysis.get("GC %", 0):.2f}%"\n')
        output.write(f'                     /note="A Count: {analysis.get("A Count", 0)} ({analysis.get("A %", 0):.2f}%)"\n')
        output.write(f'                     /note="T Count: {analysis.get("T Count", 0)} ({analysis.get("T %", 0):.2f}%)"\n')
        output.write(f'                     /note="G Count: {analysis.get("G Count", 0)} ({analysis.get("G %", 0):.2f}%)"\n')
        output.write(f'                     /note="C Count: {analysis.get("C Count", 0)} ({analysis.get("C %", 0):.2f}%)"\n')
        output.write(f'                     /note="AT Content: {analysis.get("AT %", 0):.2f}%"\n')
        output.write("ORIGIN\n")
        for i in range(0, len(frag_seq), 60):
            chunk = frag_seq[i:i+60]
            output.write(f"     {i+1:9d} {chunk}\n")
        output.write("//\n")
    return output.getvalue()

def main():
    st.set_page_config(page_title="SNGC", page_icon=None, layout="wide")
    
    # Load custom CSS
    load_css()
    
    # Password protection
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        # Login screen
        st.markdown("""
            <div class="main-header">
                <h1>SNGC</h1>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown("""
            <div class="section-container">
                <h2 class="section-header">Authentication Required</h2>
            </div>
        """, unsafe_allow_html=True)
        
        # Center the login form
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("""
                <div style='text-align: center; margin: 2rem 0;'>
                    <h3>Please enter the password to access the application</h3>
                </div>
            """, unsafe_allow_html=True)
            
            password = st.text_input("Password", type="password", key="password_input")
            
            if st.button("Login", key="login_button"):
                if password == "TroyDNA2024":
                    st.session_state.authenticated = True
                    st.success("Authentication successful! Redirecting...")
                    st.rerun()
                else:
                    st.error("Incorrect password. Please try again.")
            
            st.markdown("""
                <div style='text-align: center; margin-top: 2rem; color: #bdc3c7;'>
                    <p>Developed by Shubh Rakesh Nahar | Troy University</p>
                </div>
            """, unsafe_allow_html=True)
        
        return
    
    # Main application (only shown after authentication)
    # Professional header with gradient background
    st.markdown("""
        <div class="main-header">
            <h1>SNGC</h1>
        </div>
    """, unsafe_allow_html=True)

    sequences = []  # Ensure sequences is always defined

    # Sidebar with logo, navigation, and about info
    st.sidebar.image("https://cdn-icons-png.flaticon.com/512/616/616494.png", width=100)
    st.sidebar.title("Navigation")
    
    # Add logout button to sidebar
    if st.sidebar.button("Logout"):
        st.session_state.authenticated = False
        st.rerun()
    
    with st.sidebar.expander("About this Application"):
        st.markdown("""
        <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
        This professional tool calculates GC content and provides comprehensive analysis for DNA sequences. 
        Upload your data or enter sequences manually to explore nucleotide composition, GC content, and more. 
        
        **Developed by Shubh Rakesh Nahar, Troy University.**
        </div>
        """, unsafe_allow_html=True)
    
    # Fun facts about DNA/genes/sequences
    facts = [
        "The human genome contains about 3 billion base pairs.",
        "GC content can affect the stability of DNA.",
        "Some bacteria have extremely high or low GC content.",
        "DNA was first isolated by Friedrich Miescher in 1869.",
        "GC-rich regions are often found near gene promoters.",
        "Genes are segments of DNA that code for proteins.",
        "The longest human gene is over 2.4 million base pairs long!",
        "Mitochondrial DNA is inherited only from your mother.",
        "Some viruses use RNA instead of DNA as their genetic material.",
        "The fruit fly has about 15,000 genes, while humans have about 20,000-25,000.",
        "DNA stands for Deoxyribonucleic Acid.",
        "The double helix structure of DNA was discovered in 1953.",
        "Some plants have much more DNA than humans!",
        "The GC content of a genome can be used to identify species.",
        "DNA can be extracted from almost any living thing, even ancient fossils!"
    ]
    if 'fun_fact_idx' not in st.session_state:
        st.session_state['fun_fact_idx'] = random.randint(0, len(facts)-1)
    if st.sidebar.button("Show Another Fun Fact"):
        prev_idx = st.session_state['fun_fact_idx']
        new_idx = prev_idx
        while new_idx == prev_idx:
            new_idx = random.randint(0, len(facts)-1)
        st.session_state['fun_fact_idx'] = new_idx
        st.rerun()
    st.sidebar.success(f"**Fun Fact:** {facts[st.session_state['fun_fact_idx']]}")

    # Glossary of Bioinformatics Terms
    glossary = {
        "GC Content": "The percentage of bases in a DNA or RNA molecule that are either guanine (G) or cytosine (C).",
        "FASTA": "A text-based format for representing nucleotide or peptide sequences.",
        "GenBank": "A rich file format for DNA sequences with annotations, used by NCBI.",
        "ORF": "Open Reading Frame, a sequence of DNA that could potentially encode a protein.",
        "Codon": "A sequence of three nucleotides that together form a unit of genetic code.",
        "Motif": "A short, recurring pattern in DNA that is presumed to have a biological function.",
        "CpG Island": "A region with a high frequency of CG dinucleotides, often found near gene promoters.",
        "SNP": "Single Nucleotide Polymorphism, a variation at a single position in a DNA sequence among individuals.",
        "BLAST": "A tool for comparing an input sequence against a database of sequences.",
        "Reverse Complement": "The sequence formed by reversing a DNA sequence and replacing each base with its complement.",
        "Translation": "The process of converting a nucleotide sequence into a protein sequence.",
        "Phylogenetic Tree": "A branching diagram showing evolutionary relationships among sequences.",
        "Alignment": "The arrangement of two or more sequences to identify regions of similarity."
    }
    with st.sidebar.expander("Glossary of Bioinformatics Terms"):
        search_term = st.text_input("Search glossary", key="glossary_search")
        for term, definition in glossary.items():
            if not search_term or search_term.lower() in term.lower():
                st.markdown(f"**{term}:** {definition}")

    # Tutorial Mode
    with st.sidebar.expander("Tutorial / Step-by-Step Guide"):
        st.markdown("""
        <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
        <h4 style='margin-top:0;'>How to Use the GC Content Calculator and Sequence Fragmenter</h4>
        <p>
        This application allows you to analyze DNA sequences for GC content and nucleotide composition, as well as to split sequences into defined-length fragments with full analysis. Follow these steps to use the tools effectively:
        </p>
        <p>
        <b>1. Choose Your Input Method:</b><br>
        You can upload your DNA sequences using a CSV or Excel file, a FASTA file, a GenBank file, or by entering them manually. For file uploads, ensure your data includes both a gene name and a sequence for each entry.
        </p>
        <p>
        <b>2. Analyze Sequences:</b><br>
        After uploading or entering your sequences, click the "Calculate GC Content & Fragment Analysis" button. The application will automatically split each sequence into 200-base fragments and process your data to display comprehensive analysis including GC content, nucleotide composition statistics, and Z-DNA propensity for both original sequences and each fragment.
        </p>
        <p>
        <b>3. Export Results:</b><br>
        Once the analysis is complete, you can download your results in Excel, FASTA, GenBank, or as a ZIP archive. Use the export section to select your preferred format.
        </p>
        <p>
        <b>4. Use the Sequence Fragmenter Tool:</b><br>
        Open the "Sequence Fragmenter (Split Tool)" section in the sidebar. Choose your input method and upload or enter your sequences. Select the desired fragment length. When you click "Split and Export Fragments," the tool will split each sequence into fragments of the specified length and calculate GC content and nucleotide statistics for each fragment. You can then download the results in Excel, FASTA, or GenBank format, each including the full analysis for every fragment.
        </p>
        <p>
        <b>5. Additional Features:</b><br>
        The sidebar also provides a random sequence generator for practice (which can also accept uploaded files in CSV, FASTA, or GenBank formats), a glossary of bioinformatics terms, and fun facts about DNA and genomics.
        </p>
        <p>
        If you need to log out, use the "Logout" button in the sidebar. For any step, ensure your input data is correctly formatted to avoid errors during analysis.
        </p>
        </div>
        """, unsafe_allow_html=True)

    # Random Sequence Generator
    with st.sidebar.expander("Random Sequence Generator"):
        st.markdown("Generate random DNA or protein sequences for practice or testing, or upload existing sequences.")
        
        # Input method selection
        rand_input_method = st.radio("Input method", ["Generate Random", "Upload CSV/Excel", "Upload FASTA", "Upload GenBank"], key="rand_input_method")
        
        if rand_input_method == "Generate Random":
            seq_type = st.selectbox("Sequence type", ["DNA", "Protein"], key="rand_seq_type")
            rand_num = st.number_input("Number of sequences", min_value=1, max_value=1000, value=1, key="rand_num")
            rand_len = st.number_input("Length of each sequence", min_value=5, max_value=40000, value=50, key="rand_len")
            if st.button("Generate Random Sequences"):
                import string
                import random as pyrandom
                if seq_type == "DNA":
                    alphabet = "ATGC"
                else:
                    alphabet = "ACDEFGHIKLMNPQRSTVWY"  # 20 amino acids
                
                # Generate sequences in chunks to manage memory
                chunk_size = 100  # Process 100 sequences at a time
                all_sequences = []
                
                for i in range(0, rand_num, chunk_size):
                    current_chunk = min(chunk_size, rand_num - i)
                    chunk_sequences = [(f"Random_{i+j+1}", ''.join(pyrandom.choices(alphabet, k=rand_len))) 
                                     for j in range(current_chunk)]
                    all_sequences.extend(chunk_sequences)
                    
                    # Show progress
                    progress = min(100, int((i + current_chunk) / rand_num * 100))
                    st.progress(progress)
                
                if 'random_sequences' not in st.session_state:
                    st.session_state['random_sequences'] = []
                st.session_state['random_sequences'] = all_sequences
                st.toast(f"{rand_num} random {seq_type} sequences of length {rand_len} generated!", icon=None)
        
        elif rand_input_method == "Upload CSV/Excel":
            rand_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="rand_file")
            if rand_file:
                try:
                    rand_file.seek(0)
                    if rand_file.name.endswith((".xlsx", ".xls")):
                        df = pd.read_excel(rand_file)
                    else:
                        df = pd.read_csv(rand_file)
                    if "Gene Name" in df.columns and "Sequence" in df.columns:
                        df['Sequence'] = df['Sequence'].fillna('').astype(str)
                        df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
                        sequences = list(zip(df["Gene Name"], df["Sequence"]))
                        if 'random_sequences' not in st.session_state:
                            st.session_state['random_sequences'] = []
                        st.session_state['random_sequences'] = sequences
                        st.success(f"Successfully loaded {len(sequences)} sequences from {rand_file.name}")
                    else:
                        st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
                except Exception as e:
                    st.warning(f"Error reading file: {str(e)}")
                    
        elif rand_input_method == "Upload FASTA":
            rand_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="rand_fasta")
            if rand_fasta:
                rand_fasta.seek(0)
                sequences = process_fasta(rand_fasta, max_sequences=1000)
                if sequences:
                    if 'random_sequences' not in st.session_state:
                        st.session_state['random_sequences'] = []
                    st.session_state['random_sequences'] = sequences
                    st.success(f"Successfully loaded {len(sequences)} sequences from FASTA file")
                    
        elif rand_input_method == "Upload GenBank":
            rand_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="rand_gb")
            if rand_gb:
                rand_gb.seek(0)
                sequences = process_genbank(rand_gb, max_sequences=1000)
                if sequences:
                    if 'random_sequences' not in st.session_state:
                        st.session_state['random_sequences'] = []
                    st.session_state['random_sequences'] = sequences
                    st.success(f"Successfully loaded {len(sequences)} sequences from GenBank file")
        
        # Option to clear sequences
        if st.button("Clear Sequences"):
            st.session_state['random_sequences'] = []
            st.toast("Sequences cleared.", icon=None)
        
        # Download random sequences if present
        if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
            # First analyze the random sequences in chunks
            chunk_size = 100  # Process 100 sequences at a time
            all_results = []
            
            for i in range(0, len(st.session_state['random_sequences']), chunk_size):
                chunk = st.session_state['random_sequences'][i:i + chunk_size]
                chunk_results = [analyze_sequence(name, seq) for name, seq in chunk]
                all_results.extend(chunk_results)
                
                # Show progress
                progress = min(100, int((i + len(chunk)) / len(st.session_state['random_sequences']) * 100))
                st.progress(progress)
            
            rand_df = pd.DataFrame(all_results)
            
            # Excel with formatting
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                rand_df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
                # Get the workbook and the worksheet
                workbook = writer.book
                worksheet = writer.sheets['GC Content Analysis']
                
                # Auto-adjust columns' width
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
                # Add some basic formatting
                for row in worksheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = cell.font.copy(bold=True)
            
            excel_buffer.seek(0)
            st.download_button(
                label="Download Random Sequences (Excel)",
                data=excel_buffer,
                file_name="random_sequences.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # FASTA
            fasta_buffer = export_to_fasta(st.session_state['random_sequences'], rand_df)
            st.download_button(
                label="Download Random Sequences (FASTA)",
                data=fasta_buffer,
                file_name="random_sequences.fasta",
                mime="text/plain"
            )
            
            # GenBank
            gb_buffer = export_to_genbank(st.session_state['random_sequences'], rand_df)
            st.download_button(
                label="Download Random Sequences (GenBank)",
                data=gb_buffer,
                file_name="random_sequences.gb",
                mime="text/plain"
            )

    # Add random sequences to main input if present
    if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
        sequences.extend(st.session_state['random_sequences'])
        st.session_state['random_sequences'] = [] # Clear after adding

    # Main content area with professional styling
    st.markdown("---")
    
    # Main Analysis Section
    with st.container():
        st.markdown("""
            <div class="section-container">
                <h2 class="section-header">SNGC - Sequence Analysis & GC Content Calculation</h2>
            </div>
        """, unsafe_allow_html=True)
        
        st.info("**Professional Tip:** You can input up to 1000 sequences at a time (CSV, Excel, FASTA, GenBank, or manual entry). If you upload more, only the first 1000 will be processed.")

        input_method = st.radio("Choose input method", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"])

        if input_method == "Upload CSV/Excel":
            uploaded_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"])
            if uploaded_file:
                try:
                    uploaded_file.seek(0) # Ensure pointer is at the start
                    if uploaded_file.name.endswith((".xlsx", ".xls")):
                        df = pd.read_excel(uploaded_file)
                    else:
                        df = pd.read_csv(uploaded_file)
                    
                    if "Gene Name" in df.columns and "Sequence" in df.columns:
                        # Convert sequences to strings and handle any NaN values
                        df['Sequence'] = df['Sequence'].fillna('').astype(str)
                        df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
                        if len(df) > 1000:
                            st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
                        sequences = list(zip(df["Gene Name"], df["Sequence"]))[:1000]
                    else:
                        st.toast("File must contain 'Gene Name' and 'Sequence' columns.", icon=None)
                except Exception as e:
                    st.toast(f"Error reading file: {str(e)}", icon=None)

        elif input_method == "Upload FASTA":
            fasta_file = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"])
            if fasta_file:
                fasta_file.seek(0)  # Reset pointer
                try:
                    sequences = process_fasta(fasta_file, max_sequences=1000)
                    if len(sequences) == 0:
                        st.warning("No sequences found in the uploaded FASTA file.")
                    elif len(sequences) == 1000:
                        st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
                except Exception as e:
                    st.toast(f"Error processing FASTA file: {str(e)}", icon=None)

        elif input_method == "Upload GenBank":
            gb_file = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"])
            if gb_file:
                gb_file.seek(0)  # Reset pointer
                try:
                    sequences = process_genbank(gb_file, max_sequences=1000)
                    if len(sequences) == 0:
                        st.warning("No sequences found in the uploaded GenBank file.")
                    elif len(sequences) == 1000:
                        st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
                except Exception as e:
                    st.toast(f"Error processing GenBank file: {str(e)}", icon=None)

        elif input_method == "Manual Entry":
            num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1)
            manual_entries = []
            for i in range(num):
                st.markdown(f"**Sequence {i+1}**")
                name = st.text_input(f"Gene Name {i+1}", key=f"name_{i}")
                seq = st.text_area(f"Sequence {i+1}", key=f"seq_{i}")
                if name and seq:
                    manual_entries.append((name, seq))
            sequences.extend(manual_entries)

        # Add random sequences to main input if present
        if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
            sequences.extend(st.session_state['random_sequences'])
            st.session_state['random_sequences'] = [] # Clear after adding

        if sequences:
            # Fragment length selector
            st.markdown("### Fragment Analysis Settings")
            fragment_length = st.selectbox(
                "Select fragment length for analysis:",
                options=[50, 100, 200, 400],
                index=2,  # Default to 200
                help="Choose the length of fragments to split your sequences into for detailed analysis"
            )
            
            if st.button("Calculate GC Content & Fragment Analysis"):
                # Analyze original sequences
                results = [analyze_sequence(name, seq) for name, seq in sequences]
                result_df = pd.DataFrame(results)

                if "Error" in result_df.columns:
                    st.toast("Some sequences were invalid and skipped.", icon=None)
                    result_df = result_df.dropna(subset=["Length"])

                # Check if there are valid sequences for output
                if len(result_df) == 0:
                     st.warning("No valid sequences were processed.")
                else:
                    # Create fragments for all sequences
                    all_fragments = []
                    fragment_analyses = []
                    
                    for name, seq in sequences:
                        if seq and len(seq) > 0:
                            # Create fragments with selected length
                            fragments = fragment_sequences([(name, seq)], fragment_length)
                            all_fragments.extend(fragments)
                            
                            # Extract fragment analyses for display
                            for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
                                fragment_analyses.append({
                                    "Gene Name": frag_name,
                                    "Parent Gene": parent_name,
                                    "Fragment Number": frag_num,
                                    "Start Position": start_pos,
                                    "End Position": start_pos + len(frag_seq) - 1,
                                    "Fragment Length": len(frag_seq),
                                    "Sequence": frag_seq,
                                    "Length": analysis.get("Length", len(frag_seq)),
                                    "A Count": analysis.get("A Count", 0),
                                    "T Count": analysis.get("T Count", 0),
                                    "G Count": analysis.get("G Count", 0),
                                    "C Count": analysis.get("C Count", 0),
                                    "A %": analysis.get("A %", 0),
                                    "T %": analysis.get("T %", 0),
                                    "G %": analysis.get("G %", 0),
                                    "C %": analysis.get("C %", 0),
                                    "GC %": analysis.get("GC %", 0),
                                    "AT %": analysis.get("AT %", 0),
                                    "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
                                })
                    
                    fragment_df = pd.DataFrame(fragment_analyses)

                    # Conditional Visualization Display
                    if len(result_df) > 50:
                        st.info("Analysis complete. Visualizations and on-page data table are disabled for more than 50 sequences to ensure performance. Please download your full results below.")
                    else:
                        st.toast("Analysis complete with automatic fragmentation!", icon=None)
                        st.markdown("---")
                        
                        # Results Section
                        st.markdown("""
                            <div class="section-container">
                                <h2 class="section-header">Results & Visualizations</h2>
                            </div>
                        """, unsafe_allow_html=True)
                        
                        # Original Sequences Analysis
                        st.markdown("### Original Sequences Analysis")
                        with st.expander("Show Original Sequences Data", expanded=True):
                            st.dataframe(result_df, use_container_width=True)
                        
                        # Fragment Analysis
                        st.markdown(f"### Fragment Analysis ({fragment_length}-base chunks)")
                        with st.expander("Show Fragment Data", expanded=True):
                            st.dataframe(fragment_df, use_container_width=True)
                            
                        # Fragment Statistics
                        if len(fragment_df) > 0:
                            st.markdown("### Fragment Statistics")
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                st.metric("Total Fragments", len(fragment_df))
                                st.metric("Average GC Content", f"{fragment_df['GC %'].mean():.2f}%")
                                
                            with col2:
                                st.metric("Average Z-DNA Propensity", f"{fragment_df['Z-DNA Propensity %'].mean():.2f}%")
                                st.metric("High GC Fragments (>60%)", len(fragment_df[fragment_df['GC %'] > 60]))
                                
                            with col3:
                                st.metric("Low GC Fragments (<40%)", len(fragment_df[fragment_df['GC %'] < 40]))
                                st.metric("High Z-DNA Fragments (>50%)", len(fragment_df[fragment_df['Z-DNA Propensity %'] > 50]))
                        with st.expander("Show GC Content Heatmap"):
                            if len(result_df) > 0:
                                create_gc_heatmap(result_df)
                            else:
                                st.info("No data to plot.")
                        with st.expander("Show Sequence Logo"):
                            # Use the original 'sequences' list for the logo, not result_df
                            if len(sequences) > 1 and all(len(seq) > 0 for _, seq in sequences):
                                # Filter original sequences based on names in result_df to match analysis results
                                valid_gene_names = result_df['Gene Name'].tolist()
                                sequences_for_logo = [(name, seq) for name, seq in sequences if name in valid_gene_names]
                                if len(sequences_for_logo) > 1:
                                    create_sequence_logo([seq for _, seq in sequences_for_logo])
                                else:
                                    st.info("Not enough valid sequences to plot a sequence logo.")
                            else:
                                st.info("Not enough data to plot a sequence logo.")

                        with st.expander("Show Interactive Plots"):
                            if len(result_df) > 1:
                                create_interactive_plots(result_df)
                            else:
                                st.info("Not enough data to plot interactive plots.")

                    # Export Results section - always shown if there are results
                    st.markdown("---")
                    st.markdown("""
                        <div class="section-container">
                            <h2 class="section-header">SNGC Export Results</h2>
                        </div>
                    """, unsafe_allow_html=True)
                    
                    output_filename = st.text_input("Enter output file name (without extension):", "gc_output_v4", key="export_filename")
                    
                    # Export buttons in columns for better layout
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        if st.button("Export All Results"):
                            zip_buffer = export_all_to_zip(result_df, output_filename)
                            st.toast("All results exported as ZIP!", icon=None)
                            st.download_button(
                                label="Download All Results (ZIP)",
                                data=zip_buffer,
                                file_name=f"{output_filename}_all_results.zip",
                                mime="application/zip"
                            )
                    
                    with col2:
                        excel_buffer = BytesIO()
                        result_df.to_excel(excel_buffer, index=False, engine='openpyxl')
                        excel_buffer.seek(0)
                        st.download_button(
                            label="Download Original Sequences (Excel)",
                            data=excel_buffer,
                            file_name=f"{output_filename}_original.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    with col3:
                        # FASTA export button
                        # Use result_df for FASTA/GenBank exports to include only analyzed sequences
                        fasta_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
                        fasta_buffer = export_to_fasta(fasta_sequences_to_export, result_df)
                        st.download_button(
                            label="Download Original Sequences (FASTA)",
                            data=fasta_buffer,
                            file_name=f"{output_filename}_original.fasta",
                            mime="text/plain"
                        )
                    
                    with col4:
                        # GenBank export button
                        gb_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
                        gb_buffer = export_to_genbank(gb_sequences_to_export, result_df)
                        st.download_button(
                            label="Download Original Sequences (GenBank)",
                            data=gb_buffer,
                            file_name=f"{output_filename}_original.gb",
                            mime="text/plain"
                        )
                    
                    # Fragment Export Section
                    if len(fragment_df) > 0:
                        st.markdown("### Fragment Export Options")
                        frag_col1, frag_col2, frag_col3, frag_col4 = st.columns(4)
                        
                        with frag_col1:
                            # Export fragments to Excel
                            frag_excel_buffer = export_fragments_to_excel(all_fragments)
                            st.download_button(
                                label="Download Fragments (Excel)",
                                data=frag_excel_buffer,
                                file_name=f"{output_filename}_fragments_{fragment_length}bp.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                        
                        with frag_col2:
                            # Export fragments to FASTA
                            frag_fasta_buffer = export_fragments_to_fasta(all_fragments)
                            st.download_button(
                                label="Download Fragments (FASTA)",
                                data=frag_fasta_buffer,
                                file_name=f"{output_filename}_fragments_{fragment_length}bp.fasta",
                                mime="text/plain"
                            )
                        
                        with frag_col3:
                            # Export fragments to GenBank
                            frag_gb_buffer = export_fragments_to_genbank(all_fragments)
                            st.download_button(
                                label="Download Fragments (GenBank)",
                                data=frag_gb_buffer,
                                file_name=f"{output_filename}_fragments_{fragment_length}bp.gb",
                                mime="text/plain"
                            )
                        
                        with frag_col4:
                            # Combined export with both original and fragments
                            combined_data = []
                            
                            # Add original sequences
                            for _, row in result_df.iterrows():
                                combined_data.append({
                                    "Type": "Original",
                                    "Gene Name": row["Gene Name"],
                                    "Sequence": row["Sequence"],
                                    "Length": row["Length"],
                                    "GC %": row["GC %"],
                                    "Z-DNA Propensity %": row.get("Z-DNA Propensity %", 0),
                                    "A Count": row["A Count"],
                                    "T Count": row["T Count"],
                                    "G Count": row["G Count"],
                                    "C Count": row["C Count"]
                                })
                            
                            # Add fragments
                            for _, row in fragment_df.iterrows():
                                combined_data.append({
                                    "Type": "Fragment",
                                    "Gene Name": row["Gene Name"],
                                    "Parent Gene": row["Parent Gene"],
                                    "Fragment Number": row["Fragment Number"],
                                    "Start Position": row["Start Position"],
                                    "End Position": row["End Position"],
                                    "Sequence": row["Sequence"],
                                    "Length": row["Length"],
                                    "GC %": row["GC %"],
                                    "Z-DNA Propensity %": row["Z-DNA Propensity %"],
                                    "A Count": row["A Count"],
                                    "T Count": row["T Count"],
                                    "G Count": row["G Count"],
                                    "C Count": row["C Count"]
                                })
                            
                            combined_df = pd.DataFrame(combined_data)
                            combined_buffer = BytesIO()
                            combined_df.to_excel(combined_buffer, index=False, engine='openpyxl')
                            combined_buffer.seek(0)
                            st.download_button(
                                label="Download Combined (Excel)",
                                data=combined_buffer,
                                file_name=f"{output_filename}_combined_{fragment_length}bp.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )

    # Add Sequence Fragmenter tool to sidebar
    with st.sidebar.expander("Sequence Fragmenter (Split Tool)"):
        st.markdown("""
        <div style='background: rgba(52,152,219,0.1); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3); margin-bottom: 1rem;'>
        <strong>Split sequences into fragments of a chosen length.</strong><br>
        Output is compatible with TOOL-1 and available in multiple formats.
        </div>
        """, unsafe_allow_html=True)
        
        frag_input_method = st.radio("Input method for fragmenter", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"], key="frag_input_method")
        frag_sequences = []
        
        if frag_input_method == "Upload CSV/Excel":
            frag_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="frag_file")
            if frag_file:
                try:
                    frag_file.seek(0)
                    if frag_file.name.endswith((".xlsx", ".xls")):
                        df = pd.read_excel(frag_file)
                    else:
                        df = pd.read_csv(frag_file)
                    if "Gene Name" in df.columns and "Sequence" in df.columns:
                        df['Sequence'] = df['Sequence'].fillna('').astype(str)
                        df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
                        frag_sequences = list(zip(df["Gene Name"], df["Sequence"]))
                        st.success(f"Successfully loaded {len(frag_sequences)} sequences from {frag_file.name}")
                    else:
                        st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
                except Exception as e:
                    st.warning(f"Error reading file: {str(e)}")
                    
        elif frag_input_method == "Upload FASTA":
            frag_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="frag_fasta")
            if frag_fasta:
                frag_fasta.seek(0)
                frag_sequences = process_fasta(frag_fasta, max_sequences=1000)
                if frag_sequences:
                    st.success(f"Successfully loaded {len(frag_sequences)} sequences from FASTA file")
                    
        elif frag_input_method == "Upload GenBank":
            frag_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="frag_gb")
            if frag_gb:
                frag_gb.seek(0)
                frag_sequences = process_genbank(frag_gb, max_sequences=1000)
                if frag_sequences:
                    st.success(f"Successfully loaded {len(frag_sequences)} sequences from GenBank file")
                    
        elif frag_input_method == "Manual Entry":
            frag_num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1, key="frag_num")
            frag_manual_entries = []
            for i in range(frag_num):
                st.markdown(f"**Sequence {i+1}**")
                name = st.text_input(f"Gene Name {i+1}", key=f"frag_name_{i}")
                seq = st.text_area(f"Sequence {i+1}", key=f"frag_seq_{i}")
                if name and seq:
                    frag_manual_entries.append((name, seq))
            frag_sequences.extend(frag_manual_entries)
            if frag_manual_entries:
                st.success(f"Successfully entered {len(frag_manual_entries)} sequences")
                
        # Fragment length selection with better styling
        st.markdown("---")
        st.markdown("**Fragment Configuration**")
        frag_length = st.selectbox("Fragment length (nucleotides)", [50, 100, 200, 400], key="frag_length")
        
        if frag_sequences:
            st.info(f"Ready to process {len(frag_sequences)} sequences into {frag_length}-nucleotide fragments")
            
            if st.button("Split and Export Fragments", key="frag_button"):
                with st.spinner("Processing fragments and calculating GC content..."):
                    fragments = fragment_sequences(frag_sequences, frag_length)
                    
                if not fragments:
                    st.warning("No fragments produced. Check your input sequences and fragment length.")
                else:
                    st.success(f"Successfully produced {len(fragments)} fragments of {frag_length} nucleotides each!")
                    
                    # Calculate overall statistics
                    total_gc_content = sum(f[5].get("GC %", 0) for f in fragments)
                    avg_gc_content = total_gc_content / len(fragments)
                    
                    # Display fragment statistics with GC content
                    st.markdown("**Fragment Statistics:**")
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        st.markdown(f"""
                        - Total fragments: {len(fragments)}
                        - Fragment length: {frag_length} nucleotides
                        - Parent sequences: {len(set([f[2] for f in fragments]))}
                        - Average GC content: {avg_gc_content:.2f}%
                        """)
                    
                    with col2:
                        # Show GC content distribution
                        gc_values = [f[5].get("GC %", 0) for f in fragments]
                        min_gc = min(gc_values)
                        max_gc = max(gc_values)
                        st.markdown(f"""
                        - GC content range: {min_gc:.2f}% - {max_gc:.2f}%
                        - High GC fragments (>60%): {sum(1 for gc in gc_values if gc > 60)}
                        - Low GC fragments (<40%): {sum(1 for gc in gc_values if gc < 40)}
                        """)
                    
                    # Export options in a more organized way
                    st.markdown("**Download Options**")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        excel_buffer = export_fragments_to_excel(fragments)
                        st.download_button(
                            label="Excel Format (with Analysis)",
                            data=excel_buffer,
                            file_name=f"fragments_{frag_length}n_with_analysis.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        fasta_buffer = export_fragments_to_fasta(fragments)
                        st.download_button(
                            label="FASTA Format (with Analysis)",
                            data=fasta_buffer,
                            file_name=f"fragments_{frag_length}n_with_analysis.fasta",
                            mime="text/plain"
                        )
                    
                    with col2:
                        gb_buffer = export_fragments_to_genbank(fragments)
                        st.download_button(
                            label="GenBank Format (with Analysis)",
                            data=gb_buffer,
                            file_name=f"fragments_{frag_length}n_with_analysis.gb",
                            mime="text/plain"
                        )
                        
                        # Show a sample of the analysis data
                        if len(fragments) > 0:
                            st.markdown("**Sample Fragment Analysis:**")
                            sample_frag = fragments[0]
                            st.markdown(f"""
                            **{sample_frag[0]}** (from {sample_frag[2]})
                            - GC Content: {sample_frag[5].get('GC %', 0):.2f}%
                            - A: {sample_frag[5].get('A Count', 0)} ({sample_frag[5].get('A %', 0):.2f}%)
                            - T: {sample_frag[5].get('T Count', 0)} ({sample_frag[5].get('T %', 0):.2f}%)
                            - G: {sample_frag[5].get('G Count', 0)} ({sample_frag[5].get('G %', 0):.2f}%)
                            - C: {sample_frag[5].get('C Count', 0)} ({sample_frag[5].get('C %', 0):.2f}%)
                            """)
        else:
            st.info("Please upload or enter sequences to begin fragmentation")

    # Professional footer
    st.markdown("---")
    st.markdown("""
        <div class="footer">
            <p>Developed by Shubh Rakesh Nahar | Troy University |  📧 shubhrakeshnahar@gmail.com </p>
        </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()





# import pandas as pd
# import streamlit as st
# import matplotlib.pyplot as plt
# from io import BytesIO
# from Bio import SeqIO
# import json
# import plotly.express as px
# import plotly.graph_objects as go
# import seaborn as sns
# import numpy as np
# from matplotlib.patches import Rectangle
# import math
# import zipfile
# import random
# from Bio.Seq import Seq
# from Bio.SeqRecord import SeqRecord
# import io
# import openpyxl

# # Custom CSS for professional styling
# def load_css():
#     st.markdown("""
#     <style>
#     /* Main background and overall styling */
#     .main {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#         color: #ffffff;
#     }
    
#     .stApp {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#     }
    
#     /* Sidebar styling - Overlay mode */
#     .css-1d391kg {
#         background: linear-gradient(180deg, #0f1419 0%, #1a2332 50%, #0f1419 100%);
#         border-right: 2px solid #3498db;
#         position: fixed !important;
#         top: 0;
#         left: 0;
#         height: 100vh;
#         z-index: 1000;
#         width: 21rem;
#         overflow-y: auto;
#         box-shadow: 2px 0 10px rgba(0,0,0,0.3);
#         transition: transform 0.3s ease;
#     }
    
#     /* Main content area - adjust for overlay sidebar */
#     .main .block-container {
#         margin-left: 0 !important;
#         padding-left: 1rem !important;
#         max-width: none !important;
#     }
    
#     /* Sidebar toggle button styling */
#     .css-1rs6os {
#         position: fixed !important;
#         top: 1rem;
#         left: 1rem;
#         z-index: 1001;
#         background: rgba(52, 152, 219, 0.9) !important;
#         border-radius: 50% !important;
#         width: 3rem !important;
#         height: 3rem !important;
#         display: flex !important;
#         align-items: center !important;
#         justify-content: center !important;
#         box-shadow: 0 2px 10px rgba(0,0,0,0.3) !important;
#     }
    
#     /* Header styling */
#     .main-header {
#         background: linear-gradient(90deg, #2c3e50 0%, #34495e 50%, #2c3e50 100%);
#         padding: 2rem;
#         border-radius: 15px;
#         margin-bottom: 2rem;
#         box-shadow: 0 8px 32px rgba(0,0,0,0.4);
#         border: 1px solid #3498db;
#     }
    
#     .main-header h1 {
#         color: #ecf0f1;
#         text-align: center;
#         font-size: 2.5rem;
#         font-weight: 700;
#         margin: 0;
#         text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
#     }
    
#     /* Section styling */
#     .section-container {
#         background: rgba(255, 255, 255, 0.05);
#         padding: 2rem;
#         border-radius: 15px;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         backdrop-filter: blur(10px);
#         box-shadow: 0 8px 32px rgba(0,0,0,0.3);
#     }
    
#     .section-header {
#         color: #3498db;
#         font-size: 1.5rem;
#         font-weight: 600;
#         margin-bottom: 1rem;
#         text-align: center;
#         text-transform: uppercase;
#         letter-spacing: 1px;
#     }
    
#     /* Button styling */
#     .stButton > button {
#         background: linear-gradient(45deg, #3498db, #2980b9);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
#     }
    
#     .stButton > button:hover {
#         background: linear-gradient(45deg, #2980b9, #1f5f8b);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
#     }
    
#     /* Input styling */
#     .stTextInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stTextArea > div > div > textarea {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stSelectbox > div > div > select {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* File uploader styling */
#     .stFileUploader > div {
#         background: rgba(255, 255, 255, 0.05);
#         border: 2px dashed rgba(52, 152, 219, 0.5);
#         border-radius: 15px;
#         padding: 2rem;
#         text-align: center;
#     }
    
#     /* Radio button styling */
#     .stRadio > div {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         padding: 1rem;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Number input styling */
#     .stNumberInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* Expander styling */
#     .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #34495e, #2c3e50);
#         color: #ecf0f1;
#         border-radius: 10px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Success/Info/Warning styling */
#     .stSuccess {
#         background: rgba(46, 204, 113, 0.2);
#         border: 1px solid #2ecc71;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #2ecc71;
#     }
    
#     .stInfo {
#         background: rgba(52, 152, 219, 0.2);
#         border: 1px solid #3498db;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #3498db;
#     }
    
#     .stWarning {
#         background: rgba(241, 196, 15, 0.2);
#         border: 1px solid #f1c40f;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #f1c40f;
#     }
    
#     /* Dataframe styling */
#     .dataframe {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         overflow: hidden;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Download button styling */
#     .stDownloadButton > button {
#         background: linear-gradient(45deg, #27ae60, #229954);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
#     }
    
#     .stDownloadButton > button:hover {
#         background: linear-gradient(45deg, #229954, #1e8449);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
#     }
    
#     /* Progress bar styling */
#     .stProgress > div > div > div {
#         background: linear-gradient(90deg, #3498db, #2980b9);
#         border-radius: 10px;
#     }
    
#     /* Sidebar expander styling */
#     .css-1d391kg .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #2c3e50, #34495e);
#         color: #ecf0f1;
#         border-radius: 8px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Footer styling */
#     .footer {
#         text-align: center;
#         color: #bdc3c7;
#         font-size: 0.9rem;
#         margin-top: 3rem;
#         padding: 1rem;
#         border-top: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Tool card styling */
#     .tool-card {
#         background: rgba(255, 255, 255, 0.08);
#         border-radius: 15px;
#         padding: 1.5rem;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         box-shadow: 0 4px 15px rgba(0,0,0,0.3);
#     }
    
#     /* Fragmenter specific styling */
#     .fragmenter-section {
#         background: linear-gradient(135deg, rgba(52, 152, 219, 0.1) 0%, rgba(41, 128, 185, 0.1) 100%);
#         border: 2px solid rgba(52, 152, 219, 0.5);
#         border-radius: 20px;
#         padding: 2rem;
#         margin: 1rem 0;
#     }
    
#     /* Responsive design */
#     @media (max-width: 768px) {
#         .main-header h1 {
#             font-size: 2rem;
#         }
        
#         .section-container {
#             padding: 1rem;
#         }
        
#         .css-1d391kg {
#             width: 100vw;
#         }
#     }
#     </style>
#     """, unsafe_allow_html=True)

# # --- Instructions ---
# def show_instructions():
#     st.markdown("""
#     ### 📋 Instructions
    
#     1. **Choose Input Method**:
#        - **Upload CSV/Excel**: Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns
#        - **Upload FASTA**: Upload a FASTA file containing DNA sequences
#        - **Upload GenBank**: Upload a GenBank file containing DNA sequences
#        - **Manual Entry**: Enter sequences manually with gene names
    
#     2. **Input Format**:
#        - Only A, T, G, C nucleotides are accepted (case insensitive)
#        - Invalid characters will be automatically removed
#        - Each sequence must have a unique gene name
    
#     3. **Analysis**:
#        - Click "Calculate GC Content" to process sequences
#        - View results in the interactive table
#        - Explore visualizations of GC content and nucleotide composition
    
#     4. **Export Options**:
#        - Export results in multiple formats (Excel, CSV, JSON)
#        - Use "Export All Results" for a complete dataset export
#        - Customize output filename before downloading
#     """)

# # --- Sequence Sanitization and Validation ---
# def sanitize_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return ''.join(filter(lambda x: x.upper() in ['A', 'T', 'G', 'C'], seq.upper()))

# def is_valid_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return all(base in ['A', 'T', 'G', 'C'] for base in seq.upper())

# # --- Z-DNA Prediction Function ---
# def calculate_z_dna_propensity(sequence):
#     """
#     Calculate Z-DNA propensity based on alternating purine-pyrimidine patterns
#     and GC-rich regions that are prone to Z-DNA formation.
#     """
#     if not sequence or len(sequence) < 6:
#         return 0.0
    
#     sequence = sequence.upper()
#     z_score = 0.0
#     total_positions = 0
    
#     # Check for alternating purine-pyrimidine patterns (Z-DNA prone)
#     for i in range(len(sequence) - 5):
#         window = sequence[i:i+6]
#         if len(window) == 6:
#             # Check for alternating pattern
#             alternating = True
#             for j in range(1, 6):
#                 current = window[j] in 'AG'  # Purine
#                 prev = window[j-1] in 'AG'   # Purine
#                 if current == prev:  # Not alternating
#                     alternating = False
#                     break
            
#             if alternating:
#                 # Calculate GC content in this window
#                 gc_count = window.count('G') + window.count('C')
#                 gc_content = gc_count / 6.0
                
#                 # Z-DNA is more likely in GC-rich alternating regions
#                 if gc_content >= 0.5:  # 50% or more GC
#                     z_score += gc_content * 2.0  # Higher weight for GC-rich
#                 else:
#                     z_score += gc_content * 1.0  # Lower weight for AT-rich
                
#                 total_positions += 1
    
#     # Normalize by sequence length and total positions checked
#     if total_positions > 0:
#         return (z_score / total_positions) * 100  # Convert to percentage
#     else:
#         return 0.0

# # --- Enhanced Nucleotide Analysis with Z-DNA ---
# def analyze_sequence(name, seq):
#     # Convert sequence to string if it's not already
#     if not isinstance(seq, str):
#         seq = str(seq)
    
#     seq = sanitize_sequence(seq)
#     length = len(seq)
#     if length == 0:
#         return {"Gene Name": name, "Error": "Invalid sequence or empty after sanitization"}

#     a_count = seq.count('A')
#     t_count = seq.count('T')
#     g_count = seq.count('G')
#     c_count = seq.count('C')
#     gc_count = g_count + c_count
#     at_count = a_count + t_count

#     # Calculate Z-DNA propensity
#     z_dna_propensity = calculate_z_dna_propensity(seq)

#     return {
#         "Gene Name": name,
#         "Sequence": seq,
#         "Length": length,
#         "A Count": a_count,
#         "T Count": t_count,
#         "G Count": g_count,
#         "C Count": c_count,
#         "A %": round((a_count / length) * 100, 2),
#         "T %": round((t_count / length) * 100, 2),
#         "G %": round((g_count / length) * 100, 2),
#         "C %": round((c_count / length) * 100, 2),
#         "GC %": round((gc_count / length) * 100, 2),
#         "AT %": round((at_count / length) * 100, 2),
#         "Z-DNA Propensity %": round(z_dna_propensity, 2),
#     }

# def process_fasta(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for FASTA
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "fasta")
#         count = 0
#         for record in records:
#             if count >= max_sequences:
#                 break
#             # Convert Seq object to string
#             sequences.append((str(record.id), str(record.seq)))
#             count += 1
#         if count == 0:
#             st.warning("No sequences found in the uploaded FASTA file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the FASTA file.")
#     except Exception as e:
#         st.error(f"Error parsing FASTA file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def process_genbank(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for GenBank
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "genbank")
#         count = 0
#         for record in records:
#              if count >= max_sequences:
#                  break
#              # Ensure molecule_type is set for GenBank export compatibility later
#              # For uploaded GenBank, try to get molecule_type from annotations, default to DNA
#              mol_type = record.annotations.get("molecule_type", "DNA")
#              sequences.append((str(record.id), str(record.seq)))
#              count += 1
#         if count == 0:
#              st.warning("No sequences found in the uploaded GenBank file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the GenBank file.")

#     except Exception as e:
#         st.error(f"Error parsing GenBank file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def display_visuals(df):
#     st.subheader("📊 GC Content Distribution")
#     fig, ax = plt.subplots()
#     df.plot.bar(x='Gene Name', y='GC %', ax=ax, color='purple', legend=False)
#     plt.ylabel('GC %')
#     plt.xticks(rotation=45, ha='right')
#     st.pyplot(fig)

#     st.subheader("🧬 Nucleotide Composition Per Sequence")
#     for _, row in df.iterrows():
#         fig, ax = plt.subplots()
#         ax.pie(
#             [row['A %'], row['T %'], row['G %'], row['C %']],
#             labels=['A %', 'T %', 'G %', 'C %'],
#             autopct='%1.1f%%',
#             startangle=90
#         )
#         ax.set_title(f"{row['Gene Name']} - Base % Composition")
#         st.pyplot(fig)

# def export_data(df, format_type):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     if format_type == "Excel":
#         towrite = BytesIO()
#         with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         towrite.seek(0)
#         return towrite, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
#     elif format_type == "CSV":
#         towrite = BytesIO()
#         # Add signature as a comment in the first line
#         towrite.write(f"# {signature}\n".encode())
#         df.to_csv(towrite, index=False)
#         towrite.seek(0)
#         return towrite, "text/csv", "csv"
#     else:  # JSON
#         towrite = BytesIO()
#         # Add signature to the JSON metadata
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         towrite.write(json_str.encode())
#         towrite.seek(0)
#         return towrite, "application/json", "json"

# def create_gc_heatmap(df):
#     st.subheader("🌡️ GC Content Heatmap")
    
#     # Create a matrix of GC content for each position
#     sequences = df['Sequence'].tolist()
#     max_len = max(len(seq) for seq in sequences)
    
#     # Initialize matrix
#     gc_matrix = np.zeros((len(sequences), max_len))
    
#     # Fill matrix with GC content for each position
#     for i, seq in enumerate(sequences):
#         for j in range(len(seq)):
#             if j < len(seq):
#                 window = seq[max(0, j-10):min(len(seq), j+11)]
#                 gc_count = window.count('G') + window.count('C')
#                 gc_matrix[i, j] = (gc_count / len(window)) * 100
    
#     # Create heatmap using plotly
#     fig = go.Figure(data=go.Heatmap(
#         z=gc_matrix,
#         x=list(range(max_len)),
#         y=df['Gene Name'].tolist(),
#         colorscale='Viridis',
#         colorbar=dict(title='GC %')
#     ))
    
#     fig.update_layout(
#         title='GC Content Distribution Across Sequences',
#         xaxis_title='Position',
#         yaxis_title='Gene Name',
#         height=400 + (len(sequences) * 20)  # Adjust height based on number of sequences
#     )
    
#     st.plotly_chart(fig, use_container_width=True)

# def calculate_information_content(freq):
#     """Calculate information content in bits."""
#     if freq == 0:
#         return 0
#     return freq * math.log2(freq * 4)  # 4 for number of nucleotides

# def create_sequence_logo(sequences):
#     st.subheader("🎨 Sequence Logo")
    
#     # Calculate position frequency matrix
#     max_len = max(len(seq) for seq in sequences)
#     pfm = np.zeros((4, max_len))  # 4 nucleotides
    
#     for seq in sequences:
#         for i, base in enumerate(seq):
#             if i < max_len:
#                 if base == 'A':
#                     pfm[0, i] += 1
#                 elif base == 'T':
#                     pfm[1, i] += 1
#                 elif base == 'G':
#                     pfm[2, i] += 1
#                 elif base == 'C':
#                     pfm[3, i] += 1
    
#     # Normalize
#     pfm = pfm / len(sequences)
    
#     # Calculate information content
#     ic = np.zeros(max_len)
#     for i in range(max_len):
#         ic[i] = sum(calculate_information_content(freq) for freq in pfm[:, i])
    
#     # Create the plot
#     fig, ax = plt.subplots(figsize=(12, 4))
    
#     # Colors for nucleotides
#     colors = {'A': '#2ecc71', 'T': '#e74c3c', 'G': '#f1c40f', 'C': '#3498db'}
#     bases = ['A', 'T', 'G', 'C']
    
#     # Plot each position
#     for i in range(max_len):
#         # Sort frequencies for this position
#         freqs = pfm[:, i]
#         sorted_indices = np.argsort(freqs)
        
#         # Plot each base
#         y_bottom = 0
#         for idx in sorted_indices:
#             if freqs[idx] > 0:
#                 height = freqs[idx] * ic[i]
#                 rect = Rectangle((i, y_bottom), 1, height,
#                                facecolor=colors[bases[idx]],
#                                edgecolor='black',
#                                linewidth=0.5)
#                 ax.add_patch(rect)
#                 y_bottom += height
    
#     # Customize the plot
#     ax.set_xlim(0, max_len)
#     ax.set_ylim(0, max(ic) * 1.1)
#     ax.set_xlabel('Position')
#     ax.set_ylabel('Bits')
#     ax.set_title('Sequence Logo')
    
#     # Add legend
#     legend_elements = [Rectangle((0, 0), 1, 1, facecolor=color, edgecolor='black')
#                       for color in colors.values()]
#     ax.legend(legend_elements, bases, loc='upper right')
    
#     # Remove spines
#     for spine in ['top', 'right']:
#         ax.spines[spine].set_visible(False)
    
#     # Show the plot
#     st.pyplot(fig)
#     plt.close(fig)

# def create_interactive_plots(df):
#     st.subheader("📊 Interactive Plots")
    
#     # GC Content Distribution
#     fig_gc = px.box(df, y='GC %', title='GC Content Distribution',
#                     hover_data=['Gene Name', 'Length', 'GC %'])
#     fig_gc.update_traces(marker_color='purple')
#     st.plotly_chart(fig_gc, use_container_width=True)
    
#     # Nucleotide Composition
#     fig_comp = px.bar(df, 
#                      x='Gene Name',
#                      y=['A %', 'T %', 'G %', 'C %'],
#                      title='Nucleotide Composition by Gene',
#                      barmode='group',
#                      hover_data=['Length', 'GC %'])
#     st.plotly_chart(fig_comp, use_container_width=True)
    
#     # Length vs GC Content
#     fig_scatter = px.scatter(df,
#                            x='Length',
#                            y='GC %',
#                            color='GC %',
#                            hover_data=['Gene Name', 'Length', 'GC %'],
#                            title='Sequence Length vs GC Content')
#     st.plotly_chart(fig_scatter, use_container_width=True)

# def export_all_to_zip(df, output_filename):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     zip_buffer = BytesIO()
#     with zipfile.ZipFile(zip_buffer, "w") as zip_file:
#         # Excel with formatting
#         excel_buffer = BytesIO()
#         with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         excel_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.xlsx", excel_buffer.read())
        
#         # CSV
#         csv_buffer = BytesIO()
#         # Add signature as a comment in the first line
#         csv_buffer.write(f"# {signature}\n".encode())
#         df.to_csv(csv_buffer, index=False)
#         csv_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.csv", csv_buffer.read())
        
#         # JSON
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         json_buffer = BytesIO()
#         json_buffer.write(json_str.encode())
#         json_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.json", json_buffer.read())
    
#     zip_buffer.seek(0)
#     return zip_buffer

# def export_to_fasta(sequences, result_df):
#     """Export sequences to FASTA format with analysis results in description"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     # Add signature as a comment at the top
#     output.write(f"# {signature}\n")
#     for idx, row in result_df.iterrows():
#         # Create a description with analysis results
#         description = f"GC Content: {row['GC %']:.2f}%, Length: {row['Length']}, "
#         description += f"A: {row['A Count']}, T: {row['T Count']}, G: {row['G Count']}, C: {row['C Count']}"
        
#         # Write sequence with description
#         output.write(f">{row['Gene Name']} {description}\n")
#         output.write(f"{row['Sequence']}\n")
    
#     # Add signature at the end
#     output.write(f"\n# {signature}\n")
#     return output.getvalue()

# def export_to_genbank(sequences, result_df):
#     """Export sequences to GenBank format with analysis results in features"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     for idx, row in result_df.iterrows():
#         # Write GenBank header
#         output.write(f"LOCUS       {row['Gene Name']}              {row['Length']} bp    DNA     linear\n")
#         output.write(f"DEFINITION  {row['Gene Name']}\n")
#         output.write(f"ACCESSION   {row['Gene Name']}\n")
#         output.write(f"VERSION     {row['Gene Name']}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
        
#         # Add signature in the features section
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write("     source          1..{}\n".format(row['Length']))
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write('                     /note="{}"\n'.format(signature))
#         output.write('     misc_feature    1..{}\n'.format(row['Length']))
#         output.write('                     /note="GC Content: {:.2f}%"\n'.format(row['GC %']))
#         output.write('                     /note="A Count: {}"\n'.format(row['A Count']))
#         output.write('                     /note="T Count: {}"\n'.format(row['T Count']))
#         output.write('                     /note="G Count: {}"\n'.format(row['G Count']))
#         output.write('                     /note="C Count: {}"\n'.format(row['C Count']))
        
#         # Write sequence
#         output.write("ORIGIN\n")
#         sequence = row['Sequence']
#         for i in range(0, len(sequence), 60):
#             chunk = sequence[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
    
#     return output.getvalue()

# def fragment_sequences(sequences, fragment_length=200):
#     """
#     Splits each sequence into fragments of the given length (default 200 bases) and calculates comprehensive analysis.
#     Returns a list of (Fragment Name, Fragment Sequence, Parent Name, Fragment Index, Start Position, Analysis Data).
#     """
#     fragments = []
#     for name, seq in sequences:
#         seq = sanitize_sequence(seq)
#         if not seq or len(seq) < fragment_length:
#             # If sequence is too short, keep as is
#             analysis = analyze_sequence(name, seq)
#             fragments.append((name, seq, name, 1, 1, analysis))
#             continue
            
#         # Split into fragments
#         num_fragments = (len(seq) + fragment_length - 1) // fragment_length
        
#         for i in range(num_fragments):
#             start = i * fragment_length
#             end = min(start + fragment_length, len(seq))
#             frag_seq = seq[start:end]
            
#             # Create fragment name with position range
#             frag_name = f"{name}_F{i+1}[{start+1}-{end}]"
            
#             # Calculate comprehensive analysis for the fragment
#             analysis = analyze_sequence(frag_name, frag_seq)
            
#             fragments.append((frag_name, frag_seq, name, i+1, start+1, analysis))
#     return fragments

# def analyze_sequence_with_fragments(name, sequence, fragment_length=200):
#     """
#     Analyze a sequence and automatically create fragments with comprehensive analysis.
#     Returns original analysis and list of fragment analyses.
#     """
#     # Analyze original sequence
#     original_analysis = analyze_sequence(name, sequence)
    
#     # Create fragments
#     fragments = fragment_sequences([(name, sequence)], fragment_length)
    
#     # Extract fragment analyses
#     fragment_analyses = []
#     for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
#         fragment_analyses.append({
#             "Fragment_Name": frag_name,
#             "Parent_Gene": parent_name,
#             "Fragment_Number": frag_num,
#             "Start_Position": start_pos,
#             "End_Position": start_pos + len(frag_seq) - 1,
#             "Fragment_Length": len(frag_seq),
#             "Fragment_Sequence": frag_seq,
#             **analysis  # Include all analysis data
#         })
    
#     return original_analysis, fragment_analyses

# def export_fragments_to_excel(fragments):
#     # Create a comprehensive DataFrame with all analysis data including Z-DNA
#     data = []
#     for frag_name, frag_seq, parent, idx, start_pos, analysis in fragments:
#         row = {
#             "Gene Name": frag_name,
#             "Sequence": frag_seq,
#             "Parent Name": parent,
#             "Fragment Index": idx,
#             "Start Position": start_pos,
#             "End Position": start_pos + len(frag_seq) - 1,
#             "Fragment Length": len(frag_seq),
#             "Length": analysis.get("Length", len(frag_seq)),
#             "A Count": analysis.get("A Count", 0),
#             "T Count": analysis.get("T Count", 0),
#             "G Count": analysis.get("G Count", 0),
#             "C Count": analysis.get("C Count", 0),
#             "A %": analysis.get("A %", 0),
#             "T %": analysis.get("T %", 0),
#             "G %": analysis.get("G %", 0),
#             "C %": analysis.get("C %", 0),
#             "GC %": analysis.get("GC %", 0),
#             "AT %": analysis.get("AT %", 0),
#             "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
#         }
#         data.append(row)
    
#     df = pd.DataFrame(data)
#     towrite = BytesIO()
#     with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Fragments with Analysis')
#         workbook = writer.book
#         worksheet = writer.sheets['Fragments with Analysis']
        
#         # Auto-adjust columns' width
#         for column in worksheet.columns:
#             max_length = 0
#             column = [cell for cell in column]
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
#         # Add formatting to headers
#         for row in worksheet.iter_rows(min_row=1, max_row=1):
#             for cell in row:
#                 cell.font = cell.font.copy(bold=True)
    
#     towrite.seek(0)
#     return towrite


# def export_fragments_to_fasta(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         # Create a comprehensive description with all analysis data
#         description = f"parent={parent} idx={idx} len={frag_len} "
#         description += f"GC_Content={analysis.get('GC %', 0):.2f}% "
#         description += f"A={analysis.get('A Count', 0)}({analysis.get('A %', 0):.2f}%) "
#         description += f"T={analysis.get('T Count', 0)}({analysis.get('T %', 0):.2f}%) "
#         description += f"G={analysis.get('G Count', 0)}({analysis.get('G %', 0):.2f}%) "
#         description += f"C={analysis.get('C Count', 0)}({analysis.get('C %', 0):.2f}%)"
        
#         output.write(f">{frag_name} {description}\n{frag_seq}\n")
#     return output.getvalue()


# def export_fragments_to_genbank(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         output.write(f"LOCUS       {frag_name} {frag_len} bp    DNA     linear\n")
#         output.write(f"DEFINITION  Fragment of {parent}, index {idx}\n")
#         output.write(f"ACCESSION   {frag_name}\n")
#         output.write(f"VERSION     {frag_name}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write(f"     source          1..{frag_len}\n")
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write(f'                     /note="Fragment of {parent}, index {idx}"\n')
#         output.write(f'                     /note="GC Content: {analysis.get("GC %", 0):.2f}%"\n')
#         output.write(f'                     /note="A Count: {analysis.get("A Count", 0)} ({analysis.get("A %", 0):.2f}%)"\n')
#         output.write(f'                     /note="T Count: {analysis.get("T Count", 0)} ({analysis.get("T %", 0):.2f}%)"\n')
#         output.write(f'                     /note="G Count: {analysis.get("G Count", 0)} ({analysis.get("G %", 0):.2f}%)"\n')
#         output.write(f'                     /note="C Count: {analysis.get("C Count", 0)} ({analysis.get("C %", 0):.2f}%)"\n')
#         output.write(f'                     /note="AT Content: {analysis.get("AT %", 0):.2f}%"\n')
#         output.write("ORIGIN\n")
#         for i in range(0, len(frag_seq), 60):
#             chunk = frag_seq[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
#     return output.getvalue()

# def main():
#     st.set_page_config(page_title="GC Content Calculator", page_icon=None, layout="wide")
    
#     # Load custom CSS
#     load_css()
    
#     # Password protection
#     if 'authenticated' not in st.session_state:
#         st.session_state.authenticated = False
    
#     if not st.session_state.authenticated:
#         # Login screen
#         st.markdown("""
#             <div class="main-header">
#                 <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Authentication Required</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         # Center the login form
#         col1, col2, col3 = st.columns([1, 2, 1])
#         with col2:
#             st.markdown("""
#                 <div style='text-align: center; margin: 2rem 0;'>
#                     <h3>Please enter the password to access the application</h3>
#                 </div>
#             """, unsafe_allow_html=True)
            
#             password = st.text_input("Password", type="password", key="password_input")
            
#             if st.button("Login", key="login_button"):
#                 if password == "TroyDNA2024":
#                     st.session_state.authenticated = True
#                     st.success("Authentication successful! Redirecting...")
#                     st.rerun()
#                 else:
#                     st.error("Incorrect password. Please try again.")
            
#             st.markdown("""
#                 <div style='text-align: center; margin-top: 2rem; color: #bdc3c7;'>
#                     <p>Developed by Shubh Rakesh Nahar | Troy University</p>
#                 </div>
#             """, unsafe_allow_html=True)
        
#         return
    
#     # Main application (only shown after authentication)
#     # Professional header with gradient background
#     st.markdown("""
#         <div class="main-header">
#             <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#         </div>
#     """, unsafe_allow_html=True)

#     sequences = []  # Ensure sequences is always defined

#     # Sidebar with logo, navigation, and about info
#     st.sidebar.image("https://cdn-icons-png.flaticon.com/512/616/616494.png", width=100)
#     st.sidebar.title("Navigation")
    
#     # Add logout button to sidebar
#     if st.sidebar.button("Logout"):
#         st.session_state.authenticated = False
#         st.rerun()
    
#     with st.sidebar.expander("About this Application"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         This professional tool calculates GC content and provides comprehensive analysis for DNA sequences. 
#         Upload your data or enter sequences manually to explore nucleotide composition, GC content, and more. 
        
#         **Developed by Shubh Rakesh Nahar, Troy University.**
#         </div>
#         """, unsafe_allow_html=True)
    
#     # Fun facts about DNA/genes/sequences
#     facts = [
#         "The human genome contains about 3 billion base pairs.",
#         "GC content can affect the stability of DNA.",
#         "Some bacteria have extremely high or low GC content.",
#         "DNA was first isolated by Friedrich Miescher in 1869.",
#         "GC-rich regions are often found near gene promoters.",
#         "Genes are segments of DNA that code for proteins.",
#         "The longest human gene is over 2.4 million base pairs long!",
#         "Mitochondrial DNA is inherited only from your mother.",
#         "Some viruses use RNA instead of DNA as their genetic material.",
#         "The fruit fly has about 15,000 genes, while humans have about 20,000-25,000.",
#         "DNA stands for Deoxyribonucleic Acid.",
#         "The double helix structure of DNA was discovered in 1953.",
#         "Some plants have much more DNA than humans!",
#         "The GC content of a genome can be used to identify species.",
#         "DNA can be extracted from almost any living thing, even ancient fossils!"
#     ]
#     if 'fun_fact_idx' not in st.session_state:
#         st.session_state['fun_fact_idx'] = random.randint(0, len(facts)-1)
#     if st.sidebar.button("Show Another Fun Fact"):
#         prev_idx = st.session_state['fun_fact_idx']
#         new_idx = prev_idx
#         while new_idx == prev_idx:
#             new_idx = random.randint(0, len(facts)-1)
#         st.session_state['fun_fact_idx'] = new_idx
#         st.rerun()
#     st.sidebar.success(f"**Fun Fact:** {facts[st.session_state['fun_fact_idx']]}")

#     # Glossary of Bioinformatics Terms
#     glossary = {
#         "GC Content": "The percentage of bases in a DNA or RNA molecule that are either guanine (G) or cytosine (C).",
#         "FASTA": "A text-based format for representing nucleotide or peptide sequences.",
#         "GenBank": "A rich file format for DNA sequences with annotations, used by NCBI.",
#         "ORF": "Open Reading Frame, a sequence of DNA that could potentially encode a protein.",
#         "Codon": "A sequence of three nucleotides that together form a unit of genetic code.",
#         "Motif": "A short, recurring pattern in DNA that is presumed to have a biological function.",
#         "CpG Island": "A region with a high frequency of CG dinucleotides, often found near gene promoters.",
#         "SNP": "Single Nucleotide Polymorphism, a variation at a single position in a DNA sequence among individuals.",
#         "BLAST": "A tool for comparing an input sequence against a database of sequences.",
#         "Reverse Complement": "The sequence formed by reversing a DNA sequence and replacing each base with its complement.",
#         "Translation": "The process of converting a nucleotide sequence into a protein sequence.",
#         "Phylogenetic Tree": "A branching diagram showing evolutionary relationships among sequences.",
#         "Alignment": "The arrangement of two or more sequences to identify regions of similarity."
#     }
#     with st.sidebar.expander("Glossary of Bioinformatics Terms"):
#         search_term = st.text_input("Search glossary", key="glossary_search")
#         for term, definition in glossary.items():
#             if not search_term or search_term.lower() in term.lower():
#                 st.markdown(f"**{term}:** {definition}")

#     # Tutorial Mode
#     with st.sidebar.expander("Tutorial / Step-by-Step Guide"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         <h4 style='margin-top:0;'>How to Use the GC Content Calculator and Sequence Fragmenter</h4>
#         <p>
#         This application allows you to analyze DNA sequences for GC content and nucleotide composition, as well as to split sequences into defined-length fragments with full analysis. Follow these steps to use the tools effectively:
#         </p>
#         <p>
#         <b>1. Choose Your Input Method:</b><br>
#         You can upload your DNA sequences using a CSV or Excel file, a FASTA file, a GenBank file, or by entering them manually. For file uploads, ensure your data includes both a gene name and a sequence for each entry.
#         </p>
#         <p>
#         <b>2. Analyze Sequences:</b><br>
#         After uploading or entering your sequences, click the "Calculate GC Content & Fragment Analysis" button. The application will automatically split each sequence into 200-base fragments and process your data to display comprehensive analysis including GC content, nucleotide composition statistics, and Z-DNA propensity for both original sequences and each fragment.
#         </p>
#         <p>
#         <b>3. Export Results:</b><br>
#         Once the analysis is complete, you can download your results in Excel, FASTA, GenBank, or as a ZIP archive. Use the export section to select your preferred format.
#         </p>
#         <p>
#         <b>4. Use the Sequence Fragmenter Tool:</b><br>
#         Open the "Sequence Fragmenter (Split Tool)" section in the sidebar. Choose your input method and upload or enter your sequences. Select the desired fragment length. When you click "Split and Export Fragments," the tool will split each sequence into fragments of the specified length and calculate GC content and nucleotide statistics for each fragment. You can then download the results in Excel, FASTA, or GenBank format, each including the full analysis for every fragment.
#         </p>
#         <p>
#         <b>5. Additional Features:</b><br>
#         The sidebar also provides a random sequence generator for practice (which can also accept uploaded files in CSV, FASTA, or GenBank formats), a glossary of bioinformatics terms, and fun facts about DNA and genomics.
#         </p>
#         <p>
#         If you need to log out, use the "Logout" button in the sidebar. For any step, ensure your input data is correctly formatted to avoid errors during analysis.
#         </p>
#         </div>
#         """, unsafe_allow_html=True)

#     # Random Sequence Generator
#     with st.sidebar.expander("Random Sequence Generator"):
#         st.markdown("Generate random DNA or protein sequences for practice or testing, or upload existing sequences.")
        
#         # Input method selection
#         rand_input_method = st.radio("Input method", ["Generate Random", "Upload CSV/Excel", "Upload FASTA", "Upload GenBank"], key="rand_input_method")
        
#         if rand_input_method == "Generate Random":
#             seq_type = st.selectbox("Sequence type", ["DNA", "Protein"], key="rand_seq_type")
#             rand_num = st.number_input("Number of sequences", min_value=1, max_value=1000, value=1, key="rand_num")
#             rand_len = st.number_input("Length of each sequence", min_value=5, max_value=40000, value=50, key="rand_len")
#             if st.button("Generate Random Sequences"):
#                 import string
#                 import random as pyrandom
#                 if seq_type == "DNA":
#                     alphabet = "ATGC"
#                 else:
#                     alphabet = "ACDEFGHIKLMNPQRSTVWY"  # 20 amino acids
                
#                 # Generate sequences in chunks to manage memory
#                 chunk_size = 100  # Process 100 sequences at a time
#                 all_sequences = []
                
#                 for i in range(0, rand_num, chunk_size):
#                     current_chunk = min(chunk_size, rand_num - i)
#                     chunk_sequences = [(f"Random_{i+j+1}", ''.join(pyrandom.choices(alphabet, k=rand_len))) 
#                                      for j in range(current_chunk)]
#                     all_sequences.extend(chunk_sequences)
                    
#                     # Show progress
#                     progress = min(100, int((i + current_chunk) / rand_num * 100))
#                     st.progress(progress)
                
#                 if 'random_sequences' not in st.session_state:
#                     st.session_state['random_sequences'] = []
#                 st.session_state['random_sequences'] = all_sequences
#                 st.toast(f"{rand_num} random {seq_type} sequences of length {rand_len} generated!", icon=None)
        
#         elif rand_input_method == "Upload CSV/Excel":
#             rand_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="rand_file")
#             if rand_file:
#                 try:
#                     rand_file.seek(0)
#                     if rand_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(rand_file)
#                     else:
#                         df = pd.read_csv(rand_file)
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         sequences = list(zip(df["Gene Name"], df["Sequence"]))
#                         if 'random_sequences' not in st.session_state:
#                             st.session_state['random_sequences'] = []
#                         st.session_state['random_sequences'] = sequences
#                         st.success(f"Successfully loaded {len(sequences)} sequences from {rand_file.name}")
#                     else:
#                         st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
#                 except Exception as e:
#                     st.warning(f"Error reading file: {str(e)}")
                    
#         elif rand_input_method == "Upload FASTA":
#             rand_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="rand_fasta")
#             if rand_fasta:
#                 rand_fasta.seek(0)
#                 sequences = process_fasta(rand_fasta, max_sequences=1000)
#                 if sequences:
#                     if 'random_sequences' not in st.session_state:
#                         st.session_state['random_sequences'] = []
#                     st.session_state['random_sequences'] = sequences
#                     st.success(f"Successfully loaded {len(sequences)} sequences from FASTA file")
                    
#         elif rand_input_method == "Upload GenBank":
#             rand_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="rand_gb")
#             if rand_gb:
#                 rand_gb.seek(0)
#                 sequences = process_genbank(rand_gb, max_sequences=1000)
#                 if sequences:
#                     if 'random_sequences' not in st.session_state:
#                         st.session_state['random_sequences'] = []
#                     st.session_state['random_sequences'] = sequences
#                     st.success(f"Successfully loaded {len(sequences)} sequences from GenBank file")
        
#         # Option to clear sequences
#         if st.button("Clear Sequences"):
#             st.session_state['random_sequences'] = []
#             st.toast("Sequences cleared.", icon=None)
        
#         # Download random sequences if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             # First analyze the random sequences in chunks
#             chunk_size = 100  # Process 100 sequences at a time
#             all_results = []
            
#             for i in range(0, len(st.session_state['random_sequences']), chunk_size):
#                 chunk = st.session_state['random_sequences'][i:i + chunk_size]
#                 chunk_results = [analyze_sequence(name, seq) for name, seq in chunk]
#                 all_results.extend(chunk_results)
                
#                 # Show progress
#                 progress = min(100, int((i + len(chunk)) / len(st.session_state['random_sequences']) * 100))
#                 st.progress(progress)
            
#             rand_df = pd.DataFrame(all_results)
            
#             # Excel with formatting
#             excel_buffer = BytesIO()
#             with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#                 rand_df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#                 # Get the workbook and the worksheet
#                 workbook = writer.book
#                 worksheet = writer.sheets['GC Content Analysis']
                
#                 # Auto-adjust columns' width
#                 for column in worksheet.columns:
#                     max_length = 0
#                     column = [cell for cell in column]
#                     for cell in column:
#                         try:
#                             if len(str(cell.value)) > max_length:
#                                 max_length = len(str(cell.value))
#                         except:
#                             pass
#                     adjusted_width = (max_length + 2)
#                     worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
#                 # Add some basic formatting
#                 for row in worksheet.iter_rows(min_row=1, max_row=1):
#                     for cell in row:
#                         cell.font = cell.font.copy(bold=True)
            
#             excel_buffer.seek(0)
#             st.download_button(
#                 label="Download Random Sequences (Excel)",
#                 data=excel_buffer,
#                 file_name="random_sequences.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
            
#             # FASTA
#             fasta_buffer = export_to_fasta(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (FASTA)",
#                 data=fasta_buffer,
#                 file_name="random_sequences.fasta",
#                 mime="text/plain"
#             )
            
#             # GenBank
#             gb_buffer = export_to_genbank(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (GenBank)",
#                 data=gb_buffer,
#                 file_name="random_sequences.gb",
#                 mime="text/plain"
#             )

#     # Add random sequences to main input if present
#     if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#         sequences.extend(st.session_state['random_sequences'])
#         st.session_state['random_sequences'] = [] # Clear after adding

#     # Main content area with professional styling
#     st.markdown("---")
    
#     # Main Analysis Section
#     with st.container():
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Sequence Analysis & GC Content Calculation</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.info("**Professional Tip:** You can input up to 1000 sequences at a time (CSV, Excel, FASTA, GenBank, or manual entry). If you upload more, only the first 1000 will be processed.")

#         input_method = st.radio("Choose input method", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"])

#         if input_method == "Upload CSV/Excel":
#             uploaded_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"])
#             if uploaded_file:
#                 try:
#                     uploaded_file.seek(0) # Ensure pointer is at the start
#                     if uploaded_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(uploaded_file)
#                     else:
#                         df = pd.read_csv(uploaded_file)
                    
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         # Convert sequences to strings and handle any NaN values
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         if len(df) > 1000:
#                             st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                         sequences = list(zip(df["Gene Name"], df["Sequence"]))[:1000]
#                     else:
#                         st.toast("File must contain 'Gene Name' and 'Sequence' columns.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error reading file: {str(e)}", icon=None)

#         elif input_method == "Upload FASTA":
#             fasta_file = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"])
#             if fasta_file:
#                 fasta_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_fasta(fasta_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded FASTA file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing FASTA file: {str(e)}", icon=None)

#         elif input_method == "Upload GenBank":
#             gb_file = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"])
#             if gb_file:
#                 gb_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_genbank(gb_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded GenBank file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing GenBank file: {str(e)}", icon=None)

#         elif input_method == "Manual Entry":
#             num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1)
#             manual_entries = []
#             for i in range(num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"seq_{i}")
#                 if name and seq:
#                     manual_entries.append((name, seq))
#             sequences.extend(manual_entries)

#         # Add random sequences to main input if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             sequences.extend(st.session_state['random_sequences'])
#             st.session_state['random_sequences'] = [] # Clear after adding

#         if sequences:
#             # Fragment length selector
#             st.markdown("### Fragment Analysis Settings")
#             fragment_length = st.selectbox(
#                 "Select fragment length for analysis:",
#                 options=[50, 100, 200, 400],
#                 index=2,  # Default to 200
#                 help="Choose the length of fragments to split your sequences into for detailed analysis"
#             )
            
#             if st.button("Calculate GC Content & Fragment Analysis"):
#                 # Analyze original sequences
#                 results = [analyze_sequence(name, seq) for name, seq in sequences]
#                 result_df = pd.DataFrame(results)

#                 if "Error" in result_df.columns:
#                     st.toast("Some sequences were invalid and skipped.", icon=None)
#                     result_df = result_df.dropna(subset=["Length"])

#                 # Check if there are valid sequences for output
#                 if len(result_df) == 0:
#                      st.warning("No valid sequences were processed.")
#                 else:
#                     # Create fragments for all sequences
#                     all_fragments = []
#                     fragment_analyses = []
                    
#                     for name, seq in sequences:
#                         if seq and len(seq) > 0:
#                             # Create fragments with selected length
#                             fragments = fragment_sequences([(name, seq)], fragment_length)
#                             all_fragments.extend(fragments)
                            
#                             # Extract fragment analyses for display
#                             for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
#                                 fragment_analyses.append({
#                                     "Gene Name": frag_name,
#                                     "Parent Gene": parent_name,
#                                     "Fragment Number": frag_num,
#                                     "Start Position": start_pos,
#                                     "End Position": start_pos + len(frag_seq) - 1,
#                                     "Fragment Length": len(frag_seq),
#                                     "Sequence": frag_seq,
#                                     "Length": analysis.get("Length", len(frag_seq)),
#                                     "A Count": analysis.get("A Count", 0),
#                                     "T Count": analysis.get("T Count", 0),
#                                     "G Count": analysis.get("G Count", 0),
#                                     "C Count": analysis.get("C Count", 0),
#                                     "A %": analysis.get("A %", 0),
#                                     "T %": analysis.get("T %", 0),
#                                     "G %": analysis.get("G %", 0),
#                                     "C %": analysis.get("C %", 0),
#                                     "GC %": analysis.get("GC %", 0),
#                                     "AT %": analysis.get("AT %", 0),
#                                     "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
#                                 })
                    
#                     fragment_df = pd.DataFrame(fragment_analyses)

#                     # Conditional Visualization Display
#                     if len(result_df) > 50:
#                         st.info("Analysis complete. Visualizations and on-page data table are disabled for more than 50 sequences to ensure performance. Please download your full results below.")
#                     else:
#                         st.toast("Analysis complete with automatic fragmentation!", icon=None)
#                         st.markdown("---")
                        
#                         # Results Section
#                         st.markdown("""
#                             <div class="section-container">
#                                 <h2 class="section-header">Results & Visualizations</h2>
#                             </div>
#                         """, unsafe_allow_html=True)
                        
#                         # Original Sequences Analysis
#                         st.markdown("### Original Sequences Analysis")
#                         with st.expander("Show Original Sequences Data", expanded=True):
#                             st.dataframe(result_df, use_container_width=True)
                        
#                         # Fragment Analysis
#                         st.markdown(f"### Fragment Analysis ({fragment_length}-base chunks)")
#                         with st.expander("Show Fragment Data", expanded=True):
#                             st.dataframe(fragment_df, use_container_width=True)
                            
#                         # Fragment Statistics
#                         if len(fragment_df) > 0:
#                             st.markdown("### Fragment Statistics")
#                             col1, col2, col3 = st.columns(3)
                            
#                             with col1:
#                                 st.metric("Total Fragments", len(fragment_df))
#                                 st.metric("Average GC Content", f"{fragment_df['GC %'].mean():.2f}%")
                                
#                             with col2:
#                                 st.metric("Average Z-DNA Propensity", f"{fragment_df['Z-DNA Propensity %'].mean():.2f}%")
#                                 st.metric("High GC Fragments (>60%)", len(fragment_df[fragment_df['GC %'] > 60]))
                                
#                             with col3:
#                                 st.metric("Low GC Fragments (<40%)", len(fragment_df[fragment_df['GC %'] < 40]))
#                                 st.metric("High Z-DNA Fragments (>50%)", len(fragment_df[fragment_df['Z-DNA Propensity %'] > 50]))
#                         with st.expander("Show GC Content Heatmap"):
#                             if len(result_df) > 0:
#                                 create_gc_heatmap(result_df)
#                             else:
#                                 st.info("No data to plot.")
#                         with st.expander("Show Sequence Logo"):
#                             # Use the original 'sequences' list for the logo, not result_df
#                             if len(sequences) > 1 and all(len(seq) > 0 for _, seq in sequences):
#                                 # Filter original sequences based on names in result_df to match analysis results
#                                 valid_gene_names = result_df['Gene Name'].tolist()
#                                 sequences_for_logo = [(name, seq) for name, seq in sequences if name in valid_gene_names]
#                                 if len(sequences_for_logo) > 1:
#                                     create_sequence_logo([seq for _, seq in sequences_for_logo])
#                                 else:
#                                     st.info("Not enough valid sequences to plot a sequence logo.")
#                             else:
#                                 st.info("Not enough data to plot a sequence logo.")

#                         with st.expander("Show Interactive Plots"):
#                             if len(result_df) > 1:
#                                 create_interactive_plots(result_df)
#                             else:
#                                 st.info("Not enough data to plot interactive plots.")

#                     # Export Results section - always shown if there are results
#                     st.markdown("---")
#                     st.markdown("""
#                         <div class="section-container">
#                             <h2 class="section-header">Export Results</h2>
#                         </div>
#                     """, unsafe_allow_html=True)
                    
#                     output_filename = st.text_input("Enter output file name (without extension):", "gc_output_v4", key="export_filename")
                    
#                     # Export buttons in columns for better layout
#                     col1, col2, col3, col4 = st.columns(4)
                    
#                     with col1:
#                         if st.button("Export All Results"):
#                             zip_buffer = export_all_to_zip(result_df, output_filename)
#                             st.toast("All results exported as ZIP!", icon=None)
#                             st.download_button(
#                                 label="Download All Results (ZIP)",
#                                 data=zip_buffer,
#                                 file_name=f"{output_filename}_all_results.zip",
#                                 mime="application/zip"
#                             )
                    
#                     with col2:
#                         excel_buffer = BytesIO()
#                         result_df.to_excel(excel_buffer, index=False, engine='openpyxl')
#                         excel_buffer.seek(0)
#                         st.download_button(
#                             label="Download Original Sequences (Excel)",
#                             data=excel_buffer,
#                             file_name=f"{output_filename}_original.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                    
#                     with col3:
#                         # FASTA export button
#                         # Use result_df for FASTA/GenBank exports to include only analyzed sequences
#                         fasta_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         fasta_buffer = export_to_fasta(fasta_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download Original Sequences (FASTA)",
#                             data=fasta_buffer,
#                             file_name=f"{output_filename}_original.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col4:
#                         # GenBank export button
#                         gb_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         gb_buffer = export_to_genbank(gb_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download Original Sequences (GenBank)",
#                             data=gb_buffer,
#                             file_name=f"{output_filename}_original.gb",
#                             mime="text/plain"
#                         )
                    
#                     # Fragment Export Section
#                     if len(fragment_df) > 0:
#                         st.markdown("### Fragment Export Options")
#                         frag_col1, frag_col2, frag_col3, frag_col4 = st.columns(4)
                        
#                         with frag_col1:
#                             # Export fragments to Excel
#                             frag_excel_buffer = export_fragments_to_excel(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (Excel)",
#                                 data=frag_excel_buffer,
#                                 file_name=f"{output_filename}_fragments_{fragment_length}bp.xlsx",
#                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                             )
                        
#                         with frag_col2:
#                             # Export fragments to FASTA
#                             frag_fasta_buffer = export_fragments_to_fasta(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (FASTA)",
#                                 data=frag_fasta_buffer,
#                                 file_name=f"{output_filename}_fragments_{fragment_length}bp.fasta",
#                                 mime="text/plain"
#                             )
                        
#                         with frag_col3:
#                             # Export fragments to GenBank
#                             frag_gb_buffer = export_fragments_to_genbank(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (GenBank)",
#                                 data=frag_gb_buffer,
#                                 file_name=f"{output_filename}_fragments_{fragment_length}bp.gb",
#                                 mime="text/plain"
#                             )
                        
#                         with frag_col4:
#                             # Combined export with both original and fragments
#                             combined_data = []
                            
#                             # Add original sequences
#                             for _, row in result_df.iterrows():
#                                 combined_data.append({
#                                     "Type": "Original",
#                                     "Gene Name": row["Gene Name"],
#                                     "Sequence": row["Sequence"],
#                                     "Length": row["Length"],
#                                     "GC %": row["GC %"],
#                                     "Z-DNA Propensity %": row.get("Z-DNA Propensity %", 0),
#                                     "A Count": row["A Count"],
#                                     "T Count": row["T Count"],
#                                     "G Count": row["G Count"],
#                                     "C Count": row["C Count"]
#                                 })
                            
#                             # Add fragments
#                             for _, row in fragment_df.iterrows():
#                                 combined_data.append({
#                                     "Type": "Fragment",
#                                     "Gene Name": row["Gene Name"],
#                                     "Parent Gene": row["Parent Gene"],
#                                     "Fragment Number": row["Fragment Number"],
#                                     "Start Position": row["Start Position"],
#                                     "End Position": row["End Position"],
#                                     "Sequence": row["Sequence"],
#                                     "Length": row["Length"],
#                                     "GC %": row["GC %"],
#                                     "Z-DNA Propensity %": row["Z-DNA Propensity %"],
#                                     "A Count": row["A Count"],
#                                     "T Count": row["T Count"],
#                                     "G Count": row["G Count"],
#                                     "C Count": row["C Count"]
#                                 })
                            
#                             combined_df = pd.DataFrame(combined_data)
#                             combined_buffer = BytesIO()
#                             combined_df.to_excel(combined_buffer, index=False, engine='openpyxl')
#                             combined_buffer.seek(0)
#                             st.download_button(
#                                 label="Download Combined (Excel)",
#                                 data=combined_buffer,
#                                 file_name=f"{output_filename}_combined_{fragment_length}bp.xlsx",
#                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                             )

#     # Add Sequence Fragmenter tool to sidebar
#     with st.sidebar.expander("Sequence Fragmenter (Split Tool)"):
#         st.markdown("""
#         <div style='background: rgba(52,152,219,0.1); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3); margin-bottom: 1rem;'>
#         <strong>Split sequences into fragments of a chosen length.</strong><br>
#         Output is compatible with TOOL-1 and available in multiple formats.
#         </div>
#         """, unsafe_allow_html=True)
        
#         frag_input_method = st.radio("Input method for fragmenter", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"], key="frag_input_method")
#         frag_sequences = []
        
#         if frag_input_method == "Upload CSV/Excel":
#             frag_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="frag_file")
#             if frag_file:
#                 try:
#                     frag_file.seek(0)
#                     if frag_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(frag_file)
#                     else:
#                         df = pd.read_csv(frag_file)
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         frag_sequences = list(zip(df["Gene Name"], df["Sequence"]))
#                         st.success(f"Successfully loaded {len(frag_sequences)} sequences from {frag_file.name}")
#                     else:
#                         st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
#                 except Exception as e:
#                     st.warning(f"Error reading file: {str(e)}")
                    
#         elif frag_input_method == "Upload FASTA":
#             frag_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="frag_fasta")
#             if frag_fasta:
#                 frag_fasta.seek(0)
#                 frag_sequences = process_fasta(frag_fasta, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from FASTA file")
                    
#         elif frag_input_method == "Upload GenBank":
#             frag_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="frag_gb")
#             if frag_gb:
#                 frag_gb.seek(0)
#                 frag_sequences = process_genbank(frag_gb, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from GenBank file")
                    
#         elif frag_input_method == "Manual Entry":
#             frag_num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1, key="frag_num")
#             frag_manual_entries = []
#             for i in range(frag_num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"frag_name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"frag_seq_{i}")
#                 if name and seq:
#                     frag_manual_entries.append((name, seq))
#             frag_sequences.extend(frag_manual_entries)
#             if frag_manual_entries:
#                 st.success(f"Successfully entered {len(frag_manual_entries)} sequences")
                
#         # Fragment length selection with better styling
#         st.markdown("---")
#         st.markdown("**Fragment Configuration**")
#         frag_length = st.selectbox("Fragment length (nucleotides)", [50, 100, 200, 400], key="frag_length")
        
#         if frag_sequences:
#             st.info(f"Ready to process {len(frag_sequences)} sequences into {frag_length}-nucleotide fragments")
            
#             if st.button("Split and Export Fragments", key="frag_button"):
#                 with st.spinner("Processing fragments and calculating GC content..."):
#                     fragments = fragment_sequences(frag_sequences, frag_length)
                    
#                 if not fragments:
#                     st.warning("No fragments produced. Check your input sequences and fragment length.")
#                 else:
#                     st.success(f"Successfully produced {len(fragments)} fragments of {frag_length} nucleotides each!")
                    
#                     # Calculate overall statistics
#                     total_gc_content = sum(f[5].get("GC %", 0) for f in fragments)
#                     avg_gc_content = total_gc_content / len(fragments)
                    
#                     # Display fragment statistics with GC content
#                     st.markdown("**Fragment Statistics:**")
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         st.markdown(f"""
#                         - Total fragments: {len(fragments)}
#                         - Fragment length: {frag_length} nucleotides
#                         - Parent sequences: {len(set([f[2] for f in fragments]))}
#                         - Average GC content: {avg_gc_content:.2f}%
#                         """)
                    
#                     with col2:
#                         # Show GC content distribution
#                         gc_values = [f[5].get("GC %", 0) for f in fragments]
#                         min_gc = min(gc_values)
#                         max_gc = max(gc_values)
#                         st.markdown(f"""
#                         - GC content range: {min_gc:.2f}% - {max_gc:.2f}%
#                         - High GC fragments (>60%): {sum(1 for gc in gc_values if gc > 60)}
#                         - Low GC fragments (<40%): {sum(1 for gc in gc_values if gc < 40)}
#                         """)
                    
#                     # Export options in a more organized way
#                     st.markdown("**Download Options**")
                    
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         excel_buffer = export_fragments_to_excel(fragments)
#                         st.download_button(
#                             label="Excel Format (with Analysis)",
#                             data=excel_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                        
#                         fasta_buffer = export_fragments_to_fasta(fragments)
#                         st.download_button(
#                             label="FASTA Format (with Analysis)",
#                             data=fasta_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col2:
#                         gb_buffer = export_fragments_to_genbank(fragments)
#                         st.download_button(
#                             label="GenBank Format (with Analysis)",
#                             data=gb_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.gb",
#                             mime="text/plain"
#                         )
                        
#                         # Show a sample of the analysis data
#                         if len(fragments) > 0:
#                             st.markdown("**Sample Fragment Analysis:**")
#                             sample_frag = fragments[0]
#                             st.markdown(f"""
#                             **{sample_frag[0]}** (from {sample_frag[2]})
#                             - GC Content: {sample_frag[5].get('GC %', 0):.2f}%
#                             - A: {sample_frag[5].get('A Count', 0)} ({sample_frag[5].get('A %', 0):.2f}%)
#                             - T: {sample_frag[5].get('T Count', 0)} ({sample_frag[5].get('T %', 0):.2f}%)
#                             - G: {sample_frag[5].get('G Count', 0)} ({sample_frag[5].get('G %', 0):.2f}%)
#                             - C: {sample_frag[5].get('C Count', 0)} ({sample_frag[5].get('C %', 0):.2f}%)
#                             """)
#         else:
#             st.info("Please upload or enter sequences to begin fragmentation")

#     # Professional footer
#     st.markdown("---")
#     st.markdown("""
#         <div class="footer">
#             <p>Developed by Shubh Rakesh Nahar | Troy University | Professional DNA Sequence Analysis Tool</p>
#         </div>
#     """, unsafe_allow_html=True)

# if __name__ == "__main__":
#     main()




# import pandas as pd
# import streamlit as st
# import matplotlib.pyplot as plt
# from io import BytesIO
# from Bio import SeqIO
# import json
# import plotly.express as px
# import plotly.graph_objects as go
# import seaborn as sns
# import numpy as np
# from matplotlib.patches import Rectangle
# import math
# import zipfile
# import random
# from Bio.Seq import Seq
# from Bio.SeqRecord import SeqRecord
# import io
# import openpyxl

# # Custom CSS for professional styling
# def load_css():
#     st.markdown("""
#     <style>
#     /* Main background and overall styling */
#     .main {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#         color: #ffffff;
#     }
    
#     .stApp {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#     }
    
#     /* Sidebar styling - Overlay mode */
#     .css-1d391kg {
#         background: linear-gradient(180deg, #0f1419 0%, #1a2332 50%, #0f1419 100%);
#         border-right: 2px solid #3498db;
#         position: fixed !important;
#         top: 0;
#         left: 0;
#         height: 100vh;
#         z-index: 1000;
#         width: 21rem;
#         overflow-y: auto;
#         box-shadow: 2px 0 10px rgba(0,0,0,0.3);
#         transition: transform 0.3s ease;
#     }
    
#     /* Main content area - adjust for overlay sidebar */
#     .main .block-container {
#         margin-left: 0 !important;
#         padding-left: 1rem !important;
#         max-width: none !important;
#     }
    
#     /* Sidebar toggle button styling */
#     .css-1rs6os {
#         position: fixed !important;
#         top: 1rem;
#         left: 1rem;
#         z-index: 1001;
#         background: rgba(52, 152, 219, 0.9) !important;
#         border-radius: 50% !important;
#         width: 3rem !important;
#         height: 3rem !important;
#         display: flex !important;
#         align-items: center !important;
#         justify-content: center !important;
#         box-shadow: 0 2px 10px rgba(0,0,0,0.3) !important;
#     }
    
#     /* Header styling */
#     .main-header {
#         background: linear-gradient(90deg, #2c3e50 0%, #34495e 50%, #2c3e50 100%);
#         padding: 2rem;
#         border-radius: 15px;
#         margin-bottom: 2rem;
#         box-shadow: 0 8px 32px rgba(0,0,0,0.4);
#         border: 1px solid #3498db;
#     }
    
#     .main-header h1 {
#         color: #ecf0f1;
#         text-align: center;
#         font-size: 2.5rem;
#         font-weight: 700;
#         margin: 0;
#         text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
#     }
    
#     /* Section styling */
#     .section-container {
#         background: rgba(255, 255, 255, 0.05);
#         padding: 2rem;
#         border-radius: 15px;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         backdrop-filter: blur(10px);
#         box-shadow: 0 8px 32px rgba(0,0,0,0.3);
#     }
    
#     .section-header {
#         color: #3498db;
#         font-size: 1.5rem;
#         font-weight: 600;
#         margin-bottom: 1rem;
#         text-align: center;
#         text-transform: uppercase;
#         letter-spacing: 1px;
#     }
    
#     /* Button styling */
#     .stButton > button {
#         background: linear-gradient(45deg, #3498db, #2980b9);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
#     }
    
#     .stButton > button:hover {
#         background: linear-gradient(45deg, #2980b9, #1f5f8b);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
#     }
    
#     /* Input styling */
#     .stTextInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stTextArea > div > div > textarea {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stSelectbox > div > div > select {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* File uploader styling */
#     .stFileUploader > div {
#         background: rgba(255, 255, 255, 0.05);
#         border: 2px dashed rgba(52, 152, 219, 0.5);
#         border-radius: 15px;
#         padding: 2rem;
#         text-align: center;
#     }
    
#     /* Radio button styling */
#     .stRadio > div {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         padding: 1rem;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Number input styling */
#     .stNumberInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* Expander styling */
#     .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #34495e, #2c3e50);
#         color: #ecf0f1;
#         border-radius: 10px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Success/Info/Warning styling */
#     .stSuccess {
#         background: rgba(46, 204, 113, 0.2);
#         border: 1px solid #2ecc71;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #2ecc71;
#     }
    
#     .stInfo {
#         background: rgba(52, 152, 219, 0.2);
#         border: 1px solid #3498db;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #3498db;
#     }
    
#     .stWarning {
#         background: rgba(241, 196, 15, 0.2);
#         border: 1px solid #f1c40f;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #f1c40f;
#     }
    
#     /* Dataframe styling */
#     .dataframe {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         overflow: hidden;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Download button styling */
#     .stDownloadButton > button {
#         background: linear-gradient(45deg, #27ae60, #229954);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
#     }
    
#     .stDownloadButton > button:hover {
#         background: linear-gradient(45deg, #229954, #1e8449);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
#     }
    
#     /* Progress bar styling */
#     .stProgress > div > div > div {
#         background: linear-gradient(90deg, #3498db, #2980b9);
#         border-radius: 10px;
#     }
    
#     /* Sidebar expander styling */
#     .css-1d391kg .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #2c3e50, #34495e);
#         color: #ecf0f1;
#         border-radius: 8px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Footer styling */
#     .footer {
#         text-align: center;
#         color: #bdc3c7;
#         font-size: 0.9rem;
#         margin-top: 3rem;
#         padding: 1rem;
#         border-top: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Tool card styling */
#     .tool-card {
#         background: rgba(255, 255, 255, 0.08);
#         border-radius: 15px;
#         padding: 1.5rem;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         box-shadow: 0 4px 15px rgba(0,0,0,0.3);
#     }
    
#     /* Fragmenter specific styling */
#     .fragmenter-section {
#         background: linear-gradient(135deg, rgba(52, 152, 219, 0.1) 0%, rgba(41, 128, 185, 0.1) 100%);
#         border: 2px solid rgba(52, 152, 219, 0.5);
#         border-radius: 20px;
#         padding: 2rem;
#         margin: 1rem 0;
#     }
    
#     /* Responsive design */
#     @media (max-width: 768px) {
#         .main-header h1 {
#             font-size: 2rem;
#         }
        
#         .section-container {
#             padding: 1rem;
#         }
        
#         .css-1d391kg {
#             width: 100vw;
#         }
#     }
#     </style>
#     """, unsafe_allow_html=True)

# # --- Instructions ---
# def show_instructions():
#     st.markdown("""
#     ### 📋 Instructions
    
#     1. **Choose Input Method**:
#        - **Upload CSV/Excel**: Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns
#        - **Upload FASTA**: Upload a FASTA file containing DNA sequences
#        - **Upload GenBank**: Upload a GenBank file containing DNA sequences
#        - **Manual Entry**: Enter sequences manually with gene names
    
#     2. **Input Format**:
#        - Only A, T, G, C nucleotides are accepted (case insensitive)
#        - Invalid characters will be automatically removed
#        - Each sequence must have a unique gene name
    
#     3. **Analysis**:
#        - Click "Calculate GC Content" to process sequences
#        - View results in the interactive table
#        - Explore visualizations of GC content and nucleotide composition
    
#     4. **Export Options**:
#        - Export results in multiple formats (Excel, CSV, JSON)
#        - Use "Export All Results" for a complete dataset export
#        - Customize output filename before downloading
#     """)

# # --- Sequence Sanitization and Validation ---
# def sanitize_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return ''.join(filter(lambda x: x.upper() in ['A', 'T', 'G', 'C'], seq.upper()))

# def is_valid_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return all(base in ['A', 'T', 'G', 'C'] for base in seq.upper())

# # --- Z-DNA Prediction Function ---
# def calculate_z_dna_propensity(sequence):
#     """
#     Calculate Z-DNA propensity based on alternating purine-pyrimidine patterns
#     and GC-rich regions that are prone to Z-DNA formation.
#     """
#     if not sequence or len(sequence) < 6:
#         return 0.0
    
#     sequence = sequence.upper()
#     z_score = 0.0
#     total_positions = 0
    
#     # Check for alternating purine-pyrimidine patterns (Z-DNA prone)
#     for i in range(len(sequence) - 5):
#         window = sequence[i:i+6]
#         if len(window) == 6:
#             # Check for alternating pattern
#             alternating = True
#             for j in range(1, 6):
#                 current = window[j] in 'AG'  # Purine
#                 prev = window[j-1] in 'AG'   # Purine
#                 if current == prev:  # Not alternating
#                     alternating = False
#                     break
            
#             if alternating:
#                 # Calculate GC content in this window
#                 gc_count = window.count('G') + window.count('C')
#                 gc_content = gc_count / 6.0
                
#                 # Z-DNA is more likely in GC-rich alternating regions
#                 if gc_content >= 0.5:  # 50% or more GC
#                     z_score += gc_content * 2.0  # Higher weight for GC-rich
#                 else:
#                     z_score += gc_content * 1.0  # Lower weight for AT-rich
                
#                 total_positions += 1
    
#     # Normalize by sequence length and total positions checked
#     if total_positions > 0:
#         return (z_score / total_positions) * 100  # Convert to percentage
#     else:
#         return 0.0

# # --- Enhanced Nucleotide Analysis with Z-DNA ---
# def analyze_sequence(name, seq):
#     # Convert sequence to string if it's not already
#     if not isinstance(seq, str):
#         seq = str(seq)
    
#     seq = sanitize_sequence(seq)
#     length = len(seq)
#     if length == 0:
#         return {"Gene Name": name, "Error": "Invalid sequence or empty after sanitization"}

#     a_count = seq.count('A')
#     t_count = seq.count('T')
#     g_count = seq.count('G')
#     c_count = seq.count('C')
#     gc_count = g_count + c_count
#     at_count = a_count + t_count

#     # Calculate Z-DNA propensity
#     z_dna_propensity = calculate_z_dna_propensity(seq)

#     return {
#         "Gene Name": name,
#         "Sequence": seq,
#         "Length": length,
#         "A Count": a_count,
#         "T Count": t_count,
#         "G Count": g_count,
#         "C Count": c_count,
#         "A %": round((a_count / length) * 100, 2),
#         "T %": round((t_count / length) * 100, 2),
#         "G %": round((g_count / length) * 100, 2),
#         "C %": round((c_count / length) * 100, 2),
#         "GC %": round((gc_count / length) * 100, 2),
#         "AT %": round((at_count / length) * 100, 2),
#         "Z-DNA Propensity %": round(z_dna_propensity, 2),
#     }

# def process_fasta(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for FASTA
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "fasta")
#         count = 0
#         for record in records:
#             if count >= max_sequences:
#                 break
#             # Convert Seq object to string
#             sequences.append((str(record.id), str(record.seq)))
#             count += 1
#         if count == 0:
#             st.warning("No sequences found in the uploaded FASTA file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the FASTA file.")
#     except Exception as e:
#         st.error(f"Error parsing FASTA file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def process_genbank(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for GenBank
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "genbank")
#         count = 0
#         for record in records:
#              if count >= max_sequences:
#                  break
#              # Ensure molecule_type is set for GenBank export compatibility later
#              # For uploaded GenBank, try to get molecule_type from annotations, default to DNA
#              mol_type = record.annotations.get("molecule_type", "DNA")
#              sequences.append((str(record.id), str(record.seq)))
#              count += 1
#         if count == 0:
#              st.warning("No sequences found in the uploaded GenBank file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the GenBank file.")

#     except Exception as e:
#         st.error(f"Error parsing GenBank file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def display_visuals(df):
#     st.subheader("📊 GC Content Distribution")
#     fig, ax = plt.subplots()
#     df.plot.bar(x='Gene Name', y='GC %', ax=ax, color='purple', legend=False)
#     plt.ylabel('GC %')
#     plt.xticks(rotation=45, ha='right')
#     st.pyplot(fig)

#     st.subheader("🧬 Nucleotide Composition Per Sequence")
#     for _, row in df.iterrows():
#         fig, ax = plt.subplots()
#         ax.pie(
#             [row['A %'], row['T %'], row['G %'], row['C %']],
#             labels=['A %', 'T %', 'G %', 'C %'],
#             autopct='%1.1f%%',
#             startangle=90
#         )
#         ax.set_title(f"{row['Gene Name']} - Base % Composition")
#         st.pyplot(fig)

# def export_data(df, format_type):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     if format_type == "Excel":
#         towrite = BytesIO()
#         with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         towrite.seek(0)
#         return towrite, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
#     elif format_type == "CSV":
#         towrite = BytesIO()
#         # Add signature as a comment in the first line
#         towrite.write(f"# {signature}\n".encode())
#         df.to_csv(towrite, index=False)
#         towrite.seek(0)
#         return towrite, "text/csv", "csv"
#     else:  # JSON
#         towrite = BytesIO()
#         # Add signature to the JSON metadata
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         towrite.write(json_str.encode())
#         towrite.seek(0)
#         return towrite, "application/json", "json"

# def create_gc_heatmap(df):
#     st.subheader("🌡️ GC Content Heatmap")
    
#     # Create a matrix of GC content for each position
#     sequences = df['Sequence'].tolist()
#     max_len = max(len(seq) for seq in sequences)
    
#     # Initialize matrix
#     gc_matrix = np.zeros((len(sequences), max_len))
    
#     # Fill matrix with GC content for each position
#     for i, seq in enumerate(sequences):
#         for j in range(len(seq)):
#             if j < len(seq):
#                 window = seq[max(0, j-10):min(len(seq), j+11)]
#                 gc_count = window.count('G') + window.count('C')
#                 gc_matrix[i, j] = (gc_count / len(window)) * 100
    
#     # Create heatmap using plotly
#     fig = go.Figure(data=go.Heatmap(
#         z=gc_matrix,
#         x=list(range(max_len)),
#         y=df['Gene Name'].tolist(),
#         colorscale='Viridis',
#         colorbar=dict(title='GC %')
#     ))
    
#     fig.update_layout(
#         title='GC Content Distribution Across Sequences',
#         xaxis_title='Position',
#         yaxis_title='Gene Name',
#         height=400 + (len(sequences) * 20)  # Adjust height based on number of sequences
#     )
    
#     st.plotly_chart(fig, use_container_width=True)

# def calculate_information_content(freq):
#     """Calculate information content in bits."""
#     if freq == 0:
#         return 0
#     return freq * math.log2(freq * 4)  # 4 for number of nucleotides

# def create_sequence_logo(sequences):
#     st.subheader("🎨 Sequence Logo")
    
#     # Calculate position frequency matrix
#     max_len = max(len(seq) for seq in sequences)
#     pfm = np.zeros((4, max_len))  # 4 nucleotides
    
#     for seq in sequences:
#         for i, base in enumerate(seq):
#             if i < max_len:
#                 if base == 'A':
#                     pfm[0, i] += 1
#                 elif base == 'T':
#                     pfm[1, i] += 1
#                 elif base == 'G':
#                     pfm[2, i] += 1
#                 elif base == 'C':
#                     pfm[3, i] += 1
    
#     # Normalize
#     pfm = pfm / len(sequences)
    
#     # Calculate information content
#     ic = np.zeros(max_len)
#     for i in range(max_len):
#         ic[i] = sum(calculate_information_content(freq) for freq in pfm[:, i])
    
#     # Create the plot
#     fig, ax = plt.subplots(figsize=(12, 4))
    
#     # Colors for nucleotides
#     colors = {'A': '#2ecc71', 'T': '#e74c3c', 'G': '#f1c40f', 'C': '#3498db'}
#     bases = ['A', 'T', 'G', 'C']
    
#     # Plot each position
#     for i in range(max_len):
#         # Sort frequencies for this position
#         freqs = pfm[:, i]
#         sorted_indices = np.argsort(freqs)
        
#         # Plot each base
#         y_bottom = 0
#         for idx in sorted_indices:
#             if freqs[idx] > 0:
#                 height = freqs[idx] * ic[i]
#                 rect = Rectangle((i, y_bottom), 1, height,
#                                facecolor=colors[bases[idx]],
#                                edgecolor='black',
#                                linewidth=0.5)
#                 ax.add_patch(rect)
#                 y_bottom += height
    
#     # Customize the plot
#     ax.set_xlim(0, max_len)
#     ax.set_ylim(0, max(ic) * 1.1)
#     ax.set_xlabel('Position')
#     ax.set_ylabel('Bits')
#     ax.set_title('Sequence Logo')
    
#     # Add legend
#     legend_elements = [Rectangle((0, 0), 1, 1, facecolor=color, edgecolor='black')
#                       for color in colors.values()]
#     ax.legend(legend_elements, bases, loc='upper right')
    
#     # Remove spines
#     for spine in ['top', 'right']:
#         ax.spines[spine].set_visible(False)
    
#     # Show the plot
#     st.pyplot(fig)
#     plt.close(fig)

# def create_interactive_plots(df):
#     st.subheader("📊 Interactive Plots")
    
#     # GC Content Distribution
#     fig_gc = px.box(df, y='GC %', title='GC Content Distribution',
#                     hover_data=['Gene Name', 'Length', 'GC %'])
#     fig_gc.update_traces(marker_color='purple')
#     st.plotly_chart(fig_gc, use_container_width=True)
    
#     # Nucleotide Composition
#     fig_comp = px.bar(df, 
#                      x='Gene Name',
#                      y=['A %', 'T %', 'G %', 'C %'],
#                      title='Nucleotide Composition by Gene',
#                      barmode='group',
#                      hover_data=['Length', 'GC %'])
#     st.plotly_chart(fig_comp, use_container_width=True)
    
#     # Length vs GC Content
#     fig_scatter = px.scatter(df,
#                            x='Length',
#                            y='GC %',
#                            color='GC %',
#                            hover_data=['Gene Name', 'Length', 'GC %'],
#                            title='Sequence Length vs GC Content')
#     st.plotly_chart(fig_scatter, use_container_width=True)

# def export_all_to_zip(df, output_filename):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     zip_buffer = BytesIO()
#     with zipfile.ZipFile(zip_buffer, "w") as zip_file:
#         # Excel with formatting
#         excel_buffer = BytesIO()
#         with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         excel_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.xlsx", excel_buffer.read())
        
#         # CSV
#         csv_buffer = BytesIO()
#         # Add signature as a comment in the first line
#         csv_buffer.write(f"# {signature}\n".encode())
#         df.to_csv(csv_buffer, index=False)
#         csv_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.csv", csv_buffer.read())
        
#         # JSON
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         json_buffer = BytesIO()
#         json_buffer.write(json_str.encode())
#         json_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.json", json_buffer.read())
    
#     zip_buffer.seek(0)
#     return zip_buffer

# def export_to_fasta(sequences, result_df):
#     """Export sequences to FASTA format with analysis results in description"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     # Add signature as a comment at the top
#     output.write(f"# {signature}\n")
#     for idx, row in result_df.iterrows():
#         # Create a description with analysis results
#         description = f"GC Content: {row['GC %']:.2f}%, Length: {row['Length']}, "
#         description += f"A: {row['A Count']}, T: {row['T Count']}, G: {row['G Count']}, C: {row['C Count']}"
        
#         # Write sequence with description
#         output.write(f">{row['Gene Name']} {description}\n")
#         output.write(f"{row['Sequence']}\n")
    
#     # Add signature at the end
#     output.write(f"\n# {signature}\n")
#     return output.getvalue()

# def export_to_genbank(sequences, result_df):
#     """Export sequences to GenBank format with analysis results in features"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     for idx, row in result_df.iterrows():
#         # Write GenBank header
#         output.write(f"LOCUS       {row['Gene Name']}              {row['Length']} bp    DNA     linear\n")
#         output.write(f"DEFINITION  {row['Gene Name']}\n")
#         output.write(f"ACCESSION   {row['Gene Name']}\n")
#         output.write(f"VERSION     {row['Gene Name']}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
        
#         # Add signature in the features section
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write("     source          1..{}\n".format(row['Length']))
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write('                     /note="{}"\n'.format(signature))
#         output.write('     misc_feature    1..{}\n'.format(row['Length']))
#         output.write('                     /note="GC Content: {:.2f}%"\n'.format(row['GC %']))
#         output.write('                     /note="A Count: {}"\n'.format(row['A Count']))
#         output.write('                     /note="T Count: {}"\n'.format(row['T Count']))
#         output.write('                     /note="G Count: {}"\n'.format(row['G Count']))
#         output.write('                     /note="C Count: {}"\n'.format(row['C Count']))
        
#         # Write sequence
#         output.write("ORIGIN\n")
#         sequence = row['Sequence']
#         for i in range(0, len(sequence), 60):
#             chunk = sequence[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
    
#     return output.getvalue()

# def fragment_sequences(sequences, fragment_length=200):
#     """
#     Splits each sequence into fragments of the given length (default 200 bases) and calculates comprehensive analysis.
#     Returns a list of (Fragment Name, Fragment Sequence, Parent Name, Fragment Index, Start Position, Analysis Data).
#     """
#     fragments = []
#     for name, seq in sequences:
#         seq = sanitize_sequence(seq)
#         if not seq or len(seq) < fragment_length:
#             # If sequence is too short, keep as is
#             analysis = analyze_sequence(name, seq)
#             fragments.append((name, seq, name, 1, 1, analysis))
#             continue
            
#         # Split into fragments
#         num_fragments = (len(seq) + fragment_length - 1) // fragment_length
        
#         for i in range(num_fragments):
#             start = i * fragment_length
#             end = min(start + fragment_length, len(seq))
#             frag_seq = seq[start:end]
            
#             # Create fragment name with position range
#             frag_name = f"{name}_F{i+1}[{start+1}-{end}]"
            
#             # Calculate comprehensive analysis for the fragment
#             analysis = analyze_sequence(frag_name, frag_seq)
            
#             fragments.append((frag_name, frag_seq, name, i+1, start+1, analysis))
#     return fragments

# def analyze_sequence_with_fragments(name, sequence):
#     """
#     Analyze a sequence and automatically create 200-base fragments with comprehensive analysis.
#     Returns original analysis and list of fragment analyses.
#     """
#     # Analyze original sequence
#     original_analysis = analyze_sequence(name, sequence)
    
#     # Create fragments
#     fragments = fragment_sequences([(name, sequence)], 200)
    
#     # Extract fragment analyses
#     fragment_analyses = []
#     for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
#         fragment_analyses.append({
#             "Fragment_Name": frag_name,
#             "Parent_Gene": parent_name,
#             "Fragment_Number": frag_num,
#             "Start_Position": start_pos,
#             "End_Position": start_pos + len(frag_seq) - 1,
#             "Fragment_Length": len(frag_seq),
#             "Fragment_Sequence": frag_seq,
#             **analysis  # Include all analysis data
#         })
    
#     return original_analysis, fragment_analyses

# def export_fragments_to_excel(fragments):
#     # Create a comprehensive DataFrame with all analysis data including Z-DNA
#     data = []
#     for frag_name, frag_seq, parent, idx, start_pos, analysis in fragments:
#         row = {
#             "Gene Name": frag_name,
#             "Sequence": frag_seq,
#             "Parent Name": parent,
#             "Fragment Index": idx,
#             "Start Position": start_pos,
#             "End Position": start_pos + len(frag_seq) - 1,
#             "Fragment Length": len(frag_seq),
#             "Length": analysis.get("Length", len(frag_seq)),
#             "A Count": analysis.get("A Count", 0),
#             "T Count": analysis.get("T Count", 0),
#             "G Count": analysis.get("G Count", 0),
#             "C Count": analysis.get("C Count", 0),
#             "A %": analysis.get("A %", 0),
#             "T %": analysis.get("T %", 0),
#             "G %": analysis.get("G %", 0),
#             "C %": analysis.get("C %", 0),
#             "GC %": analysis.get("GC %", 0),
#             "AT %": analysis.get("AT %", 0),
#             "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
#         }
#         data.append(row)
    
#     df = pd.DataFrame(data)
#     towrite = BytesIO()
#     with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Fragments with Analysis')
#         workbook = writer.book
#         worksheet = writer.sheets['Fragments with Analysis']
        
#         # Auto-adjust columns' width
#         for column in worksheet.columns:
#             max_length = 0
#             column = [cell for cell in column]
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
#         # Add formatting to headers
#         for row in worksheet.iter_rows(min_row=1, max_row=1):
#             for cell in row:
#                 cell.font = cell.font.copy(bold=True)
    
#     towrite.seek(0)
#     return towrite


# def export_fragments_to_fasta(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         # Create a comprehensive description with all analysis data
#         description = f"parent={parent} idx={idx} len={frag_len} "
#         description += f"GC_Content={analysis.get('GC %', 0):.2f}% "
#         description += f"A={analysis.get('A Count', 0)}({analysis.get('A %', 0):.2f}%) "
#         description += f"T={analysis.get('T Count', 0)}({analysis.get('T %', 0):.2f}%) "
#         description += f"G={analysis.get('G Count', 0)}({analysis.get('G %', 0):.2f}%) "
#         description += f"C={analysis.get('C Count', 0)}({analysis.get('C %', 0):.2f}%)"
        
#         output.write(f">{frag_name} {description}\n{frag_seq}\n")
#     return output.getvalue()


# def export_fragments_to_genbank(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         output.write(f"LOCUS       {frag_name} {frag_len} bp    DNA     linear\n")
#         output.write(f"DEFINITION  Fragment of {parent}, index {idx}\n")
#         output.write(f"ACCESSION   {frag_name}\n")
#         output.write(f"VERSION     {frag_name}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write(f"     source          1..{frag_len}\n")
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write(f'                     /note="Fragment of {parent}, index {idx}"\n')
#         output.write(f'                     /note="GC Content: {analysis.get("GC %", 0):.2f}%"\n')
#         output.write(f'                     /note="A Count: {analysis.get("A Count", 0)} ({analysis.get("A %", 0):.2f}%)"\n')
#         output.write(f'                     /note="T Count: {analysis.get("T Count", 0)} ({analysis.get("T %", 0):.2f}%)"\n')
#         output.write(f'                     /note="G Count: {analysis.get("G Count", 0)} ({analysis.get("G %", 0):.2f}%)"\n')
#         output.write(f'                     /note="C Count: {analysis.get("C Count", 0)} ({analysis.get("C %", 0):.2f}%)"\n')
#         output.write(f'                     /note="AT Content: {analysis.get("AT %", 0):.2f}%"\n')
#         output.write("ORIGIN\n")
#         for i in range(0, len(frag_seq), 60):
#             chunk = frag_seq[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
#     return output.getvalue()

# def main():
#     st.set_page_config(page_title="GC Content Calculator", page_icon=None, layout="wide")
    
#     # Load custom CSS
#     load_css()
    
#     # Password protection
#     if 'authenticated' not in st.session_state:
#         st.session_state.authenticated = False
    
#     if not st.session_state.authenticated:
#         # Login screen
#         st.markdown("""
#             <div class="main-header">
#                 <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Authentication Required</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         # Center the login form
#         col1, col2, col3 = st.columns([1, 2, 1])
#         with col2:
#             st.markdown("""
#                 <div style='text-align: center; margin: 2rem 0;'>
#                     <h3>Please enter the password to access the application</h3>
#                 </div>
#             """, unsafe_allow_html=True)
            
#             password = st.text_input("Password", type="password", key="password_input")
            
#             if st.button("Login", key="login_button"):
#                 if password == "TroyDNA2024":
#                     st.session_state.authenticated = True
#                     st.success("Authentication successful! Redirecting...")
#                     st.rerun()
#                 else:
#                     st.error("Incorrect password. Please try again.")
            
#             st.markdown("""
#                 <div style='text-align: center; margin-top: 2rem; color: #bdc3c7;'>
#                     <p>Developed by Shubh Rakesh Nahar | Troy University</p>
#                 </div>
#             """, unsafe_allow_html=True)
        
#         return
    
#     # Main application (only shown after authentication)
#     # Professional header with gradient background
#     st.markdown("""
#         <div class="main-header">
#             <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#         </div>
#     """, unsafe_allow_html=True)

#     sequences = []  # Ensure sequences is always defined

#     # Sidebar with logo, navigation, and about info
#     st.sidebar.image("https://cdn-icons-png.flaticon.com/512/616/616494.png", width=100)
#     st.sidebar.title("Navigation")
    
#     # Add logout button to sidebar
#     if st.sidebar.button("Logout"):
#         st.session_state.authenticated = False
#         st.rerun()
    
#     with st.sidebar.expander("About this Application"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         This professional tool calculates GC content and provides comprehensive analysis for DNA sequences. 
#         Upload your data or enter sequences manually to explore nucleotide composition, GC content, and more. 
        
#         **Developed by Shubh Rakesh Nahar, Troy University.**
#         </div>
#         """, unsafe_allow_html=True)
    
#     # Fun facts about DNA/genes/sequences
#     facts = [
#         "The human genome contains about 3 billion base pairs.",
#         "GC content can affect the stability of DNA.",
#         "Some bacteria have extremely high or low GC content.",
#         "DNA was first isolated by Friedrich Miescher in 1869.",
#         "GC-rich regions are often found near gene promoters.",
#         "Genes are segments of DNA that code for proteins.",
#         "The longest human gene is over 2.4 million base pairs long!",
#         "Mitochondrial DNA is inherited only from your mother.",
#         "Some viruses use RNA instead of DNA as their genetic material.",
#         "The fruit fly has about 15,000 genes, while humans have about 20,000-25,000.",
#         "DNA stands for Deoxyribonucleic Acid.",
#         "The double helix structure of DNA was discovered in 1953.",
#         "Some plants have much more DNA than humans!",
#         "The GC content of a genome can be used to identify species.",
#         "DNA can be extracted from almost any living thing, even ancient fossils!"
#     ]
#     if 'fun_fact_idx' not in st.session_state:
#         st.session_state['fun_fact_idx'] = random.randint(0, len(facts)-1)
#     if st.sidebar.button("Show Another Fun Fact"):
#         prev_idx = st.session_state['fun_fact_idx']
#         new_idx = prev_idx
#         while new_idx == prev_idx:
#             new_idx = random.randint(0, len(facts)-1)
#         st.session_state['fun_fact_idx'] = new_idx
#         st.rerun()
#     st.sidebar.success(f"**Fun Fact:** {facts[st.session_state['fun_fact_idx']]}")

#     # Glossary of Bioinformatics Terms
#     glossary = {
#         "GC Content": "The percentage of bases in a DNA or RNA molecule that are either guanine (G) or cytosine (C).",
#         "FASTA": "A text-based format for representing nucleotide or peptide sequences.",
#         "GenBank": "A rich file format for DNA sequences with annotations, used by NCBI.",
#         "ORF": "Open Reading Frame, a sequence of DNA that could potentially encode a protein.",
#         "Codon": "A sequence of three nucleotides that together form a unit of genetic code.",
#         "Motif": "A short, recurring pattern in DNA that is presumed to have a biological function.",
#         "CpG Island": "A region with a high frequency of CG dinucleotides, often found near gene promoters.",
#         "SNP": "Single Nucleotide Polymorphism, a variation at a single position in a DNA sequence among individuals.",
#         "BLAST": "A tool for comparing an input sequence against a database of sequences.",
#         "Reverse Complement": "The sequence formed by reversing a DNA sequence and replacing each base with its complement.",
#         "Translation": "The process of converting a nucleotide sequence into a protein sequence.",
#         "Phylogenetic Tree": "A branching diagram showing evolutionary relationships among sequences.",
#         "Alignment": "The arrangement of two or more sequences to identify regions of similarity."
#     }
#     with st.sidebar.expander("Glossary of Bioinformatics Terms"):
#         search_term = st.text_input("Search glossary", key="glossary_search")
#         for term, definition in glossary.items():
#             if not search_term or search_term.lower() in term.lower():
#                 st.markdown(f"**{term}:** {definition}")

#     # Tutorial Mode
#     with st.sidebar.expander("Tutorial / Step-by-Step Guide"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         <h4 style='margin-top:0;'>How to Use the GC Content Calculator and Sequence Fragmenter</h4>
#         <p>
#         This application allows you to analyze DNA sequences for GC content and nucleotide composition, as well as to split sequences into defined-length fragments with full analysis. Follow these steps to use the tools effectively:
#         </p>
#         <p>
#         <b>1. Choose Your Input Method:</b><br>
#         You can upload your DNA sequences using a CSV or Excel file, a FASTA file, a GenBank file, or by entering them manually. For file uploads, ensure your data includes both a gene name and a sequence for each entry.
#         </p>
#         <p>
#         <b>2. Analyze Sequences:</b><br>
#         After uploading or entering your sequences, click the "Calculate GC Content & Fragment Analysis" button. The application will automatically split each sequence into 200-base fragments and process your data to display comprehensive analysis including GC content, nucleotide composition statistics, and Z-DNA propensity for both original sequences and each fragment.
#         </p>
#         <p>
#         <b>3. Export Results:</b><br>
#         Once the analysis is complete, you can download your results in Excel, FASTA, GenBank, or as a ZIP archive. Use the export section to select your preferred format.
#         </p>
#         <p>
#         <b>4. Use the Sequence Fragmenter Tool:</b><br>
#         Open the "Sequence Fragmenter (Split Tool)" section in the sidebar. Choose your input method and upload or enter your sequences. Select the desired fragment length. When you click "Split and Export Fragments," the tool will split each sequence into fragments of the specified length and calculate GC content and nucleotide statistics for each fragment. You can then download the results in Excel, FASTA, or GenBank format, each including the full analysis for every fragment.
#         </p>
#         <p>
#         <b>5. Additional Features:</b><br>
#         The sidebar also provides a random sequence generator for practice (which can also accept uploaded files in CSV, FASTA, or GenBank formats), a glossary of bioinformatics terms, and fun facts about DNA and genomics.
#         </p>
#         <p>
#         If you need to log out, use the "Logout" button in the sidebar. For any step, ensure your input data is correctly formatted to avoid errors during analysis.
#         </p>
#         </div>
#         """, unsafe_allow_html=True)

#     # Random Sequence Generator
#     with st.sidebar.expander("Random Sequence Generator"):
#         st.markdown("Generate random DNA or protein sequences for practice or testing, or upload existing sequences.")
        
#         # Input method selection
#         rand_input_method = st.radio("Input method", ["Generate Random", "Upload CSV/Excel", "Upload FASTA", "Upload GenBank"], key="rand_input_method")
        
#         if rand_input_method == "Generate Random":
#             seq_type = st.selectbox("Sequence type", ["DNA", "Protein"], key="rand_seq_type")
#             rand_num = st.number_input("Number of sequences", min_value=1, max_value=1000, value=1, key="rand_num")
#             rand_len = st.number_input("Length of each sequence", min_value=5, max_value=40000, value=50, key="rand_len")
#             if st.button("Generate Random Sequences"):
#                 import string
#                 import random as pyrandom
#                 if seq_type == "DNA":
#                     alphabet = "ATGC"
#                 else:
#                     alphabet = "ACDEFGHIKLMNPQRSTVWY"  # 20 amino acids
                
#                 # Generate sequences in chunks to manage memory
#                 chunk_size = 100  # Process 100 sequences at a time
#                 all_sequences = []
                
#                 for i in range(0, rand_num, chunk_size):
#                     current_chunk = min(chunk_size, rand_num - i)
#                     chunk_sequences = [(f"Random_{i+j+1}", ''.join(pyrandom.choices(alphabet, k=rand_len))) 
#                                      for j in range(current_chunk)]
#                     all_sequences.extend(chunk_sequences)
                    
#                     # Show progress
#                     progress = min(100, int((i + current_chunk) / rand_num * 100))
#                     st.progress(progress)
                
#                 if 'random_sequences' not in st.session_state:
#                     st.session_state['random_sequences'] = []
#                 st.session_state['random_sequences'] = all_sequences
#                 st.toast(f"{rand_num} random {seq_type} sequences of length {rand_len} generated!", icon=None)
        
#         elif rand_input_method == "Upload CSV/Excel":
#             rand_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="rand_file")
#             if rand_file:
#                 try:
#                     rand_file.seek(0)
#                     if rand_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(rand_file)
#                     else:
#                         df = pd.read_csv(rand_file)
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         sequences = list(zip(df["Gene Name"], df["Sequence"]))
#                         if 'random_sequences' not in st.session_state:
#                             st.session_state['random_sequences'] = []
#                         st.session_state['random_sequences'] = sequences
#                         st.success(f"Successfully loaded {len(sequences)} sequences from {rand_file.name}")
#                     else:
#                         st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
#                 except Exception as e:
#                     st.warning(f"Error reading file: {str(e)}")
                    
#         elif rand_input_method == "Upload FASTA":
#             rand_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="rand_fasta")
#             if rand_fasta:
#                 rand_fasta.seek(0)
#                 sequences = process_fasta(rand_fasta, max_sequences=1000)
#                 if sequences:
#                     if 'random_sequences' not in st.session_state:
#                         st.session_state['random_sequences'] = []
#                     st.session_state['random_sequences'] = sequences
#                     st.success(f"Successfully loaded {len(sequences)} sequences from FASTA file")
                    
#         elif rand_input_method == "Upload GenBank":
#             rand_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="rand_gb")
#             if rand_gb:
#                 rand_gb.seek(0)
#                 sequences = process_genbank(rand_gb, max_sequences=1000)
#                 if sequences:
#                     if 'random_sequences' not in st.session_state:
#                         st.session_state['random_sequences'] = []
#                     st.session_state['random_sequences'] = sequences
#                     st.success(f"Successfully loaded {len(sequences)} sequences from GenBank file")
        
#         # Option to clear sequences
#         if st.button("Clear Sequences"):
#             st.session_state['random_sequences'] = []
#             st.toast("Sequences cleared.", icon=None)
        
#         # Download random sequences if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             # First analyze the random sequences in chunks
#             chunk_size = 100  # Process 100 sequences at a time
#             all_results = []
            
#             for i in range(0, len(st.session_state['random_sequences']), chunk_size):
#                 chunk = st.session_state['random_sequences'][i:i + chunk_size]
#                 chunk_results = [analyze_sequence(name, seq) for name, seq in chunk]
#                 all_results.extend(chunk_results)
                
#                 # Show progress
#                 progress = min(100, int((i + len(chunk)) / len(st.session_state['random_sequences']) * 100))
#                 st.progress(progress)
            
#             rand_df = pd.DataFrame(all_results)
            
#             # Excel with formatting
#             excel_buffer = BytesIO()
#             with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#                 rand_df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#                 # Get the workbook and the worksheet
#                 workbook = writer.book
#                 worksheet = writer.sheets['GC Content Analysis']
                
#                 # Auto-adjust columns' width
#                 for column in worksheet.columns:
#                     max_length = 0
#                     column = [cell for cell in column]
#                     for cell in column:
#                         try:
#                             if len(str(cell.value)) > max_length:
#                                 max_length = len(str(cell.value))
#                         except:
#                             pass
#                     adjusted_width = (max_length + 2)
#                     worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
#                 # Add some basic formatting
#                 for row in worksheet.iter_rows(min_row=1, max_row=1):
#                     for cell in row:
#                         cell.font = cell.font.copy(bold=True)
            
#             excel_buffer.seek(0)
#             st.download_button(
#                 label="Download Random Sequences (Excel)",
#                 data=excel_buffer,
#                 file_name="random_sequences.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
            
#             # FASTA
#             fasta_buffer = export_to_fasta(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (FASTA)",
#                 data=fasta_buffer,
#                 file_name="random_sequences.fasta",
#                 mime="text/plain"
#             )
            
#             # GenBank
#             gb_buffer = export_to_genbank(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (GenBank)",
#                 data=gb_buffer,
#                 file_name="random_sequences.gb",
#                 mime="text/plain"
#             )

#     # Add random sequences to main input if present
#     if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#         sequences.extend(st.session_state['random_sequences'])
#         st.session_state['random_sequences'] = [] # Clear after adding

#     # Main content area with professional styling
#     st.markdown("---")
    
#     # Main Analysis Section
#     with st.container():
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Sequence Analysis & GC Content Calculation</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.info("**Professional Tip:** You can input up to 1000 sequences at a time (CSV, Excel, FASTA, GenBank, or manual entry). If you upload more, only the first 1000 will be processed.")

#         input_method = st.radio("Choose input method", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"])

#         if input_method == "Upload CSV/Excel":
#             uploaded_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"])
#             if uploaded_file:
#                 try:
#                     uploaded_file.seek(0) # Ensure pointer is at the start
#                     if uploaded_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(uploaded_file)
#                     else:
#                         df = pd.read_csv(uploaded_file)
                    
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         # Convert sequences to strings and handle any NaN values
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         if len(df) > 1000:
#                             st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                         sequences = list(zip(df["Gene Name"], df["Sequence"]))[:1000]
#                     else:
#                         st.toast("File must contain 'Gene Name' and 'Sequence' columns.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error reading file: {str(e)}", icon=None)

#         elif input_method == "Upload FASTA":
#             fasta_file = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"])
#             if fasta_file:
#                 fasta_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_fasta(fasta_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded FASTA file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing FASTA file: {str(e)}", icon=None)

#         elif input_method == "Upload GenBank":
#             gb_file = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"])
#             if gb_file:
#                 gb_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_genbank(gb_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded GenBank file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing GenBank file: {str(e)}", icon=None)

#         elif input_method == "Manual Entry":
#             num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1)
#             manual_entries = []
#             for i in range(num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"seq_{i}")
#                 if name and seq:
#                     manual_entries.append((name, seq))
#             sequences.extend(manual_entries)

#         # Add random sequences to main input if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             sequences.extend(st.session_state['random_sequences'])
#             st.session_state['random_sequences'] = [] # Clear after adding

#         if sequences:
#             if st.button("Calculate GC Content & Fragment Analysis"):
#                 # Analyze original sequences
#                 results = [analyze_sequence(name, seq) for name, seq in sequences]
#                 result_df = pd.DataFrame(results)

#                 if "Error" in result_df.columns:
#                     st.toast("Some sequences were invalid and skipped.", icon=None)
#                     result_df = result_df.dropna(subset=["Length"])

#                 # Check if there are valid sequences for output
#                 if len(result_df) == 0:
#                      st.warning("No valid sequences were processed.")
#                 else:
#                     # Create fragments for all sequences
#                     all_fragments = []
#                     fragment_analyses = []
                    
#                     for name, seq in sequences:
#                         if seq and len(seq) > 0:
#                             # Create 200-base fragments
#                             fragments = fragment_sequences([(name, seq)], 200)
#                             all_fragments.extend(fragments)
                            
#                             # Extract fragment analyses for display
#                             for frag_name, frag_seq, parent_name, frag_num, start_pos, analysis in fragments:
#                                 fragment_analyses.append({
#                                     "Gene Name": frag_name,
#                                     "Parent Gene": parent_name,
#                                     "Fragment Number": frag_num,
#                                     "Start Position": start_pos,
#                                     "End Position": start_pos + len(frag_seq) - 1,
#                                     "Fragment Length": len(frag_seq),
#                                     "Sequence": frag_seq,
#                                     "Length": analysis.get("Length", len(frag_seq)),
#                                     "A Count": analysis.get("A Count", 0),
#                                     "T Count": analysis.get("T Count", 0),
#                                     "G Count": analysis.get("G Count", 0),
#                                     "C Count": analysis.get("C Count", 0),
#                                     "A %": analysis.get("A %", 0),
#                                     "T %": analysis.get("T %", 0),
#                                     "G %": analysis.get("G %", 0),
#                                     "C %": analysis.get("C %", 0),
#                                     "GC %": analysis.get("GC %", 0),
#                                     "AT %": analysis.get("AT %", 0),
#                                     "Z-DNA Propensity %": analysis.get("Z-DNA Propensity %", 0)
#                                 })
                    
#                     fragment_df = pd.DataFrame(fragment_analyses)

#                     # Conditional Visualization Display
#                     if len(result_df) > 50:
#                         st.info("Analysis complete. Visualizations and on-page data table are disabled for more than 50 sequences to ensure performance. Please download your full results below.")
#                     else:
#                         st.toast("Analysis complete with automatic fragmentation!", icon=None)
#                         st.markdown("---")
                        
#                         # Results Section
#                         st.markdown("""
#                             <div class="section-container">
#                                 <h2 class="section-header">Results & Visualizations</h2>
#                             </div>
#                         """, unsafe_allow_html=True)
                        
#                         # Original Sequences Analysis
#                         st.markdown("### Original Sequences Analysis")
#                         with st.expander("Show Original Sequences Data", expanded=True):
#                             st.dataframe(result_df, use_container_width=True)
                        
#                         # Fragment Analysis
#                         st.markdown("### Fragment Analysis (200-base chunks)")
#                         with st.expander("Show Fragment Data", expanded=True):
#                             st.dataframe(fragment_df, use_container_width=True)
                            
#                         # Fragment Statistics
#                         if len(fragment_df) > 0:
#                             st.markdown("### Fragment Statistics")
#                             col1, col2, col3 = st.columns(3)
                            
#                             with col1:
#                                 st.metric("Total Fragments", len(fragment_df))
#                                 st.metric("Average GC Content", f"{fragment_df['GC %'].mean():.2f}%")
                                
#                             with col2:
#                                 st.metric("Average Z-DNA Propensity", f"{fragment_df['Z-DNA Propensity %'].mean():.2f}%")
#                                 st.metric("High GC Fragments (>60%)", len(fragment_df[fragment_df['GC %'] > 60]))
                                
#                             with col3:
#                                 st.metric("Low GC Fragments (<40%)", len(fragment_df[fragment_df['GC %'] < 40]))
#                                 st.metric("High Z-DNA Fragments (>50%)", len(fragment_df[fragment_df['Z-DNA Propensity %'] > 50]))
#                         with st.expander("Show GC Content Heatmap"):
#                             if len(result_df) > 0:
#                                 create_gc_heatmap(result_df)
#                             else:
#                                 st.info("No data to plot.")
#                         with st.expander("Show Sequence Logo"):
#                             # Use the original 'sequences' list for the logo, not result_df
#                             if len(sequences) > 1 and all(len(seq) > 0 for _, seq in sequences):
#                                 # Filter original sequences based on names in result_df to match analysis results
#                                 valid_gene_names = result_df['Gene Name'].tolist()
#                                 sequences_for_logo = [(name, seq) for name, seq in sequences if name in valid_gene_names]
#                                 if len(sequences_for_logo) > 1:
#                                     create_sequence_logo([seq for _, seq in sequences_for_logo])
#                                 else:
#                                     st.info("Not enough valid sequences to plot a sequence logo.")
#                             else:
#                                 st.info("Not enough data to plot a sequence logo.")

#                         with st.expander("Show Interactive Plots"):
#                             if len(result_df) > 1:
#                                 create_interactive_plots(result_df)
#                             else:
#                                 st.info("Not enough data to plot interactive plots.")

#                     # Export Results section - always shown if there are results
#                     st.markdown("---")
#                     st.markdown("""
#                         <div class="section-container">
#                             <h2 class="section-header">Export Results</h2>
#                         </div>
#                     """, unsafe_allow_html=True)
                    
#                     output_filename = st.text_input("Enter output file name (without extension):", "gc_output_v4", key="export_filename")
                    
#                     # Export buttons in columns for better layout
#                     col1, col2, col3, col4 = st.columns(4)
                    
#                     with col1:
#                         if st.button("Export All Results"):
#                             zip_buffer = export_all_to_zip(result_df, output_filename)
#                             st.toast("All results exported as ZIP!", icon=None)
#                             st.download_button(
#                                 label="Download All Results (ZIP)",
#                                 data=zip_buffer,
#                                 file_name=f"{output_filename}_all_results.zip",
#                                 mime="application/zip"
#                             )
                    
#                     with col2:
#                         excel_buffer = BytesIO()
#                         result_df.to_excel(excel_buffer, index=False, engine='openpyxl')
#                         excel_buffer.seek(0)
#                         st.download_button(
#                             label="Download Original Sequences (Excel)",
#                             data=excel_buffer,
#                             file_name=f"{output_filename}_original.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                    
#                     with col3:
#                         # FASTA export button
#                         # Use result_df for FASTA/GenBank exports to include only analyzed sequences
#                         fasta_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         fasta_buffer = export_to_fasta(fasta_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download Original Sequences (FASTA)",
#                             data=fasta_buffer,
#                             file_name=f"{output_filename}_original.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col4:
#                         # GenBank export button
#                         gb_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         gb_buffer = export_to_genbank(gb_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download Original Sequences (GenBank)",
#                             data=gb_buffer,
#                             file_name=f"{output_filename}_original.gb",
#                             mime="text/plain"
#                         )
                    
#                     # Fragment Export Section
#                     if len(fragment_df) > 0:
#                         st.markdown("### Fragment Export Options")
#                         frag_col1, frag_col2, frag_col3, frag_col4 = st.columns(4)
                        
#                         with frag_col1:
#                             # Export fragments to Excel
#                             frag_excel_buffer = export_fragments_to_excel(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (Excel)",
#                                 data=frag_excel_buffer,
#                                 file_name=f"{output_filename}_fragments.xlsx",
#                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                             )
                        
#                         with frag_col2:
#                             # Export fragments to FASTA
#                             frag_fasta_buffer = export_fragments_to_fasta(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (FASTA)",
#                                 data=frag_fasta_buffer,
#                                 file_name=f"{output_filename}_fragments.fasta",
#                                 mime="text/plain"
#                             )
                        
#                         with frag_col3:
#                             # Export fragments to GenBank
#                             frag_gb_buffer = export_fragments_to_genbank(all_fragments)
#                             st.download_button(
#                                 label="Download Fragments (GenBank)",
#                                 data=frag_gb_buffer,
#                                 file_name=f"{output_filename}_fragments.gb",
#                                 mime="text/plain"
#                             )
                        
#                         with frag_col4:
#                             # Combined export with both original and fragments
#                             combined_data = []
                            
#                             # Add original sequences
#                             for _, row in result_df.iterrows():
#                                 combined_data.append({
#                                     "Type": "Original",
#                                     "Gene Name": row["Gene Name"],
#                                     "Sequence": row["Sequence"],
#                                     "Length": row["Length"],
#                                     "GC %": row["GC %"],
#                                     "Z-DNA Propensity %": row.get("Z-DNA Propensity %", 0),
#                                     "A Count": row["A Count"],
#                                     "T Count": row["T Count"],
#                                     "G Count": row["G Count"],
#                                     "C Count": row["C Count"]
#                                 })
                            
#                             # Add fragments
#                             for _, row in fragment_df.iterrows():
#                                 combined_data.append({
#                                     "Type": "Fragment",
#                                     "Gene Name": row["Gene Name"],
#                                     "Parent Gene": row["Parent Gene"],
#                                     "Fragment Number": row["Fragment Number"],
#                                     "Start Position": row["Start Position"],
#                                     "End Position": row["End Position"],
#                                     "Sequence": row["Sequence"],
#                                     "Length": row["Length"],
#                                     "GC %": row["GC %"],
#                                     "Z-DNA Propensity %": row["Z-DNA Propensity %"],
#                                     "A Count": row["A Count"],
#                                     "T Count": row["T Count"],
#                                     "G Count": row["G Count"],
#                                     "C Count": row["C Count"]
#                                 })
                            
#                             combined_df = pd.DataFrame(combined_data)
#                             combined_buffer = BytesIO()
#                             combined_df.to_excel(combined_buffer, index=False, engine='openpyxl')
#                             combined_buffer.seek(0)
#                             st.download_button(
#                                 label="Download Combined (Excel)",
#                                 data=combined_buffer,
#                                 file_name=f"{output_filename}_combined.xlsx",
#                                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                             )

#     # Add Sequence Fragmenter tool to sidebar
#     with st.sidebar.expander("Sequence Fragmenter (Split Tool)"):
#         st.markdown("""
#         <div style='background: rgba(52,152,219,0.1); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3); margin-bottom: 1rem;'>
#         <strong>Split sequences into fragments of a chosen length.</strong><br>
#         Output is compatible with TOOL-1 and available in multiple formats.
#         </div>
#         """, unsafe_allow_html=True)
        
#         frag_input_method = st.radio("Input method for fragmenter", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"], key="frag_input_method")
#         frag_sequences = []
        
#         if frag_input_method == "Upload CSV/Excel":
#             frag_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="frag_file")
#             if frag_file:
#                 try:
#                     frag_file.seek(0)
#                     if frag_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(frag_file)
#                     else:
#                         df = pd.read_csv(frag_file)
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         frag_sequences = list(zip(df["Gene Name"], df["Sequence"]))
#                         st.success(f"Successfully loaded {len(frag_sequences)} sequences from {frag_file.name}")
#                     else:
#                         st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
#                 except Exception as e:
#                     st.warning(f"Error reading file: {str(e)}")
                    
#         elif frag_input_method == "Upload FASTA":
#             frag_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="frag_fasta")
#             if frag_fasta:
#                 frag_fasta.seek(0)
#                 frag_sequences = process_fasta(frag_fasta, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from FASTA file")
                    
#         elif frag_input_method == "Upload GenBank":
#             frag_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="frag_gb")
#             if frag_gb:
#                 frag_gb.seek(0)
#                 frag_sequences = process_genbank(frag_gb, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from GenBank file")
                    
#         elif frag_input_method == "Manual Entry":
#             frag_num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1, key="frag_num")
#             frag_manual_entries = []
#             for i in range(frag_num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"frag_name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"frag_seq_{i}")
#                 if name and seq:
#                     frag_manual_entries.append((name, seq))
#             frag_sequences.extend(frag_manual_entries)
#             if frag_manual_entries:
#                 st.success(f"Successfully entered {len(frag_manual_entries)} sequences")
                
#         # Fragment length selection with better styling
#         st.markdown("---")
#         st.markdown("**Fragment Configuration**")
#         frag_length = st.selectbox("Fragment length (nucleotides)", [25, 50, 100, 200, 400], key="frag_length")
        
#         if frag_sequences:
#             st.info(f"Ready to process {len(frag_sequences)} sequences into {frag_length}-nucleotide fragments")
            
#             if st.button("Split and Export Fragments", key="frag_button"):
#                 with st.spinner("Processing fragments and calculating GC content..."):
#                     fragments = fragment_sequences(frag_sequences, frag_length)
                    
#                 if not fragments:
#                     st.warning("No fragments produced. Check your input sequences and fragment length.")
#                 else:
#                     st.success(f"Successfully produced {len(fragments)} fragments of {frag_length} nucleotides each!")
                    
#                     # Calculate overall statistics
#                     total_gc_content = sum(f[5].get("GC %", 0) for f in fragments)
#                     avg_gc_content = total_gc_content / len(fragments)
                    
#                     # Display fragment statistics with GC content
#                     st.markdown("**Fragment Statistics:**")
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         st.markdown(f"""
#                         - Total fragments: {len(fragments)}
#                         - Fragment length: {frag_length} nucleotides
#                         - Parent sequences: {len(set([f[2] for f in fragments]))}
#                         - Average GC content: {avg_gc_content:.2f}%
#                         """)
                    
#                     with col2:
#                         # Show GC content distribution
#                         gc_values = [f[5].get("GC %", 0) for f in fragments]
#                         min_gc = min(gc_values)
#                         max_gc = max(gc_values)
#                         st.markdown(f"""
#                         - GC content range: {min_gc:.2f}% - {max_gc:.2f}%
#                         - High GC fragments (>60%): {sum(1 for gc in gc_values if gc > 60)}
#                         - Low GC fragments (<40%): {sum(1 for gc in gc_values if gc < 40)}
#                         """)
                    
#                     # Export options in a more organized way
#                     st.markdown("**Download Options**")
                    
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         excel_buffer = export_fragments_to_excel(fragments)
#                         st.download_button(
#                             label="Excel Format (with Analysis)",
#                             data=excel_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                        
#                         fasta_buffer = export_fragments_to_fasta(fragments)
#                         st.download_button(
#                             label="FASTA Format (with Analysis)",
#                             data=fasta_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col2:
#                         gb_buffer = export_fragments_to_genbank(fragments)
#                         st.download_button(
#                             label="GenBank Format (with Analysis)",
#                             data=gb_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.gb",
#                             mime="text/plain"
#                         )
                        
#                         # Show a sample of the analysis data
#                         if len(fragments) > 0:
#                             st.markdown("**Sample Fragment Analysis:**")
#                             sample_frag = fragments[0]
#                             st.markdown(f"""
#                             **{sample_frag[0]}** (from {sample_frag[2]})
#                             - GC Content: {sample_frag[5].get('GC %', 0):.2f}%
#                             - A: {sample_frag[5].get('A Count', 0)} ({sample_frag[5].get('A %', 0):.2f}%)
#                             - T: {sample_frag[5].get('T Count', 0)} ({sample_frag[5].get('T %', 0):.2f}%)
#                             - G: {sample_frag[5].get('G Count', 0)} ({sample_frag[5].get('G %', 0):.2f}%)
#                             - C: {sample_frag[5].get('C Count', 0)} ({sample_frag[5].get('C %', 0):.2f}%)
#                             """)
#         else:
#             st.info("Please upload or enter sequences to begin fragmentation")

#     # Professional footer
#     st.markdown("---")
#     st.markdown("""
#         <div class="footer">
#             <p>Developed by Shubh Rakesh Nahar | Troy University | Professional DNA Sequence Analysis Tool</p>
#         </div>
#     """, unsafe_allow_html=True)

# if __name__ == "__main__":
#     main()





# import pandas as pd
# import streamlit as st
# import matplotlib.pyplot as plt
# from io import BytesIO
# from Bio import SeqIO
# import json
# import plotly.express as px
# import plotly.graph_objects as go
# import seaborn as sns
# import numpy as np
# from matplotlib.patches import Rectangle
# import math
# import zipfile
# import random
# from Bio.Seq import Seq
# from Bio.SeqRecord import SeqRecord
# import io
# import openpyxl

# # Custom CSS for professional styling
# def load_css():
#     st.markdown("""
#     <style>
#     /* Main background and overall styling */
#     .main {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#         color: #ffffff;
#     }
    
#     .stApp {
#         background: linear-gradient(135deg, #1a1a2e 0%, #16213e 25%, #0f3460 50%, #16213e 75%, #1a1a2e 100%);
#     }
    
#     /* Sidebar styling */
#     .css-1d391kg {
#         background: linear-gradient(180deg, #0f1419 0%, #1a2332 50%, #0f1419 100%);
#         border-right: 2px solid #3498db;
#     }
    
#     /* Header styling */
#     .main-header {
#         background: linear-gradient(90deg, #2c3e50 0%, #34495e 50%, #2c3e50 100%);
#         padding: 2rem;
#         border-radius: 15px;
#         margin-bottom: 2rem;
#         box-shadow: 0 8px 32px rgba(0,0,0,0.4);
#         border: 1px solid #3498db;
#     }
    
#     .main-header h1 {
#         color: #ecf0f1;
#         text-align: center;
#         font-size: 2.5rem;
#         font-weight: 700;
#         margin: 0;
#         text-shadow: 2px 2px 4px rgba(0,0,0,0.5);
#     }
    
#     /* Section styling */
#     .section-container {
#         background: rgba(255, 255, 255, 0.05);
#         padding: 2rem;
#         border-radius: 15px;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         backdrop-filter: blur(10px);
#         box-shadow: 0 8px 32px rgba(0,0,0,0.3);
#     }
    
#     .section-header {
#         color: #3498db;
#         font-size: 1.5rem;
#         font-weight: 600;
#         margin-bottom: 1rem;
#         text-align: center;
#         text-transform: uppercase;
#         letter-spacing: 1px;
#     }
    
#     /* Button styling */
#     .stButton > button {
#         background: linear-gradient(45deg, #3498db, #2980b9);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
#     }
    
#     .stButton > button:hover {
#         background: linear-gradient(45deg, #2980b9, #1f5f8b);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(52, 152, 219, 0.4);
#     }
    
#     /* Input styling */
#     .stTextInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stTextArea > div > div > textarea {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     .stSelectbox > div > div > select {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* File uploader styling */
#     .stFileUploader > div {
#         background: rgba(255, 255, 255, 0.05);
#         border: 2px dashed rgba(52, 152, 219, 0.5);
#         border-radius: 15px;
#         padding: 2rem;
#         text-align: center;
#     }
    
#     /* Radio button styling */
#     .stRadio > div {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         padding: 1rem;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Number input styling */
#     .stNumberInput > div > div > input {
#         background: rgba(255, 255, 255, 0.1);
#         border: 1px solid rgba(52, 152, 219, 0.5);
#         border-radius: 10px;
#         color: white;
#         padding: 0.75rem;
#     }
    
#     /* Expander styling */
#     .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #34495e, #2c3e50);
#         color: #ecf0f1;
#         border-radius: 10px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Success/Info/Warning styling */
#     .stSuccess {
#         background: rgba(46, 204, 113, 0.2);
#         border: 1px solid #2ecc71;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #2ecc71;
#     }
    
#     .stInfo {
#         background: rgba(52, 152, 219, 0.2);
#         border: 1px solid #3498db;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #3498db;
#     }
    
#     .stWarning {
#         background: rgba(241, 196, 15, 0.2);
#         border: 1px solid #f1c40f;
#         border-radius: 10px;
#         padding: 1rem;
#         color: #f1c40f;
#     }
    
#     /* Dataframe styling */
#     .dataframe {
#         background: rgba(255, 255, 255, 0.05);
#         border-radius: 10px;
#         overflow: hidden;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Download button styling */
#     .stDownloadButton > button {
#         background: linear-gradient(45deg, #27ae60, #229954);
#         color: white;
#         border: none;
#         border-radius: 25px;
#         padding: 0.75rem 2rem;
#         font-weight: 600;
#         font-size: 1rem;
#         transition: all 0.3s ease;
#         box-shadow: 0 4px 15px rgba(39, 174, 96, 0.3);
#     }
    
#     .stDownloadButton > button:hover {
#         background: linear-gradient(45deg, #229954, #1e8449);
#         transform: translateY(-2px);
#         box-shadow: 0 6px 20px rgba(39, 174, 96, 0.4);
#     }
    
#     /* Progress bar styling */
#     .stProgress > div > div > div {
#         background: linear-gradient(90deg, #3498db, #2980b9);
#         border-radius: 10px;
#     }
    
#     /* Sidebar expander styling */
#     .css-1d391kg .streamlit-expanderHeader {
#         background: linear-gradient(45deg, #2c3e50, #34495e);
#         color: #ecf0f1;
#         border-radius: 8px;
#         font-weight: 600;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Footer styling */
#     .footer {
#         text-align: center;
#         color: #bdc3c7;
#         font-size: 0.9rem;
#         margin-top: 3rem;
#         padding: 1rem;
#         border-top: 1px solid rgba(52, 152, 219, 0.3);
#     }
    
#     /* Tool card styling */
#     .tool-card {
#         background: rgba(255, 255, 255, 0.08);
#         border-radius: 15px;
#         padding: 1.5rem;
#         margin: 1rem 0;
#         border: 1px solid rgba(52, 152, 219, 0.3);
#         box-shadow: 0 4px 15px rgba(0,0,0,0.3);
#     }
    
#     /* Fragmenter specific styling */
#     .fragmenter-section {
#         background: linear-gradient(135deg, rgba(52, 152, 219, 0.1) 0%, rgba(41, 128, 185, 0.1) 100%);
#         border: 2px solid rgba(52, 152, 219, 0.5);
#         border-radius: 20px;
#         padding: 2rem;
#         margin: 1rem 0;
#     }
    
#     /* Responsive design */
#     @media (max-width: 768px) {
#         .main-header h1 {
#             font-size: 2rem;
#         }
        
#         .section-container {
#             padding: 1rem;
#         }
#     }
#     </style>
#     """, unsafe_allow_html=True)

# # --- Instructions ---
# def show_instructions():
#     st.markdown("""
#     ### 📋 Instructions
    
#     1. **Choose Input Method**:
#        - **Upload CSV/Excel**: Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns
#        - **Upload FASTA**: Upload a FASTA file containing DNA sequences
#        - **Upload GenBank**: Upload a GenBank file containing DNA sequences
#        - **Manual Entry**: Enter sequences manually with gene names
    
#     2. **Input Format**:
#        - Only A, T, G, C nucleotides are accepted (case insensitive)
#        - Invalid characters will be automatically removed
#        - Each sequence must have a unique gene name
    
#     3. **Analysis**:
#        - Click "Calculate GC Content" to process sequences
#        - View results in the interactive table
#        - Explore visualizations of GC content and nucleotide composition
    
#     4. **Export Options**:
#        - Export results in multiple formats (Excel, CSV, JSON)
#        - Use "Export All Results" for a complete dataset export
#        - Customize output filename before downloading
#     """)

# # --- Sequence Sanitization and Validation ---
# def sanitize_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return ''.join(filter(lambda x: x.upper() in ['A', 'T', 'G', 'C'], seq.upper()))

# def is_valid_sequence(seq):
#     # Convert to string if not already
#     if not isinstance(seq, str):
#         seq = str(seq)
#     return all(base in ['A', 'T', 'G', 'C'] for base in seq.upper())

# # --- Nucleotide Analysis ---
# def analyze_sequence(name, seq):
#     # Convert sequence to string if it's not already
#     if not isinstance(seq, str):
#         seq = str(seq)
    
#     seq = sanitize_sequence(seq)
#     length = len(seq)
#     if length == 0:
#         return {"Gene Name": name, "Error": "Invalid sequence or empty after sanitization"}

#     a_count = seq.count('A')
#     t_count = seq.count('T')
#     g_count = seq.count('G')
#     c_count = seq.count('C')
#     gc_count = g_count + c_count
#     at_count = a_count + t_count

#     return {
#         "Gene Name": name,
#         "Sequence": seq,
#         "Length": length,
#         "A Count": a_count,
#         "T Count": t_count,
#         "G Count": g_count,
#         "C Count": c_count,
#         "A %": round((a_count / length) * 100, 2),
#         "T %": round((t_count / length) * 100, 2),
#         "G %": round((g_count / length) * 100, 2),
#         "C %": round((c_count / length) * 100, 2),
#         "GC %": round((gc_count / length) * 100, 2),
#         "AT %": round((at_count / length) * 100, 2),
#     }

# def process_fasta(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for FASTA
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "fasta")
#         count = 0
#         for record in records:
#             if count >= max_sequences:
#                 break
#             # Convert Seq object to string
#             sequences.append((str(record.id), str(record.seq)))
#             count += 1
#         if count == 0:
#             st.warning("No sequences found in the uploaded FASTA file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the FASTA file.")
#     except Exception as e:
#         st.error(f"Error parsing FASTA file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def process_genbank(file, max_sequences=1000):
#     sequences = []
#     try:
#         # Ensure the file is read from the beginning
#         file.seek(0)
#         # Read binary and decode to text for GenBank
#         text = file.read().decode("utf-8")
#         text_io = io.StringIO(text)
#         records = SeqIO.parse(text_io, "genbank")
#         count = 0
#         for record in records:
#              if count >= max_sequences:
#                  break
#              # Ensure molecule_type is set for GenBank export compatibility later
#              # For uploaded GenBank, try to get molecule_type from annotations, default to DNA
#              mol_type = record.annotations.get("molecule_type", "DNA")
#              sequences.append((str(record.id), str(record.seq)))
#              count += 1
#         if count == 0:
#              st.warning("No sequences found in the uploaded GenBank file after parsing.")
#         elif count >= max_sequences:
#              st.info(f"Processed the first {max_sequences} sequences from the GenBank file.")

#     except Exception as e:
#         st.error(f"Error parsing GenBank file: {e}")
#         # Return empty list on error
#         return []

#     return sequences

# def display_visuals(df):
#     st.subheader("📊 GC Content Distribution")
#     fig, ax = plt.subplots()
#     df.plot.bar(x='Gene Name', y='GC %', ax=ax, color='purple', legend=False)
#     plt.ylabel('GC %')
#     plt.xticks(rotation=45, ha='right')
#     st.pyplot(fig)

#     st.subheader("🧬 Nucleotide Composition Per Sequence")
#     for _, row in df.iterrows():
#         fig, ax = plt.subplots()
#         ax.pie(
#             [row['A %'], row['T %'], row['G %'], row['C %']],
#             labels=['A %', 'T %', 'G %', 'C %'],
#             autopct='%1.1f%%',
#             startangle=90
#         )
#         ax.set_title(f"{row['Gene Name']} - Base % Composition")
#         st.pyplot(fig)

# def export_data(df, format_type):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     if format_type == "Excel":
#         towrite = BytesIO()
#         with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         towrite.seek(0)
#         return towrite, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
#     elif format_type == "CSV":
#         towrite = BytesIO()
#         # Add signature as a comment in the first line
#         towrite.write(f"# {signature}\n".encode())
#         df.to_csv(towrite, index=False)
#         towrite.seek(0)
#         return towrite, "text/csv", "csv"
#     else:  # JSON
#         towrite = BytesIO()
#         # Add signature to the JSON metadata
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         towrite.write(json_str.encode())
#         towrite.seek(0)
#         return towrite, "application/json", "json"

# def create_gc_heatmap(df):
#     st.subheader("🌡️ GC Content Heatmap")
    
#     # Create a matrix of GC content for each position
#     sequences = df['Sequence'].tolist()
#     max_len = max(len(seq) for seq in sequences)
    
#     # Initialize matrix
#     gc_matrix = np.zeros((len(sequences), max_len))
    
#     # Fill matrix with GC content for each position
#     for i, seq in enumerate(sequences):
#         for j in range(len(seq)):
#             if j < len(seq):
#                 window = seq[max(0, j-10):min(len(seq), j+11)]
#                 gc_count = window.count('G') + window.count('C')
#                 gc_matrix[i, j] = (gc_count / len(window)) * 100
    
#     # Create heatmap using plotly
#     fig = go.Figure(data=go.Heatmap(
#         z=gc_matrix,
#         x=list(range(max_len)),
#         y=df['Gene Name'].tolist(),
#         colorscale='Viridis',
#         colorbar=dict(title='GC %')
#     ))
    
#     fig.update_layout(
#         title='GC Content Distribution Across Sequences',
#         xaxis_title='Position',
#         yaxis_title='Gene Name',
#         height=400 + (len(sequences) * 20)  # Adjust height based on number of sequences
#     )
    
#     st.plotly_chart(fig, use_container_width=True)

# def calculate_information_content(freq):
#     """Calculate information content in bits."""
#     if freq == 0:
#         return 0
#     return freq * math.log2(freq * 4)  # 4 for number of nucleotides

# def create_sequence_logo(sequences):
#     st.subheader("🎨 Sequence Logo")
    
#     # Calculate position frequency matrix
#     max_len = max(len(seq) for seq in sequences)
#     pfm = np.zeros((4, max_len))  # 4 nucleotides
    
#     for seq in sequences:
#         for i, base in enumerate(seq):
#             if i < max_len:
#                 if base == 'A':
#                     pfm[0, i] += 1
#                 elif base == 'T':
#                     pfm[1, i] += 1
#                 elif base == 'G':
#                     pfm[2, i] += 1
#                 elif base == 'C':
#                     pfm[3, i] += 1
    
#     # Normalize
#     pfm = pfm / len(sequences)
    
#     # Calculate information content
#     ic = np.zeros(max_len)
#     for i in range(max_len):
#         ic[i] = sum(calculate_information_content(freq) for freq in pfm[:, i])
    
#     # Create the plot
#     fig, ax = plt.subplots(figsize=(12, 4))
    
#     # Colors for nucleotides
#     colors = {'A': '#2ecc71', 'T': '#e74c3c', 'G': '#f1c40f', 'C': '#3498db'}
#     bases = ['A', 'T', 'G', 'C']
    
#     # Plot each position
#     for i in range(max_len):
#         # Sort frequencies for this position
#         freqs = pfm[:, i]
#         sorted_indices = np.argsort(freqs)
        
#         # Plot each base
#         y_bottom = 0
#         for idx in sorted_indices:
#             if freqs[idx] > 0:
#                 height = freqs[idx] * ic[i]
#                 rect = Rectangle((i, y_bottom), 1, height,
#                                facecolor=colors[bases[idx]],
#                                edgecolor='black',
#                                linewidth=0.5)
#                 ax.add_patch(rect)
#                 y_bottom += height
    
#     # Customize the plot
#     ax.set_xlim(0, max_len)
#     ax.set_ylim(0, max(ic) * 1.1)
#     ax.set_xlabel('Position')
#     ax.set_ylabel('Bits')
#     ax.set_title('Sequence Logo')
    
#     # Add legend
#     legend_elements = [Rectangle((0, 0), 1, 1, facecolor=color, edgecolor='black')
#                       for color in colors.values()]
#     ax.legend(legend_elements, bases, loc='upper right')
    
#     # Remove spines
#     for spine in ['top', 'right']:
#         ax.spines[spine].set_visible(False)
    
#     # Show the plot
#     st.pyplot(fig)
#     plt.close(fig)

# def create_interactive_plots(df):
#     st.subheader("📊 Interactive Plots")
    
#     # GC Content Distribution
#     fig_gc = px.box(df, y='GC %', title='GC Content Distribution',
#                     hover_data=['Gene Name', 'Length', 'GC %'])
#     fig_gc.update_traces(marker_color='purple')
#     st.plotly_chart(fig_gc, use_container_width=True)
    
#     # Nucleotide Composition
#     fig_comp = px.bar(df, 
#                      x='Gene Name',
#                      y=['A %', 'T %', 'G %', 'C %'],
#                      title='Nucleotide Composition by Gene',
#                      barmode='group',
#                      hover_data=['Length', 'GC %'])
#     st.plotly_chart(fig_comp, use_container_width=True)
    
#     # Length vs GC Content
#     fig_scatter = px.scatter(df,
#                            x='Length',
#                            y='GC %',
#                            color='GC %',
#                            hover_data=['Gene Name', 'Length', 'GC %'],
#                            title='Sequence Length vs GC Content')
#     st.plotly_chart(fig_scatter, use_container_width=True)

# def export_all_to_zip(df, output_filename):
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     zip_buffer = BytesIO()
#     with zipfile.ZipFile(zip_buffer, "w") as zip_file:
#         # Excel with formatting
#         excel_buffer = BytesIO()
#         with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#             df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#             # Get the workbook and the worksheet
#             workbook = writer.book
#             worksheet = writer.sheets['GC Content Analysis']
            
#             # Add signature as a comment in cell A1
#             worksheet.cell(row=1, column=1).comment = openpyxl.comments.Comment(signature, "Author")
            
#             # Auto-adjust columns' width
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(str(cell.value))
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add some basic formatting
#             for row in worksheet.iter_rows(min_row=1, max_row=1):
#                 for cell in row:
#                     cell.font = cell.font.copy(bold=True)
            
#             # Add signature in the last row
#             last_row = len(df) + 2
#             worksheet.cell(row=last_row, column=1, value=signature)
        
#         excel_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.xlsx", excel_buffer.read())
        
#         # CSV
#         csv_buffer = BytesIO()
#         # Add signature as a comment in the first line
#         csv_buffer.write(f"# {signature}\n".encode())
#         df.to_csv(csv_buffer, index=False)
#         csv_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.csv", csv_buffer.read())
        
#         # JSON
#         json_data = {
#             "metadata": {
#                 "signature": signature,
#                 "timestamp": pd.Timestamp.now().isoformat()
#             },
#             "data": df.to_dict(orient='records')
#         }
#         json_str = json.dumps(json_data, indent=2)
#         json_buffer = BytesIO()
#         json_buffer.write(json_str.encode())
#         json_buffer.seek(0)
#         zip_file.writestr(f"{output_filename}.json", json_buffer.read())
    
#     zip_buffer.seek(0)
#     return zip_buffer

# def export_to_fasta(sequences, result_df):
#     """Export sequences to FASTA format with analysis results in description"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     # Add signature as a comment at the top
#     output.write(f"# {signature}\n")
#     for idx, row in result_df.iterrows():
#         # Create a description with analysis results
#         description = f"GC Content: {row['GC %']:.2f}%, Length: {row['Length']}, "
#         description += f"A: {row['A Count']}, T: {row['T Count']}, G: {row['G Count']}, C: {row['C Count']}"
        
#         # Write sequence with description
#         output.write(f">{row['Gene Name']} {description}\n")
#         output.write(f"{row['Sequence']}\n")
    
#     # Add signature at the end
#     output.write(f"\n# {signature}\n")
#     return output.getvalue()

# def export_to_genbank(sequences, result_df):
#     """Export sequences to GenBank format with analysis results in features"""
#     signature = "Made by Shubh Rakesh Nahar / Troy University"
#     output = io.StringIO()
#     for idx, row in result_df.iterrows():
#         # Write GenBank header
#         output.write(f"LOCUS       {row['Gene Name']}              {row['Length']} bp    DNA     linear\n")
#         output.write(f"DEFINITION  {row['Gene Name']}\n")
#         output.write(f"ACCESSION   {row['Gene Name']}\n")
#         output.write(f"VERSION     {row['Gene Name']}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
        
#         # Add signature in the features section
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write("     source          1..{}\n".format(row['Length']))
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write('                     /note="{}"\n'.format(signature))
#         output.write('     misc_feature    1..{}\n'.format(row['Length']))
#         output.write('                     /note="GC Content: {:.2f}%"\n'.format(row['GC %']))
#         output.write('                     /note="A Count: {}"\n'.format(row['A Count']))
#         output.write('                     /note="T Count: {}"\n'.format(row['T Count']))
#         output.write('                     /note="G Count: {}"\n'.format(row['G Count']))
#         output.write('                     /note="C Count: {}"\n'.format(row['C Count']))
        
#         # Write sequence
#         output.write("ORIGIN\n")
#         sequence = row['Sequence']
#         for i in range(0, len(sequence), 60):
#             chunk = sequence[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
    
#     return output.getvalue()

# def fragment_sequences(sequences, fragment_length):
#     """
#     Splits each sequence into fragments of the given length and calculates GC content analysis.
#     Returns a list of (Fragment Name, Fragment Sequence, Parent Name, Fragment Index, Fragment Length, Analysis Data).
#     """
#     fragments = []
#     for name, seq in sequences:
#         seq = sanitize_sequence(seq)
#         for i in range(0, len(seq), fragment_length):
#             frag_seq = seq[i:i+fragment_length]
#             if len(frag_seq) == fragment_length:
#                 frag_name = f"{name}_frag_{i//fragment_length+1}"
#                 # Calculate GC content analysis for the fragment
#                 analysis = analyze_sequence(frag_name, frag_seq)
#                 fragments.append((frag_name, frag_seq, name, i//fragment_length+1, fragment_length, analysis))
#     return fragments

# def export_fragments_to_excel(fragments):
#     # Create a comprehensive DataFrame with all analysis data
#     data = []
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         row = {
#             "Gene Name": frag_name,
#             "Sequence": frag_seq,
#             "Parent Name": parent,
#             "Fragment Index": idx,
#             "Fragment Length": frag_len,
#             "Length": analysis.get("Length", len(frag_seq)),
#             "A Count": analysis.get("A Count", 0),
#             "T Count": analysis.get("T Count", 0),
#             "G Count": analysis.get("G Count", 0),
#             "C Count": analysis.get("C Count", 0),
#             "A %": analysis.get("A %", 0),
#             "T %": analysis.get("T %", 0),
#             "G %": analysis.get("G %", 0),
#             "C %": analysis.get("C %", 0),
#             "GC %": analysis.get("GC %", 0),
#             "AT %": analysis.get("AT %", 0)
#         }
#         data.append(row)
    
#     df = pd.DataFrame(data)
#     towrite = BytesIO()
#     with pd.ExcelWriter(towrite, engine='openpyxl') as writer:
#         df.to_excel(writer, index=False, sheet_name='Fragments with Analysis')
#         workbook = writer.book
#         worksheet = writer.sheets['Fragments with Analysis']
        
#         # Auto-adjust columns' width
#         for column in worksheet.columns:
#             max_length = 0
#             column = [cell for cell in column]
#             for cell in column:
#                 try:
#                     if len(str(cell.value)) > max_length:
#                         max_length = len(str(cell.value))
#                 except:
#                     pass
#             adjusted_width = (max_length + 2)
#             worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
#         # Add formatting to headers
#         for row in worksheet.iter_rows(min_row=1, max_row=1):
#             for cell in row:
#                 cell.font = cell.font.copy(bold=True)
    
#     towrite.seek(0)
#     return towrite


# def export_fragments_to_fasta(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         # Create a comprehensive description with all analysis data
#         description = f"parent={parent} idx={idx} len={frag_len} "
#         description += f"GC_Content={analysis.get('GC %', 0):.2f}% "
#         description += f"A={analysis.get('A Count', 0)}({analysis.get('A %', 0):.2f}%) "
#         description += f"T={analysis.get('T Count', 0)}({analysis.get('T %', 0):.2f}%) "
#         description += f"G={analysis.get('G Count', 0)}({analysis.get('G %', 0):.2f}%) "
#         description += f"C={analysis.get('C Count', 0)}({analysis.get('C %', 0):.2f}%)"
        
#         output.write(f">{frag_name} {description}\n{frag_seq}\n")
#     return output.getvalue()


# def export_fragments_to_genbank(fragments):
#     output = io.StringIO()
#     for frag_name, frag_seq, parent, idx, frag_len, analysis in fragments:
#         output.write(f"LOCUS       {frag_name} {frag_len} bp    DNA     linear\n")
#         output.write(f"DEFINITION  Fragment of {parent}, index {idx}\n")
#         output.write(f"ACCESSION   {frag_name}\n")
#         output.write(f"VERSION     {frag_name}\n")
#         output.write(f"SOURCE      .\n")
#         output.write(f"  ORGANISM  .\n")
#         output.write(f"            .\n")
#         output.write("FEATURES             Location/Qualifiers\n")
#         output.write(f"     source          1..{frag_len}\n")
#         output.write('                     /organism="."\n')
#         output.write('                     /mol_type="genomic DNA"\n')
#         output.write(f'                     /note="Fragment of {parent}, index {idx}"\n')
#         output.write(f'                     /note="GC Content: {analysis.get("GC %", 0):.2f}%"\n')
#         output.write(f'                     /note="A Count: {analysis.get("A Count", 0)} ({analysis.get("A %", 0):.2f}%)"\n')
#         output.write(f'                     /note="T Count: {analysis.get("T Count", 0)} ({analysis.get("T %", 0):.2f}%)"\n')
#         output.write(f'                     /note="G Count: {analysis.get("G Count", 0)} ({analysis.get("G %", 0):.2f}%)"\n')
#         output.write(f'                     /note="C Count: {analysis.get("C Count", 0)} ({analysis.get("C %", 0):.2f}%)"\n')
#         output.write(f'                     /note="AT Content: {analysis.get("AT %", 0):.2f}%"\n')
#         output.write("ORIGIN\n")
#         for i in range(0, len(frag_seq), 60):
#             chunk = frag_seq[i:i+60]
#             output.write(f"     {i+1:9d} {chunk}\n")
#         output.write("//\n")
#     return output.getvalue()

# def main():
#     st.set_page_config(page_title="GC Content Calculator", page_icon=None, layout="wide")
    
#     # Load custom CSS
#     load_css()
    
#     # Password protection
#     if 'authenticated' not in st.session_state:
#         st.session_state.authenticated = False
    
#     if not st.session_state.authenticated:
#         # Login screen
#         st.markdown("""
#             <div class="main-header">
#                 <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Authentication Required</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         # Center the login form
#         col1, col2, col3 = st.columns([1, 2, 1])
#         with col2:
#             st.markdown("""
#                 <div style='text-align: center; margin: 2rem 0;'>
#                     <h3>Please enter the password to access the application</h3>
#                 </div>
#             """, unsafe_allow_html=True)
            
#             password = st.text_input("Password", type="password", key="password_input")
            
#             if st.button("Login", key="login_button"):
#                 if password == "TroyDNA2024":
#                     st.session_state.authenticated = True
#                     st.success("Authentication successful! Redirecting...")
#                     st.rerun()
#                 else:
#                     st.error("Incorrect password. Please try again.")
            
#             st.markdown("""
#                 <div style='text-align: center; margin-top: 2rem; color: #bdc3c7;'>
#                     <p>Developed by Shubh Rakesh Nahar | Troy University</p>
#                 </div>
#             """, unsafe_allow_html=True)
        
#         return
    
#     # Main application (only shown after authentication)
#     # Professional header with gradient background
#     st.markdown("""
#         <div class="main-header">
#             <h1>GC Content Calculator & Sequence Analysis Tool</h1>
#         </div>
#     """, unsafe_allow_html=True)

#     sequences = []  # Ensure sequences is always defined

#     # Sidebar with logo, navigation, and about info
#     st.sidebar.image("https://cdn-icons-png.flaticon.com/512/616/616494.png", width=100)
#     st.sidebar.title("Navigation")
    
#     # Add logout button to sidebar
#     if st.sidebar.button("Logout"):
#         st.session_state.authenticated = False
#         st.rerun()
    
#     with st.sidebar.expander("About this Application"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         This professional tool calculates GC content and provides comprehensive analysis for DNA sequences. 
#         Upload your data or enter sequences manually to explore nucleotide composition, GC content, and more. 
        
#         **Developed by Shubh Rakesh Nahar, Troy University.**
#         </div>
#         """, unsafe_allow_html=True)
    
#     # Fun facts about DNA/genes/sequences
#     facts = [
#         "The human genome contains about 3 billion base pairs.",
#         "GC content can affect the stability of DNA.",
#         "Some bacteria have extremely high or low GC content.",
#         "DNA was first isolated by Friedrich Miescher in 1869.",
#         "GC-rich regions are often found near gene promoters.",
#         "Genes are segments of DNA that code for proteins.",
#         "The longest human gene is over 2.4 million base pairs long!",
#         "Mitochondrial DNA is inherited only from your mother.",
#         "Some viruses use RNA instead of DNA as their genetic material.",
#         "The fruit fly has about 15,000 genes, while humans have about 20,000-25,000.",
#         "DNA stands for Deoxyribonucleic Acid.",
#         "The double helix structure of DNA was discovered in 1953.",
#         "Some plants have much more DNA than humans!",
#         "The GC content of a genome can be used to identify species.",
#         "DNA can be extracted from almost any living thing, even ancient fossils!"
#     ]
#     if 'fun_fact_idx' not in st.session_state:
#         st.session_state['fun_fact_idx'] = random.randint(0, len(facts)-1)
#     if st.sidebar.button("Show Another Fun Fact"):
#         prev_idx = st.session_state['fun_fact_idx']
#         new_idx = prev_idx
#         while new_idx == prev_idx:
#             new_idx = random.randint(0, len(facts)-1)
#         st.session_state['fun_fact_idx'] = new_idx
#         st.rerun()
#     st.sidebar.success(f"**Fun Fact:** {facts[st.session_state['fun_fact_idx']]}")

#     # Glossary of Bioinformatics Terms
#     glossary = {
#         "GC Content": "The percentage of bases in a DNA or RNA molecule that are either guanine (G) or cytosine (C).",
#         "FASTA": "A text-based format for representing nucleotide or peptide sequences.",
#         "GenBank": "A rich file format for DNA sequences with annotations, used by NCBI.",
#         "ORF": "Open Reading Frame, a sequence of DNA that could potentially encode a protein.",
#         "Codon": "A sequence of three nucleotides that together form a unit of genetic code.",
#         "Motif": "A short, recurring pattern in DNA that is presumed to have a biological function.",
#         "CpG Island": "A region with a high frequency of CG dinucleotides, often found near gene promoters.",
#         "SNP": "Single Nucleotide Polymorphism, a variation at a single position in a DNA sequence among individuals.",
#         "BLAST": "A tool for comparing an input sequence against a database of sequences.",
#         "Reverse Complement": "The sequence formed by reversing a DNA sequence and replacing each base with its complement.",
#         "Translation": "The process of converting a nucleotide sequence into a protein sequence.",
#         "Phylogenetic Tree": "A branching diagram showing evolutionary relationships among sequences.",
#         "Alignment": "The arrangement of two or more sequences to identify regions of similarity."
#     }
#     with st.sidebar.expander("Glossary of Bioinformatics Terms"):
#         search_term = st.text_input("Search glossary", key="glossary_search")
#         for term, definition in glossary.items():
#             if not search_term or search_term.lower() in term.lower():
#                 st.markdown(f"**{term}:** {definition}")

#     # Tutorial Mode
#     with st.sidebar.expander("Tutorial / Step-by-Step Guide"):
#         st.markdown("""
#         <div style='background: rgba(255,255,255,0.05); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3);'>
#         <h4 style='margin-top:0;'>How to Use the GC Content Calculator and Sequence Fragmenter</h4>
#         <p>
#         This application allows you to analyze DNA sequences for GC content and nucleotide composition, as well as to split sequences into defined-length fragments with full analysis. Follow these steps to use the tools effectively:
#         </p>
#         <p>
#         <b>1. Choose Your Input Method:</b><br>
#         You can upload your DNA sequences using a CSV or Excel file, a FASTA file, a GenBank file, or by entering them manually. For file uploads, ensure your data includes both a gene name and a sequence for each entry.
#         </p>
#         <p>
#         <b>2. Analyze Sequences:</b><br>
#         After uploading or entering your sequences, click the "Calculate GC Content" button. The application will process your data and display a summary table, visualizations, and nucleotide composition statistics.
#         </p>
#         <p>
#         <b>3. Export Results:</b><br>
#         Once the analysis is complete, you can download your results in Excel, FASTA, GenBank, or as a ZIP archive. Use the export section to select your preferred format.
#         </p>
#         <p>
#         <b>4. Use the Sequence Fragmenter Tool:</b><br>
#         Open the "Sequence Fragmenter (Split Tool)" section in the sidebar. Choose your input method and upload or enter your sequences. Select the desired fragment length. When you click "Split and Export Fragments," the tool will split each sequence into fragments of the specified length and calculate GC content and nucleotide statistics for each fragment. You can then download the results in Excel, FASTA, or GenBank format, each including the full analysis for every fragment.
#         </p>
#         <p>
#         <b>5. Additional Features:</b><br>
#         The sidebar also provides a random sequence generator for practice, a glossary of bioinformatics terms, and fun facts about DNA and genomics.
#         </p>
#         <p>
#         If you need to log out, use the "Logout" button in the sidebar. For any step, ensure your input data is correctly formatted to avoid errors during analysis.
#         </p>
#         </div>
#         """, unsafe_allow_html=True)

#     # Random Sequence Generator
#     with st.sidebar.expander("Random Sequence Generator"):
#         st.markdown("Generate random DNA or protein sequences for practice or testing.")
#         seq_type = st.selectbox("Sequence type", ["DNA", "Protein"], key="rand_seq_type")
#         rand_num = st.number_input("Number of sequences", min_value=1, max_value=1000, value=1, key="rand_num")
#         rand_len = st.number_input("Length of each sequence", min_value=5, max_value=40000, value=50, key="rand_len")
#         if st.button("Generate Random Sequences"):
#             import string
#             import random as pyrandom
#             if seq_type == "DNA":
#                 alphabet = "ATGC"
#             else:
#                 alphabet = "ACDEFGHIKLMNPQRSTVWY"  # 20 amino acids
            
#             # Generate sequences in chunks to manage memory
#             chunk_size = 100  # Process 100 sequences at a time
#             all_sequences = []
            
#             for i in range(0, rand_num, chunk_size):
#                 current_chunk = min(chunk_size, rand_num - i)
#                 chunk_sequences = [(f"Random_{i+j+1}", ''.join(pyrandom.choices(alphabet, k=rand_len))) 
#                                  for j in range(current_chunk)]
#                 all_sequences.extend(chunk_sequences)
                
#                 # Show progress
#                 progress = min(100, int((i + current_chunk) / rand_num * 100))
#                 st.progress(progress)
            
#             if 'random_sequences' not in st.session_state:
#                 st.session_state['random_sequences'] = []
#             st.session_state['random_sequences'] = all_sequences
#             st.toast(f"{rand_num} random {seq_type} sequences of length {rand_len} generated!", icon=None)
        
#         # Option to clear generated sequences
#         if st.button("Clear Random Sequences"):
#             st.session_state['random_sequences'] = []
#             st.toast("Random sequences cleared.", icon=None)
        
#         # Download random sequences if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             # First analyze the random sequences in chunks
#             chunk_size = 100  # Process 100 sequences at a time
#             all_results = []
            
#             for i in range(0, len(st.session_state['random_sequences']), chunk_size):
#                 chunk = st.session_state['random_sequences'][i:i + chunk_size]
#                 chunk_results = [analyze_sequence(name, seq) for name, seq in chunk]
#                 all_results.extend(chunk_results)
                
#                 # Show progress
#                 progress = min(100, int((i + len(chunk)) / len(st.session_state['random_sequences']) * 100))
#                 st.progress(progress)
            
#             rand_df = pd.DataFrame(all_results)
            
#             # Excel with formatting
#             excel_buffer = BytesIO()
#             with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
#                 rand_df.to_excel(writer, index=False, sheet_name='GC Content Analysis')
#                 # Get the workbook and the worksheet
#                 workbook = writer.book
#                 worksheet = writer.sheets['GC Content Analysis']
                
#                 # Auto-adjust columns' width
#                 for column in worksheet.columns:
#                     max_length = 0
#                     column = [cell for cell in column]
#                     for cell in column:
#                         try:
#                             if len(str(cell.value)) > max_length:
#                                 max_length = len(str(cell.value))
#                         except:
#                             pass
#                     adjusted_width = (max_length + 2)
#                     worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                
#                 # Add some basic formatting
#                 for row in worksheet.iter_rows(min_row=1, max_row=1):
#                     for cell in row:
#                         cell.font = cell.font.copy(bold=True)
            
#             excel_buffer.seek(0)
#             st.download_button(
#                 label="Download Random Sequences (Excel)",
#                 data=excel_buffer,
#                 file_name="random_sequences.xlsx",
#                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#             )
            
#             # FASTA
#             fasta_buffer = export_to_fasta(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (FASTA)",
#                 data=fasta_buffer,
#                 file_name="random_sequences.fasta",
#                 mime="text/plain"
#             )
            
#             # GenBank
#             gb_buffer = export_to_genbank(st.session_state['random_sequences'], rand_df)
#             st.download_button(
#                 label="Download Random Sequences (GenBank)",
#                 data=gb_buffer,
#                 file_name="random_sequences.gb",
#                 mime="text/plain"
#             )

#     # Add random sequences to main input if present
#     if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#         sequences.extend(st.session_state['random_sequences'])
#         st.session_state['random_sequences'] = [] # Clear after adding

#     # Main content area with professional styling
#     st.markdown("---")
    
#     # Main Analysis Section
#     with st.container():
#         st.markdown("""
#             <div class="section-container">
#                 <h2 class="section-header">Sequence Analysis & GC Content Calculation</h2>
#             </div>
#         """, unsafe_allow_html=True)
        
#         st.info("**Professional Tip:** You can input up to 1000 sequences at a time (CSV, Excel, FASTA, GenBank, or manual entry). If you upload more, only the first 1000 will be processed.")

#         input_method = st.radio("Choose input method", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"])

#         if input_method == "Upload CSV/Excel":
#             uploaded_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"])
#             if uploaded_file:
#                 try:
#                     uploaded_file.seek(0) # Ensure pointer is at the start
#                     if uploaded_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(uploaded_file)
#                     else:
#                         df = pd.read_csv(uploaded_file)
                    
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         # Convert sequences to strings and handle any NaN values
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         if len(df) > 1000:
#                             st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                         sequences = list(zip(df["Gene Name"], df["Sequence"]))[:1000]
#                     else:
#                         st.toast("File must contain 'Gene Name' and 'Sequence' columns.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error reading file: {str(e)}", icon=None)

#         elif input_method == "Upload FASTA":
#             fasta_file = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"])
#             if fasta_file:
#                 fasta_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_fasta(fasta_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded FASTA file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing FASTA file: {str(e)}", icon=None)

#         elif input_method == "Upload GenBank":
#             gb_file = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"])
#             if gb_file:
#                 gb_file.seek(0)  # Reset pointer
#                 try:
#                     sequences = process_genbank(gb_file, max_sequences=1000)
#                     if len(sequences) == 0:
#                         st.warning("No sequences found in the uploaded GenBank file.")
#                     elif len(sequences) == 1000:
#                         st.toast("More than 1000 sequences detected. Only the first 1000 will be processed.", icon=None)
#                 except Exception as e:
#                     st.toast(f"Error processing GenBank file: {str(e)}", icon=None)

#         elif input_method == "Manual Entry":
#             num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1)
#             manual_entries = []
#             for i in range(num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"seq_{i}")
#                 if name and seq:
#                     manual_entries.append((name, seq))
#             sequences.extend(manual_entries)

#         # Add random sequences to main input if present
#         if 'random_sequences' in st.session_state and st.session_state['random_sequences']:
#             sequences.extend(st.session_state['random_sequences'])
#             st.session_state['random_sequences'] = [] # Clear after adding

#         if sequences:
#             if st.button("Calculate GC Content"):
#                 results = [analyze_sequence(name, seq) for name, seq in sequences]
#                 result_df = pd.DataFrame(results)

#                 if "Error" in result_df.columns:
#                     st.toast("Some sequences were invalid and skipped.", icon=None)
#                     result_df = result_df.dropna(subset=["Length"])

#                 # Check if there are valid sequences for output
#                 if len(result_df) == 0:
#                      st.warning("No valid sequences were processed.")
#                 else:

#                     # Conditional Visualization Display
#                     if len(result_df) > 50:
#                         st.info("Analysis complete. Visualizations and on-page data table are disabled for more than 50 sequences to ensure performance. Please download your full results below.")
#                     else:
#                         st.toast("Analysis complete!", icon=None)
#                         st.markdown("---")
                        
#                         # Results Section
#                         st.markdown("""
#                             <div class="section-container">
#                                 <h2 class="section-header">Results & Visualizations</h2>
#                             </div>
#                         """, unsafe_allow_html=True)
                        
#                         with st.expander("Show Data Table", expanded=True):
#                             st.dataframe(result_df, use_container_width=True)
#                         with st.expander("Show GC Content Heatmap"):
#                             if len(result_df) > 0:
#                                 create_gc_heatmap(result_df)
#                             else:
#                                 st.info("No data to plot.")
#                         with st.expander("Show Sequence Logo"):
#                             # Use the original 'sequences' list for the logo, not result_df
#                             if len(sequences) > 1 and all(len(seq) > 0 for _, seq in sequences):
#                                 # Filter original sequences based on names in result_df to match analysis results
#                                 valid_gene_names = result_df['Gene Name'].tolist()
#                                 sequences_for_logo = [(name, seq) for name, seq in sequences if name in valid_gene_names]
#                                 if len(sequences_for_logo) > 1:
#                                     create_sequence_logo([seq for _, seq in sequences_for_logo])
#                                 else:
#                                     st.info("Not enough valid sequences to plot a sequence logo.")
#                             else:
#                                 st.info("Not enough data to plot a sequence logo.")

#                         with st.expander("Show Interactive Plots"):
#                             if len(result_df) > 1:
#                                 create_interactive_plots(result_df)
#                             else:
#                                 st.info("Not enough data to plot interactive plots.")

#                     # Export Results section - always shown if there are results
#                     st.markdown("---")
#                     st.markdown("""
#                         <div class="section-container">
#                             <h2 class="section-header">Export Results</h2>
#                         </div>
#                     """, unsafe_allow_html=True)
                    
#                     output_filename = st.text_input("Enter output file name (without extension):", "gc_output_v4", key="export_filename")
                    
#                     # Export buttons in columns for better layout
#                     col1, col2, col3, col4 = st.columns(4)
                    
#                     with col1:
#                         if st.button("Export All Results"):
#                             zip_buffer = export_all_to_zip(result_df, output_filename)
#                             st.toast("All results exported as ZIP!", icon=None)
#                             st.download_button(
#                                 label="Download All Results (ZIP)",
#                                 data=zip_buffer,
#                                 file_name=f"{output_filename}_all_results.zip",
#                                 mime="application/zip"
#                             )
                    
#                     with col2:
#                         excel_buffer = BytesIO()
#                         result_df.to_excel(excel_buffer, index=False, engine='openpyxl')
#                         excel_buffer.seek(0)
#                         st.download_button(
#                             label="Download Excel Only",
#                             data=excel_buffer,
#                             file_name=f"{output_filename}.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                    
#                     with col3:
#                         # FASTA export button
#                         # Use result_df for FASTA/GenBank exports to include only analyzed sequences
#                         fasta_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         fasta_buffer = export_to_fasta(fasta_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download as FASTA",
#                             data=fasta_buffer,
#                             file_name=f"{output_filename}.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col4:
#                         # GenBank export button
#                         gb_sequences_to_export = [(row["Gene Name"], row["Sequence"]) for _, row in result_df.iterrows()]
#                         gb_buffer = export_to_genbank(gb_sequences_to_export, result_df)
#                         st.download_button(
#                             label="Download as GenBank",
#                             data=gb_buffer,
#                             file_name=f"{output_filename}.gb",
#                             mime="text/plain"
#                         )

#     # Add Sequence Fragmenter tool to sidebar
#     with st.sidebar.expander("Sequence Fragmenter (Split Tool)"):
#         st.markdown("""
#         <div style='background: rgba(52,152,219,0.1); padding: 1rem; border-radius: 10px; border: 1px solid rgba(52,152,219,0.3); margin-bottom: 1rem;'>
#         <strong>Split sequences into fragments of a chosen length.</strong><br>
#         Output is compatible with TOOL-1 and available in multiple formats.
#         </div>
#         """, unsafe_allow_html=True)
        
#         frag_input_method = st.radio("Input method for fragmenter", ["Upload CSV/Excel", "Upload FASTA", "Upload GenBank", "Manual Entry"], key="frag_input_method")
#         frag_sequences = []
        
#         if frag_input_method == "Upload CSV/Excel":
#             frag_file = st.file_uploader("Upload a CSV or Excel file with 'Gene Name' and 'Sequence' columns", type=["csv", "xlsx", "xls"], key="frag_file")
#             if frag_file:
#                 try:
#                     frag_file.seek(0)
#                     if frag_file.name.endswith((".xlsx", ".xls")):
#                         df = pd.read_excel(frag_file)
#                     else:
#                         df = pd.read_csv(frag_file)
#                     if "Gene Name" in df.columns and "Sequence" in df.columns:
#                         df['Sequence'] = df['Sequence'].fillna('').astype(str)
#                         df['Gene Name'] = df['Gene Name'].fillna('').astype(str)
#                         frag_sequences = list(zip(df["Gene Name"], df["Sequence"]))
#                         st.success(f"Successfully loaded {len(frag_sequences)} sequences from {frag_file.name}")
#                     else:
#                         st.warning("File must contain 'Gene Name' and 'Sequence' columns.")
#                 except Exception as e:
#                     st.warning(f"Error reading file: {str(e)}")
                    
#         elif frag_input_method == "Upload FASTA":
#             frag_fasta = st.file_uploader("Upload a FASTA file", type=["fasta", "fa"], key="frag_fasta")
#             if frag_fasta:
#                 frag_fasta.seek(0)
#                 frag_sequences = process_fasta(frag_fasta, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from FASTA file")
                    
#         elif frag_input_method == "Upload GenBank":
#             frag_gb = st.file_uploader("Upload a GenBank file", type=["gb", "gbk"], key="frag_gb")
#             if frag_gb:
#                 frag_gb.seek(0)
#                 frag_sequences = process_genbank(frag_gb, max_sequences=1000)
#                 if frag_sequences:
#                     st.success(f"Successfully loaded {len(frag_sequences)} sequences from GenBank file")
                    
#         elif frag_input_method == "Manual Entry":
#             frag_num = st.number_input("How many sequences would you like to enter?", min_value=1, max_value=1000, value=1, key="frag_num")
#             frag_manual_entries = []
#             for i in range(frag_num):
#                 st.markdown(f"**Sequence {i+1}**")
#                 name = st.text_input(f"Gene Name {i+1}", key=f"frag_name_{i}")
#                 seq = st.text_area(f"Sequence {i+1}", key=f"frag_seq_{i}")
#                 if name and seq:
#                     frag_manual_entries.append((name, seq))
#             frag_sequences.extend(frag_manual_entries)
#             if frag_manual_entries:
#                 st.success(f"Successfully entered {len(frag_manual_entries)} sequences")
                
#         # Fragment length selection with better styling
#         st.markdown("---")
#         st.markdown("**Fragment Configuration**")
#         frag_length = st.selectbox("Fragment length (nucleotides)", [25, 50, 100, 200, 400], key="frag_length")
        
#         if frag_sequences:
#             st.info(f"Ready to process {len(frag_sequences)} sequences into {frag_length}-nucleotide fragments")
            
#             if st.button("Split and Export Fragments", key="frag_button"):
#                 with st.spinner("Processing fragments and calculating GC content..."):
#                     fragments = fragment_sequences(frag_sequences, frag_length)
                    
#                 if not fragments:
#                     st.warning("No fragments produced. Check your input sequences and fragment length.")
#                 else:
#                     st.success(f"Successfully produced {len(fragments)} fragments of {frag_length} nucleotides each!")
                    
#                     # Calculate overall statistics
#                     total_gc_content = sum(f[5].get("GC %", 0) for f in fragments)
#                     avg_gc_content = total_gc_content / len(fragments)
                    
#                     # Display fragment statistics with GC content
#                     st.markdown("**Fragment Statistics:**")
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         st.markdown(f"""
#                         - Total fragments: {len(fragments)}
#                         - Fragment length: {frag_length} nucleotides
#                         - Parent sequences: {len(set([f[2] for f in fragments]))}
#                         - Average GC content: {avg_gc_content:.2f}%
#                         """)
                    
#                     with col2:
#                         # Show GC content distribution
#                         gc_values = [f[5].get("GC %", 0) for f in fragments]
#                         min_gc = min(gc_values)
#                         max_gc = max(gc_values)
#                         st.markdown(f"""
#                         - GC content range: {min_gc:.2f}% - {max_gc:.2f}%
#                         - High GC fragments (>60%): {sum(1 for gc in gc_values if gc > 60)}
#                         - Low GC fragments (<40%): {sum(1 for gc in gc_values if gc < 40)}
#                         """)
                    
#                     # Export options in a more organized way
#                     st.markdown("**Download Options**")
                    
#                     col1, col2 = st.columns(2)
                    
#                     with col1:
#                         excel_buffer = export_fragments_to_excel(fragments)
#                         st.download_button(
#                             label="Excel Format (with Analysis)",
#                             data=excel_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.xlsx",
#                             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#                         )
                        
#                         fasta_buffer = export_fragments_to_fasta(fragments)
#                         st.download_button(
#                             label="FASTA Format (with Analysis)",
#                             data=fasta_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.fasta",
#                             mime="text/plain"
#                         )
                    
#                     with col2:
#                         gb_buffer = export_fragments_to_genbank(fragments)
#                         st.download_button(
#                             label="GenBank Format (with Analysis)",
#                             data=gb_buffer,
#                             file_name=f"fragments_{frag_length}n_with_analysis.gb",
#                             mime="text/plain"
#                         )
                        
#                         # Show a sample of the analysis data
#                         if len(fragments) > 0:
#                             st.markdown("**Sample Fragment Analysis:**")
#                             sample_frag = fragments[0]
#                             st.markdown(f"""
#                             **{sample_frag[0]}** (from {sample_frag[2]})
#                             - GC Content: {sample_frag[5].get('GC %', 0):.2f}%
#                             - A: {sample_frag[5].get('A Count', 0)} ({sample_frag[5].get('A %', 0):.2f}%)
#                             - T: {sample_frag[5].get('T Count', 0)} ({sample_frag[5].get('T %', 0):.2f}%)
#                             - G: {sample_frag[5].get('G Count', 0)} ({sample_frag[5].get('G %', 0):.2f}%)
#                             - C: {sample_frag[5].get('C Count', 0)} ({sample_frag[5].get('C %', 0):.2f}%)
#                             """)
#         else:
#             st.info("Please upload or enter sequences to begin fragmentation")

#     # Professional footer
#     st.markdown("---")
#     st.markdown("""
#         <div class="footer">
#             <p>Developed by Shubh Rakesh Nahar | Troy University | Professional DNA Sequence Analysis Tool</p>
#         </div>
#     """, unsafe_allow_html=True)

# if __name__ == "__main__":
#     main()
